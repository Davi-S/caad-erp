"""Utility for initializing the CAAD ERP master workbook.

The module doubles as a script (``python setup_excel.py``) and as a library
used by tests or other tooling. Shared helpers keep the workbook bootstrap
logic consistent regardless of the execution path.
"""

from __future__ import annotations

import argparse
import configparser
from dataclasses import dataclass
from pathlib import Path
from typing import Mapping, MutableMapping, Sequence
import sys

import openpyxl
from openpyxl.chart import BarChart, DoughnutChart, PieChart, Reference
from openpyxl.styles import Font

# Define the schema exactly as specified in the architecture
SHEET_COLUMNS: Mapping[str, Sequence[str]] = {
    "Products": [
        "ProductID",
        "ProductName",
        "SellPrice",
        "IsActive",
    ],
    "Salesmen": [
        "SalesmanID",
        "SalesmanName",
        "IsActive",
    ],
    "TransactionLog": [
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ],
}

# Define the default salesman
DEFAULT_SALESMAN: MutableMapping[str, object] = {
    "SalesmanID": "GRR00000000",
    "SalesmanName": "Lounge Sale",
    "IsActive": True,
}

CONFIG_FILE = "config.ini"


@dataclass(frozen=True)
class SetupSettings:
    """Type-safe representation of configuration values used during setup."""

    data_file: Path
    default_salesman_id: str


def load_settings(config_path: Path) -> SetupSettings:
    """Read ``config.ini`` and produce :class:`SetupSettings`.

    Relative paths inside the config file are resolved against the config file's
    directory so the behavior matches the previous script.
    """

    if not config_path.exists():
        raise FileNotFoundError(f"Configuration file not found: {config_path}")

    parser = configparser.ConfigParser()
    parser.read(config_path)

    try:
        data_file_raw = parser.get("System", "DataFile")
        default_salesman_id = parser.get("Defaults", "DefaultSalesman")
    except (configparser.NoSectionError, configparser.NoOptionError) as exc:
        raise KeyError(f"Missing required configuration entry: {exc}") from exc

    data_file_path = Path(data_file_raw)
    if not data_file_path.is_absolute():
        data_file_path = (config_path.parent / data_file_path).resolve()

    return SetupSettings(
        data_file=data_file_path,
        default_salesman_id=default_salesman_id,
    )


def _create_dashboard_sheet(workbook: openpyxl.Workbook) -> None:
    """Create and format the executive Dashboard sheet as the first tab in the workbook."""

    ws = workbook.create_sheet(title="Dashboard", index=0)

    bold_title_font = Font(bold=True, size=14)
    bold_header_font = Font(bold=True, size=11)
    bold_font = Font(bold=True)

    # ---------------------------------------------------------
    # 1. Executive KPI Summary Cards (A1:B6)
    # ---------------------------------------------------------
    ws["A1"] = "CAAD ERP Executive Summary"
    ws["A1"].font = bold_title_font

    kpis = [
        ("Total Revenue", "=SUM(TransactionLog!H:H)/100", "$#,##0.00"),
        ("Total Costs / Expenses", "=SUM(TransactionLog!I:I)/100", "$#,##0.00"),
        ("Net Profit", "=B2+B3", "$#,##0.00"),
        ("Profit Margin", "=IFERROR((B2+B3)/B2, 0)", "0.0%"),
        (
            "Outstanding Debts",
            '=(SUMIF(TransactionLog!F:F, "OnCredit", TransactionLog!H:H) - SUMIF(TransactionLog!C:C, "CREDIT_PAYMENT", TransactionLog!H:H))/100',
            "$#,##0.00",
        ),
    ]

    for idx, (label, formula, num_format) in enumerate(kpis, start=2):
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = bold_font
        val_cell = ws.cell(row=idx, column=2, value=formula)
        val_cell.number_format = num_format

    # ---------------------------------------------------------
    # 2. Payment Method Breakdown & Pie Chart (D1:F5)
    # ---------------------------------------------------------
    ws["D1"] = "Payment Type"
    ws["D1"].font = bold_header_font
    ws["E1"] = "Revenue"
    ws["E1"].font = bold_header_font
    ws["F1"] = "Share %"
    ws["F1"].font = bold_header_font

    payment_types = ["Cash", "PIX", "OnCredit", "Other"]
    for idx, ptype in enumerate(payment_types, start=2):
        ws.cell(row=idx, column=4, value=ptype)
        rev_cell = ws.cell(
            row=idx,
            column=5,
            value=f'=SUMIFS(TransactionLog!H:H, TransactionLog!F:F, "{ptype}", TransactionLog!C:C, "SALE")/100',
        )
        rev_cell.number_format = "$#,##0.00"
        share_cell = ws.cell(
            row=idx,
            column=6,
            value=f"=IFERROR(E{idx}/$B$2, 0)",
        )
        share_cell.number_format = "0.0%"

    chart_payment = PieChart()
    chart_payment.title = "Revenue Distribution by Payment Method"
    labels_payment = Reference(ws, min_col=4, min_row=2, max_row=5)
    data_payment = Reference(ws, min_col=5, min_row=1, max_row=5)
    chart_payment.add_data(data_payment, titles_from_data=True)
    chart_payment.set_categories(labels_payment)
    chart_payment.width = 14
    chart_payment.height = 7
    ws.add_chart(chart_payment, "H1")

    # ---------------------------------------------------------
    # 3. Sales Leaderboard & Bar Chart (A16:F26)
    # ---------------------------------------------------------
    ws["A16"] = "Salesman Name"
    ws["A16"].font = bold_header_font
    ws["B16"] = "Deals Closed"
    ws["B16"].font = bold_header_font
    ws["C16"] = "Deals Rank"
    ws["C16"].font = bold_header_font
    ws["D16"] = "Total Revenue"
    ws["D16"].font = bold_header_font
    ws["E16"] = "Revenue Rank"
    ws["E16"].font = bold_header_font
    ws["F16"] = "% of Total"
    ws["F16"].font = bold_header_font

    # Populate dynamic slots linking up to 10 salesmen rows
    for r in range(2, 12):
        k = 15 + r  # Rows 17 to 26
        ws.cell(row=k, column=1, value=f'=IF(Salesmen!B{r}="","",Salesmen!B{r})')
        ws.cell(
            row=k,
            column=2,
            value=f'=IF(A{k}="","",COUNTIFS(TransactionLog!E:E, Salesmen!A{r}, TransactionLog!C:C, "SALE"))',
        )
        ws.cell(
            row=k,
            column=3,
            value=f'=IF(A{k}="","",RANK(B{k}, $B$17:$B$26))',
        )
        rev_cell = ws.cell(
            row=k,
            column=4,
            value=f'=IF(A{k}="","",SUMIFS(TransactionLog!H:H, TransactionLog!E:E, Salesmen!A{r}, TransactionLog!C:C, "SALE")/100)',
        )
        rev_cell.number_format = "$#,##0.00"
        ws.cell(
            row=k,
            column=5,
            value=f'=IF(A{k}="","",RANK(D{k}, $D$17:$D$26))',
        )
        share_cell = ws.cell(
            row=k, column=6, value=f'=IF(A{k}="","",IFERROR(D{k}/$B$2, 0))'
        )
        share_cell.number_format = "0.0%"

    # ---------------------------------------------------------
    # 4. Dynamic Product Performance & Inventory Table (A28:E78)
    # ---------------------------------------------------------
    ws["A28"] = "ProductID"
    ws["A28"].font = bold_header_font
    ws["B28"] = "ProductName"
    ws["B28"].font = bold_header_font
    ws["C28"] = "Stock On Hand"
    ws["C28"].font = bold_header_font
    ws["D28"] = "Total Sales Revenue"
    ws["D28"].font = bold_header_font
    ws["E28"] = "Stock Status Alert"
    ws["E28"].font = bold_header_font

    # Populate dynamic formula templates up to 50 products
    for r in range(2, 52):
        i = 27 + r  # Dashboard row 29 to 78
        ws.cell(row=i, column=1, value=f'=IF(Products!A{r}="","",Products!A{r})')
        ws.cell(row=i, column=2, value=f'=IF(A{i}="","",Products!B{r})')
        ws.cell(
            row=i,
            column=3,
            value=f'=IF(A{i}="","",SUMIF(TransactionLog!D:D, A{i}, TransactionLog!G:G))',
        )
        rev_cell = ws.cell(
            row=i,
            column=4,
            value=f'=IF(A{i}="","",SUMIFS(TransactionLog!H:H, TransactionLog!D:D, A{i}, TransactionLog!C:C, "SALE")/100)',
        )
        rev_cell.number_format = "$#,##0.00"
        ws.cell(
            row=i,
            column=5,
            value=f'=IF(A{i}="","",IF(C{i}<=0, "OUT OF STOCK", IF(C{i}<=5, "LOW STOCK", "OK")))',
        )

    column_widths = {
        "A": 22,
        "B": 24,
        "C": 14,
        "D": 22,
        "E": 18,
        "F": 14,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    workbook.active = ws


def create_master_workbook(
    destination: Path,
    *,
    default_salesman_id: str | None = None,
    sheet_columns: Mapping[str, Sequence[str]] = SHEET_COLUMNS,
    default_salesman_template: Mapping[str, object] | None = None,
    overwrite: bool = False,
) -> Path:
    """Create the CAAD ERP master workbook at ``destination``.

    Parameters are overridable to facilitate testing. When ``overwrite`` is
    ``False`` (the default) this function raises ``FileExistsError`` if the
    target already exists.
    """

    destination = destination.expanduser().resolve()
    if destination.exists() and not overwrite:
        raise FileExistsError(
            f"Refusing to overwrite existing master workbook: {destination}"
        )

    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()

    # Remove the default sheet openpyxl generates so we can create ours.
    if workbook.active and workbook.active.title == "Sheet":
        workbook.remove(workbook.active)

    bold_font = Font(bold=True)

    for sheet_name, columns in sheet_columns.items():
        worksheet = workbook.create_sheet(title=sheet_name)
        for column_index, column_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=column_index)
            cell.value = column_name
            cell.font = bold_font

    _create_dashboard_sheet(workbook)

    workbook.save(destination)
    return destination


def run_from_config(config_path: Path, *, overwrite: bool = False) -> Path:
    """Convenience helper mirroring the original CLI behavior."""

    settings = load_settings(config_path)
    return create_master_workbook(
        settings.data_file,
        default_salesman_id=settings.default_salesman_id,
        overwrite=overwrite,
    )


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    """Parse CLI arguments for the setup script."""

    parser = argparse.ArgumentParser(description="Initialize CAAD ERP data file")
    parser.add_argument(
        "--config",
        default=CONFIG_FILE,
        help="Path to configuration file (default: config.ini)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite the target workbook if it already exists.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    """Entry point for the CLI script."""

    args = parse_args(argv)
    config_path = Path(args.config).expanduser().resolve()

    print("--- CAAD ERP Setup Script ---")
    print(f"Using configuration: {config_path}")

    try:
        output_path = run_from_config(config_path, overwrite=args.force)
    except FileNotFoundError as exc:
        print(f"\n[ERROR] {exc}")
        return 1
    except KeyError as exc:
        print(f"\n[ERROR] {exc}")
        return 1
    except FileExistsError as exc:
        print(f"\n[ERROR] {exc}")
        print("Run with --force to overwrite the existing file if appropriate.")
        return 1
    except (PermissionError, OSError) as exc:
        print(f"\n[ERROR] Unable to write workbook: {exc}")
        return 1

    print(f"\n[SUCCESS] Created master workbook at '{output_path}'.")
    return 0


if __name__ == "__main__":  # pragma: no cover - exercised via manual runs
    sys.exit(main())
