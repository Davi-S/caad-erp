from pathlib import Path
import openpyxl
import pytest

from setup_excel import create_master_workbook, run_from_config


def test_create_master_workbook_creates_dashboard_sheet(tmp_path: Path):
    target_path = tmp_path / "master_workbook.xlsx"
    create_master_workbook(target_path)

    assert target_path.exists()
    wb = openpyxl.load_workbook(target_path)

    # 1. Verify sheet ordering and active sheet
    assert wb.sheetnames[0] == "Dashboard"
    assert "Products" in wb.sheetnames
    assert "Salesmen" in wb.sheetnames
    assert "TransactionLog" in wb.sheetnames
    assert wb.active.title == "Dashboard"

    dashboard = wb["Dashboard"]

    # 2. Verify charts presence
    assert len(dashboard._charts) == 1

    # 3. Verify KPI card cell contents starting at A1 (cents divided by 100)
    assert dashboard["A1"].value == "CAAD ERP Executive Summary"
    assert dashboard["A2"].value == "Total Revenue"
    assert dashboard["B2"].value == "=SUM(TransactionLog!H:H)/100"
    assert dashboard["B3"].value == "=SUM(TransactionLog!I:I)/100"
    assert dashboard["B4"].value == "=B2+B3"
    assert dashboard["B5"].value == "=IFERROR((B2+B3)/B2, 0)"

    # 4. Verify Payment Breakdown table (D1:F5)
    assert dashboard["D2"].value == "Cash"
    assert dashboard["E2"].value == '=SUMIFS(TransactionLog!H:H, TransactionLog!F:F, "Cash", TransactionLog!C:C, "SALE")/100'
    assert dashboard["F2"].value == "=IFERROR(E2/$B$2, 0)"

    # 5. Verify Sales Leaderboard with dual rankings
    assert dashboard["A17"].value == '=IF(Salesmen!B2="","",Salesmen!B2)'
    assert dashboard["C17"].value == '=IF(A17="","",RANK(B17, $B$17:$B$26))'
    assert dashboard["D17"].value == '=IF(A17="","",SUMIFS(TransactionLog!H:H, TransactionLog!E:E, Salesmen!A2, TransactionLog!C:C, "SALE")/100)'
    assert dashboard["E17"].value == '=IF(A17="","",RANK(D17, $D$17:$D$26))'

    # 6. Verify Dynamic Product Table (A28+)
    assert dashboard["A29"].value == '=IF(Products!A2="","",Products!A2)'
    assert dashboard["C29"].value == '=IF(A29="","",SUMIF(TransactionLog!D:D, A29, TransactionLog!G:G))'
    assert dashboard["D29"].value == '=IF(A29="","",SUMIFS(TransactionLog!H:H, TransactionLog!D:D, A29, TransactionLog!C:C, "SALE")/100)'


def test_run_from_config_generates_dashboard(tmp_path: Path):
    config_path = tmp_path / "config.ini"
    data_file = tmp_path / "data" / "master_workbook.xlsx"
    config_path.write_text(
        f"[System]\nDataFile = {data_file}\n\n[Defaults]\nDefaultSalesman = GRR00000000\n"
    )

    output = run_from_config(config_path, overwrite=True)
    assert output == data_file.resolve()

    wb = openpyxl.load_workbook(output)
    assert wb.sheetnames[0] == "Dashboard"
