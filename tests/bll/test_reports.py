from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook import Workbook

from caad_erp import constants, dal
from caad_erp.bll import reports, runtime
from caad_erp.settings import AppSettings


def _make_workbook() -> Workbook:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    products_sheet = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products_sheet.append(
        ["ProductID", "ProductName", "SellPrice", "IsActive"])

    salesmen_sheet = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen_sheet.append(["SalesmanID", "SalesmanName", "IsActive"])

    tx_sheet = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx_sheet.append(
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ]
    )
    return wb


def _make_context(workbook: Workbook) -> runtime.RuntimeContext:
    settings = AppSettings(
        data_file=Path("/tmp/data.xlsx"),
        lounge_name="Test Lounge",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return runtime.RuntimeContext(settings=settings, workbook=workbook)


def _seed_product(workbook: Workbook, product_id: str, sell_price: int) -> None:
    dal.append_product(
        workbook,
        dal.ProductRow(
            product_id=product_id,
            product_name=f"Product {product_id}",
            sell_price=sell_price,
            is_active=True,
        ),
    )


def _seed_transaction(
    workbook: Workbook,
    transaction_id: str,
    transaction_type: str,
    product_id: str | None,
    payment_type: str | None,
    quantity_change: int,
    total_revenue: int,
    total_cost: int,
    linked_transaction_id: str | None = None,
) -> None:
    dal.append_transaction(
        workbook,
        dal.TransactionRow(
            transaction_id=transaction_id,
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=transaction_type,
            product_id=product_id,
            salesman_id="S001",
            payment_type=payment_type,
            quantity_change=quantity_change,
            total_revenue=total_revenue,
            total_cost=total_cost,
            linked_transaction_id=linked_transaction_id,
            notes=None,
        ),
    )


def test_calculate_inventory_sums_quantity_by_product() -> None:
    """
    GIVEN transaction cache entries with product ids and quantity deltas
    WHEN calculate_inventory is called
    THEN inventory totals are accumulated per product id
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_transaction(
        workbook,
        "T1",
        constants.TransactionType.RESTOCK.value,
        "P1",
        None,
        5,
        0,
        -1000,
    )
    _seed_transaction(
        workbook,
        "T2",
        constants.TransactionType.SALE.value,
        "P1",
        constants.PaymentType.CASH.value,
        -2,
        600,
        0,
    )
    _seed_transaction(
        workbook,
        "T3",
        constants.TransactionType.RESTOCK.value,
        "P2",
        None,
        3,
        0,
        -400,
    )

    # Act
    inventory = reports.calculate_inventory(context)

    # Assert
    assert inventory == {"P1": 3, "P2": 3}


def test_calculate_inventory_ignores_entries_without_product_id() -> None:
    """
    GIVEN transaction cache entries where some rows have product_id set to None
    WHEN calculate_inventory is called
    THEN rows without product_id are ignored
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_transaction(
        workbook,
        "T1",
        constants.TransactionType.CREDIT_PAYMENT.value,
        None,
        constants.PaymentType.PIX.value,
        0,
        1000,
        0,
        linked_transaction_id="SALE1",
    )

    # Act
    inventory = reports.calculate_inventory(context)

    # Assert
    assert inventory == {}


def test_calculate_inventory_returns_empty_mapping_for_no_transactions() -> None:
    """
    GIVEN an empty transaction cache list
    WHEN calculate_inventory is called
    THEN an empty inventory mapping is returned
    """
    # Arrange
    context = _make_context(_make_workbook())

    # Act
    inventory = reports.calculate_inventory(context)

    # Assert
    assert inventory == {}


def test_calculate_profit_summary_returns_revenue_cost_and_profit() -> None:
    """
    GIVEN transaction cache entries with total_revenue and total_cost values
    WHEN calculate_profit_summary is called
    THEN total_revenue total_cost and profit keys are returned
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_transaction(
        workbook,
        "T1",
        constants.TransactionType.SALE.value,
        "P1",
        constants.PaymentType.CASH.value,
        -1,
        1000,
        0,
    )
    _seed_transaction(
        workbook,
        "T2",
        constants.TransactionType.RESTOCK.value,
        "P1",
        None,
        2,
        0,
        -300,
    )

    # Act
    summary = reports.calculate_profit_summary(context)

    # Assert
    assert set(summary.keys()) == {"total_revenue", "total_cost", "profit"}
    assert summary["total_revenue"] == 1000
    assert summary["total_cost"] == -300
    assert summary["profit"] == 700


def test_calculate_profit_summary_uses_additive_profit_formula() -> None:
    """
    GIVEN transaction cache totals where costs are stored as negative amounts
    WHEN calculate_profit_summary is called
    THEN profit is computed as total_revenue plus total_cost
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_transaction(
        workbook,
        "T1",
        constants.TransactionType.SALE.value,
        "P1",
        constants.PaymentType.CASH.value,
        -1,
        5000,
        0,
    )
    _seed_transaction(
        workbook,
        "T2",
        constants.TransactionType.RESTOCK.value,
        "P1",
        None,
        1,
        0,
        -2000,
    )

    # Act
    summary = reports.calculate_profit_summary(context)

    # Assert
    assert summary["profit"] == summary["total_revenue"] + \
        summary["total_cost"]
    assert summary["profit"] == 3000


def test_calculate_profit_summary_returns_zeros_for_empty_transactions() -> None:
    """
    GIVEN an empty transaction cache list
    WHEN calculate_profit_summary is called
    THEN total_revenue total_cost and profit are all zero
    """
    # Arrange
    context = _make_context(_make_workbook())

    # Act
    summary = reports.calculate_profit_summary(context)

    # Assert
    assert summary == {
        "total_revenue": 0,
        "total_cost": 0,
        "profit": 0,
    }


def test_calculate_outstanding_debts_returns_balances_and_total() -> None:
    """
    GIVEN cached transactions containing unresolved credit sales and payments
    WHEN calculate_outstanding_debts is called
    THEN balances list and total_outstanding value are returned
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", 800)
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "P1",
        constants.PaymentType.ON_CREDIT.value,
        -2,
        0,
        0,
    )
    _seed_transaction(
        workbook,
        "PAY1",
        constants.TransactionType.CREDIT_PAYMENT.value,
        "P1",
        constants.PaymentType.CASH.value,
        0,
        500,
        0,
        linked_transaction_id="SALE1",
    )

    # Act
    result = reports.calculate_outstanding_debts(context)

    # Assert
    assert len(result["balances"]) == 1
    debt = result["balances"][0]
    assert debt.transaction_id == "SALE1"
    assert debt.expected_amount == 1600
    assert debt.amount_paid == 500
    assert debt.balance == 1100
    assert result["total_outstanding"] == 1100


def test_calculate_outstanding_debts_returns_empty_result_for_no_transactions() -> None:
    """
    GIVEN an empty transaction cache list
    WHEN calculate_outstanding_debts is called
    THEN balances is empty and total_outstanding is zero
    """
    # Arrange
    context = _make_context(_make_workbook())

    # Act
    result = reports.calculate_outstanding_debts(context)

    # Assert
    assert result["balances"] == []
    assert result["total_outstanding"] == 0


def test_calculate_outstanding_debts_ignores_non_credit_sales() -> None:
    """
    GIVEN cached sale transactions not marked as OnCredit
    WHEN calculate_outstanding_debts is called
    THEN non-credit sales are excluded from debt balances
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", 1000)
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "P1",
        constants.PaymentType.CASH.value,
        -2,
        2000,
        0,
    )

    # Act
    result = reports.calculate_outstanding_debts(context)

    # Assert
    assert result["balances"] == []
    assert result["total_outstanding"] == 0


def test_calculate_outstanding_debts_ignores_voided_sales() -> None:
    """
    GIVEN cached credit sales that have corresponding VOID transactions
    WHEN calculate_outstanding_debts is called
    THEN voided sales are excluded from debt balances
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", 1000)
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "P1",
        constants.PaymentType.ON_CREDIT.value,
        -1,
        0,
        0,
    )
    _seed_transaction(
        workbook,
        "VOID1",
        constants.TransactionType.VOID.value,
        "P1",
        constants.PaymentType.ON_CREDIT.value,
        1,
        0,
        0,
        linked_transaction_id="SALE1",
    )

    # Act
    result = reports.calculate_outstanding_debts(context)

    # Assert
    assert result["balances"] == []


def test_calculate_outstanding_debts_ignores_non_sale_transactions() -> None:
    """
    GIVEN cached transactions where transaction_type is not SALE
    WHEN calculate_outstanding_debts is called
    THEN non-sale entries are ignored for debt generation
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_transaction(
        workbook,
        "TX1",
        constants.TransactionType.RESTOCK.value,
        "P1",
        None,
        2,
        0,
        -500,
    )

    # Act
    result = reports.calculate_outstanding_debts(context)

    # Assert
    assert result["balances"] == []


@pytest.mark.parametrize("expected_source", ["transaction", "price"])
def test_calculate_outstanding_debts_prefers_expected_amount_source(expected_source) -> None:
    """
    GIVEN credit sales with possible expected amount from transaction and product price
    WHEN calculate_outstanding_debts is called
    THEN expected amount selection follows the implemented precedence rules
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", 900)

    if expected_source == "transaction":
        # Expected amount should come from negative total_revenue magnitude (12)
        _seed_transaction(
            workbook,
            "SALE1",
            constants.TransactionType.SALE.value,
            "P1",
            constants.PaymentType.ON_CREDIT.value,
            -2,
            -1200,
            0,
        )
        expected_amount = 1200
    else:
        # Expected amount should come from price * quantity (9 * 2 = 18)
        _seed_transaction(
            workbook,
            "SALE1",
            constants.TransactionType.SALE.value,
            "P1",
            constants.PaymentType.ON_CREDIT.value,
            -2,
            0,
            0,
        )
        expected_amount = 1800

    # Act
    result = reports.calculate_outstanding_debts(context)

    # Assert
    assert len(result["balances"]) == 1
    assert result["balances"][0].expected_amount == expected_amount


def test_calculate_outstanding_debts_skips_fully_settled_or_nonpositive_balances() -> None:
    """
    GIVEN credit sales whose payments satisfy or exceed expected amount
    WHEN calculate_outstanding_debts is called
    THEN entries with nonpositive balances are not included
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", 1000)
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "P1",
        constants.PaymentType.ON_CREDIT.value,
        -1,
        0,
        0,
    )
    _seed_transaction(
        workbook,
        "PAY1",
        constants.TransactionType.CREDIT_PAYMENT.value,
        "P1",
        constants.PaymentType.CASH.value,
        0,
        1000,
        0,
        linked_transaction_id="SALE1",
    )

    # Act
    result = reports.calculate_outstanding_debts(context)

    # Assert
    assert result["balances"] == []
    assert result["total_outstanding"] == 0


def test_calculate_outstanding_debts_skips_entries_with_nonpositive_expected_amount() -> None:
    """
    GIVEN credit sales whose expected amount resolves to zero or below
    WHEN calculate_outstanding_debts is called
    THEN those sales are excluded from balances
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", 0)
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "P1",
        constants.PaymentType.ON_CREDIT.value,
        -2,
        0,
        0,
    )

    # Act
    result = reports.calculate_outstanding_debts(context)

    # Assert
    assert result["balances"] == []


def test_calculate_outstanding_debts_handles_missing_product_reference_gracefully() -> None:
    """
    GIVEN a credit sale whose product lookup raises MissingReferenceError
    WHEN calculate_outstanding_debts is called
    THEN calculation continues without crashing and applies remaining rules
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "UNKNOWN_PRODUCT",
        constants.PaymentType.ON_CREDIT.value,
        -1,
        0,
        0,
    )

    # Act
    result = reports.calculate_outstanding_debts(context)

    # Assert
    assert result["balances"] == []
    assert result["total_outstanding"] == 0
