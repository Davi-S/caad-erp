from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook import Workbook

from caad_erp import constants, dal
from caad_erp.bll import runtime
from caad_erp.settings import AppSettings


def _make_context(data_file: Path, schema_version: str = constants.EXPECTED_SCHEMA_VERSION) -> runtime.RuntimeContext:
    settings = AppSettings(
        data_file=data_file,
        lounge_name="Test Lounge",
        schema_version=schema_version,
        default_salesman_id="S001",
    )
    return runtime.RuntimeContext(settings=settings, workbook=Workbook())


def _write_config(config_path: Path, data_file: Path, schema_version: str = constants.EXPECTED_SCHEMA_VERSION) -> None:
    config_path.write_text(
        "\n".join(
            [
                "[System]",
                f"DataFile = {data_file}",
                "LoungeName = Test Lounge",
                f"SchemaVersion = {schema_version}",
                "",
                "[Defaults]",
                "DefaultSalesman = S001",
                "",
            ]
        ),
        encoding="utf-8",
    )


def test_persist_context_saves_to_settings_data_file(tmp_path: Path) -> None:
    """
    GIVEN a runtime context with workbook and configured data file path
    WHEN persist_context is called
    THEN dal.save_workbook is invoked with the workbook and configured destination
    """
    # Arrange
    destination = tmp_path / "saved.xlsx"
    context = _make_context(destination)
    context.workbook.active.title = "Main"

    # Act
    runtime.persist_context(context)

    # Assert
    assert destination.exists()
    opened = dal.open_workbook(destination)
    assert "Main" in opened.sheetnames


def test_persist_context_propagates_save_errors(tmp_path: Path) -> None:
    """
    GIVEN a runtime context where DAL save raises a filesystem error
    WHEN persist_context is called
    THEN the original save error is propagated
    """
    # Arrange
    invalid_destination = tmp_path / "as_dir"
    invalid_destination.mkdir()
    context = _make_context(invalid_destination)

    # Act / Assert
    with pytest.raises((IsADirectoryError, PermissionError, OSError)):
        runtime.persist_context(context)


def test_get_cache_bucket_returns_existing_bucket(tmp_path: Path) -> None:
    """
    GIVEN a context cache containing the requested bucket name
    WHEN get_cache_bucket is called
    THEN the existing mutable bucket is returned
    """
    # Arrange
    context = _make_context(tmp_path / "dummy.xlsx")
    existing = {"key": "value"}
    context._cache["products"] = existing

    # Act
    result = runtime.get_cache_bucket(context, "products")

    # Assert
    assert result is existing


def test_get_cache_bucket_creates_missing_bucket(tmp_path: Path) -> None:
    """
    GIVEN a context cache without the requested bucket name
    WHEN get_cache_bucket is called
    THEN a new empty bucket is created stored and returned
    """
    # Arrange
    context = _make_context(tmp_path / "dummy.xlsx")

    # Act
    result = runtime.get_cache_bucket(context, "products")

    # Assert
    assert result == {}
    assert context._cache["products"] is result


def test_invalidate_cache_no_names_is_no_op(tmp_path: Path) -> None:
    """
    GIVEN a context cache with populated buckets
    WHEN invalidate_cache is called without names
    THEN no buckets are removed
    """
    # Arrange
    context = _make_context(tmp_path / "dummy.xlsx")
    context._cache["products"] = {"all": []}
    context._cache["salesmen"] = {"all": []}

    # Act
    runtime.invalidate_cache(context)

    # Assert
    assert "products" in context._cache
    assert "salesmen" in context._cache


@pytest.mark.parametrize(
    "bucket_names",
    [
        ["products"],
        ["salesmen"],
        ["products", "salesmen"],
        ["products", "transactions"],
    ],
)
def test_invalidate_cache_removes_requested_buckets(tmp_path: Path, bucket_names) -> None:
    """
    GIVEN a context cache with multiple buckets
    WHEN invalidate_cache is called with specific bucket names
    THEN only those buckets are removed
    """
    # Arrange
    context = _make_context(tmp_path / "dummy.xlsx")
    context._cache["products"] = {"all": []}
    context._cache["salesmen"] = {"all": []}
    context._cache["transactions"] = {"all": []}

    # Act
    runtime.invalidate_cache(context, *bucket_names)

    # Assert
    for name in bucket_names:
        assert name not in context._cache


def test_invalidate_cache_ignores_unknown_bucket_names(tmp_path: Path) -> None:
    """
    GIVEN a context cache without a requested bucket name
    WHEN invalidate_cache is called
    THEN unknown names are ignored without raising
    """
    # Arrange
    context = _make_context(tmp_path / "dummy.xlsx")

    # Act / Assert
    runtime.invalidate_cache(context, "missing")


def test_load_context_returns_runtime_context(tmp_path: Path) -> None:
    """
    GIVEN valid settings and a loadable workbook
    WHEN load_context is called
    THEN it returns a RuntimeContext with settings workbook and empty cache
    """
    # Arrange
    workbook_path = tmp_path / "data.xlsx"
    openpyxl.Workbook().save(workbook_path)
    config_path = tmp_path / "config.ini"
    _write_config(config_path, workbook_path)

    # Act
    result = runtime.load_context(config_path=config_path)

    # Assert
    assert isinstance(result, runtime.RuntimeContext)
    assert result.settings.data_file == workbook_path.resolve()
    assert result._cache == {}
    assert result.workbook is not None


def test_load_context_uses_explicit_config_path_when_provided(tmp_path: Path) -> None:
    """
    GIVEN an explicit configuration path argument
    WHEN load_context is called
    THEN settings are loaded from that path before opening the workbook
    """
    # Arrange
    workbook_path = tmp_path / "data.xlsx"
    openpyxl.Workbook().save(workbook_path)
    config_path = tmp_path / "custom_config.ini"
    _write_config(config_path, workbook_path)

    # Act
    result = runtime.load_context(config_path=config_path)

    # Assert
    assert result.settings.data_file == workbook_path.resolve()


def test_load_context_uses_discovery_when_config_path_is_none(tmp_path: Path, monkeypatch) -> None:
    """
    GIVEN no explicit config path argument
    WHEN load_context is called
    THEN settings loader discovery flow is used
    """
    # Arrange
    workbook_path = tmp_path / "data.xlsx"
    openpyxl.Workbook().save(workbook_path)
    config_path = tmp_path / "config.ini"
    _write_config(config_path, workbook_path)
    monkeypatch.chdir(tmp_path)

    # Act
    result = runtime.load_context(config_path=None)

    # Assert
    assert result.settings.data_file == workbook_path.resolve()


@pytest.mark.parametrize("loader_exception", ["missing_config", "missing_workbook", "bad_config"])
def test_load_context_propagates_settings_and_workbook_errors(
    tmp_path: Path,
    loader_exception: str,
) -> None:
    """
    GIVEN settings or workbook loader raises a domain-relevant exception
    WHEN load_context is called
    THEN the original exception is propagated to the caller
    """
    # Arrange / Act / Assert
    if loader_exception == "missing_config":
        with pytest.raises(FileNotFoundError):
            runtime.load_context(config_path=tmp_path / "missing.ini")
    elif loader_exception == "missing_workbook":
        config_path = tmp_path / "config.ini"
        _write_config(config_path, tmp_path / "missing.xlsx")
        with pytest.raises(FileNotFoundError):
            runtime.load_context(config_path=config_path)
    else:
        config_path = tmp_path / "config.ini"
        config_path.write_text(
            "[System]\nDataFile = data.xlsx\n", encoding="utf-8")
        with pytest.raises(KeyError):
            runtime.load_context(config_path=config_path)


def test_ensure_schema_version_accepts_expected_version(tmp_path: Path) -> None:
    """
    GIVEN a runtime context whose schema version matches EXPECTED_SCHEMA_VERSION
    WHEN ensure_schema_version is called
    THEN no exception is raised
    """
    # Arrange
    context = _make_context(tmp_path / "dummy.xlsx",
                            schema_version=constants.EXPECTED_SCHEMA_VERSION)

    # Act / Assert
    runtime.ensure_schema_version(context)


def test_ensure_schema_version_raises_on_mismatch(tmp_path: Path) -> None:
    """
    GIVEN a runtime context whose schema version differs from EXPECTED_SCHEMA_VERSION
    WHEN ensure_schema_version is called
    THEN RuntimeError is raised with mismatch details
    """
    # Arrange
    context = _make_context(tmp_path / "dummy.xlsx", schema_version="0.9.9")

    # Act / Assert
    with pytest.raises(RuntimeError, match="schema mismatch"):
        runtime.ensure_schema_version(context)
