import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook import Workbook

from caad_erp import constants, dal
from caad_erp.bll import runtime, transactions
from caad_erp.exceptions import BusinessRuleViolation, MissingReferenceError
from caad_erp.settings import AppSettings


def _make_workbook() -> Workbook:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    products_sheet = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products_sheet.append(
        ["ProductID", "ProductName", "SellPrice", "IsActive"])

    salesmen_sheet = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen_sheet.append(["SalesmanID", "SalesmanName", "IsActive"])

    tx_sheet = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx_sheet.append(
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ]
    )
    return wb


def _make_context(workbook: Workbook) -> runtime.RuntimeContext:
    settings = AppSettings(
        data_file=Path("/tmp/data.xlsx"),
        lounge_name="Test Lounge",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return runtime.RuntimeContext(settings=settings, workbook=workbook)


def _seed_product(workbook: Workbook, product_id: str, is_active: bool = True) -> None:
    dal.append_product(
        workbook,
        dal.ProductRow(
            product_id=product_id,
            product_name=f"Product {product_id}",
            sell_price=Decimal("7.00"),
            is_active=is_active,
        ),
    )


def _seed_salesman(workbook: Workbook, salesman_id: str, is_active: bool = True) -> None:
    dal.append_salesman(
        workbook,
        dal.SalesmanRow(
            salesman_id=salesman_id,
            salesman_name=f"Salesman {salesman_id}",
            is_active=is_active,
        ),
    )


def _seed_transaction(
    workbook: Workbook,
    transaction_id: str,
    transaction_type: str,
    product_id: str | None,
    salesman_id: str | None,
    payment_type: str | None,
    quantity_change: Decimal,
    total_revenue: Decimal,
    total_cost: Decimal,
    linked_transaction_id: str | None = None,
) -> dal.TransactionRow:
    row = dal.TransactionRow(
        transaction_id=transaction_id,
        timestamp_iso="2026-03-15T10:00:00+00:00",
        transaction_type=transaction_type,
        product_id=product_id,
        salesman_id=salesman_id,
        payment_type=payment_type,
        quantity_change=quantity_change,
        total_revenue=total_revenue,
        total_cost=total_cost,
        linked_transaction_id=linked_transaction_id,
        notes=None,
    )
    dal.append_transaction(workbook, row)
    return row


def test_ensure_transactions_cache_populates_missing_cache() -> None:
    """
    GIVEN a runtime context with an empty transactions cache bucket
    WHEN _ensure_transactions_cache is called
    THEN all and by_id structures are populated from DAL iteration
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_transaction(
        workbook,
        "T1",
        constants.TransactionType.SALE.value,
        "P1",
        "S1",
        constants.PaymentType.CASH.value,
        Decimal("-1"),
        Decimal("5"),
        Decimal("0"),
    )

    # Act
    bucket = transactions._ensure_transactions_cache(context)

    # Assert
    assert len(bucket["all"]) == 1
    assert bucket["by_id"]["T1"].transaction_type == constants.TransactionType.SALE.value


def test_ensure_transactions_cache_reuses_existing_cache() -> None:
    """
    GIVEN a runtime context with a pre-populated transactions cache bucket
    WHEN _ensure_transactions_cache is called
    THEN existing cache structures are reused without re-iterating DAL
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_transaction(
        workbook,
        "T1",
        constants.TransactionType.SALE.value,
        "P1",
        "S1",
        constants.PaymentType.CASH.value,
        Decimal("-1"),
        Decimal("5"),
        Decimal("0"),
    )
    first = transactions._ensure_transactions_cache(context)
    _seed_transaction(
        workbook,
        "T2",
        constants.TransactionType.RESTOCK.value,
        "P1",
        "S1",
        None,
        Decimal("1"),
        Decimal("0"),
        Decimal("-3"),
    )

    # Act
    second = transactions._ensure_transactions_cache(context)

    # Assert
    assert second is first
    assert len(second["all"]) == 1
    assert "T2" not in second["by_id"]


def test_generate_transaction_id_uses_expected_timestamp_format() -> None:
    """
    GIVEN a datetime input
    WHEN _generate_transaction_id is called
    THEN identifier format is YYYYMMDDHHMMSSffffff
    """
    # Arrange
    when = datetime.datetime(2026, 3, 15, 12, 34, 56,
                             123456, tzinfo=datetime.UTC)

    # Act
    transaction_id = transactions._generate_transaction_id(when)

    # Assert
    assert transaction_id == "20260315123456123456"
    assert len(transaction_id) == 20


def test_generate_transaction_id_is_lexicographically_sortable() -> None:
    """
    GIVEN two datetimes in chronological order
    WHEN _generate_transaction_id is called for both
    THEN earlier datetime yields lexicographically smaller id
    """
    # Arrange
    first_dt = datetime.datetime(2026, 3, 15, 10, 0, 0, 1, tzinfo=datetime.UTC)
    second_dt = datetime.datetime(
        2026, 3, 15, 10, 0, 0, 2, tzinfo=datetime.UTC)

    # Act
    first_id = transactions._generate_transaction_id(first_dt)
    second_id = transactions._generate_transaction_id(second_dt)

    # Assert
    assert first_id < second_id


@pytest.mark.parametrize("quantity", [Decimal("0.0001"), Decimal("1"), Decimal("99")])
def test_require_positive_quantity_accepts_positive_values(quantity: Decimal) -> None:
    """
    GIVEN a strictly positive quantity
    WHEN _require_positive_quantity is called
    THEN no exception is raised
    """
    # Arrange / Act / Assert
    transactions._require_positive_quantity(quantity)


@pytest.mark.parametrize("quantity", [Decimal("0"), Decimal("-0.1"), Decimal("-5")])
def test_require_positive_quantity_rejects_zero_or_negative(quantity: Decimal) -> None:
    """
    GIVEN a zero or negative quantity
    WHEN _require_positive_quantity is called
    THEN ValueError is raised
    """
    # Arrange / Act / Assert
    with pytest.raises(ValueError):
        transactions._require_positive_quantity(quantity)


@pytest.mark.parametrize("amount", [Decimal("0"), Decimal("0.01"), Decimal("5")])
def test_require_nonnegative_money_accepts_nonnegative_values(amount: Decimal) -> None:
    """
    GIVEN a nonnegative monetary amount
    WHEN _require_nonnegative_money is called
    THEN no exception is raised
    """
    # Arrange / Act / Assert
    transactions._require_nonnegative_money(amount)


@pytest.mark.parametrize("amount", [Decimal("-0.01"), Decimal("-5")])
def test_require_nonnegative_money_rejects_negative_values(amount: Decimal) -> None:
    """
    GIVEN a negative monetary amount
    WHEN _require_nonnegative_money is called
    THEN ValueError is raised
    """
    # Arrange / Act / Assert
    with pytest.raises(ValueError):
        transactions._require_nonnegative_money(amount)


def test_list_transactions_returns_copy_of_cached_log() -> None:
    """
    GIVEN a transactions cache bucket with all entries
    WHEN list_transactions is called
    THEN a list copy of cached transactions is returned
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_transaction(
        workbook,
        "T1",
        constants.TransactionType.SALE.value,
        "P1",
        "S1",
        constants.PaymentType.CASH.value,
        Decimal("-1"),
        Decimal("5"),
        Decimal("0"),
    )

    # Act
    result = transactions.list_transactions(context)

    # Assert
    assert len(result) == 1
    assert result is not context._cache["transactions"]["all"]


def test_get_transaction_returns_matching_row() -> None:
    """
    GIVEN a transaction id present in cached by_id map
    WHEN get_transaction is called
    THEN the matching TransactionRow is returned
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    expected = _seed_transaction(
        workbook,
        "T1",
        constants.TransactionType.SALE.value,
        "P1",
        "S1",
        constants.PaymentType.CASH.value,
        Decimal("-1"),
        Decimal("5"),
        Decimal("0"),
    )

    # Act
    result = transactions.get_transaction(context, "T1")

    # Assert
    assert result == expected


def test_get_transaction_raises_missing_reference_for_unknown_id() -> None:
    """
    GIVEN a transaction id missing from cached by_id map
    WHEN get_transaction is called
    THEN MissingReferenceError is raised
    """
    # Arrange
    context = _make_context(_make_workbook())

    # Act / Assert
    with pytest.raises(MissingReferenceError):
        transactions.get_transaction(context, "UNKNOWN")


def test_record_sale_appends_transaction_and_invalidates_cache() -> None:
    """
    GIVEN an active product active salesman and valid sale command
    WHEN record_sale is called
    THEN SALE transaction is appended and transactions cache is invalidated
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", is_active=True)
    _seed_salesman(workbook, "S1", is_active=True)
    transactions.list_transactions(context)  # warm cache

    # Act
    created = transactions.record_sale(
        context,
        transactions.SaleCommand(
            product_id="P1",
            salesman_id="S1",
            quantity=Decimal("2"),
            total_revenue=Decimal("14"),
            payment_type=constants.PaymentType.CASH,
        ),
    )

    # Assert
    assert created.transaction_type == constants.TransactionType.SALE.value
    assert created.quantity_change == Decimal("-2")
    assert created.total_revenue == Decimal("14")
    assert "transactions" not in context._cache


@pytest.mark.parametrize("inactive_party", ["product", "salesman"])
def test_record_sale_rejects_inactive_product_or_salesman(inactive_party) -> None:
    """
    GIVEN a sale command referencing an inactive product or salesman
    WHEN record_sale is called
    THEN BusinessRuleViolation is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", is_active=(inactive_party != "product"))
    _seed_salesman(workbook, "S1", is_active=(inactive_party != "salesman"))

    # Act / Assert
    with pytest.raises(BusinessRuleViolation):
        transactions.record_sale(
            context,
            transactions.SaleCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=Decimal("1"),
                total_revenue=Decimal("7"),
                payment_type=constants.PaymentType.CASH,
            ),
        )


@pytest.mark.parametrize("missing_reference", ["product", "salesman"])
def test_record_sale_propagates_product_or_salesman_missing_reference(
    missing_reference,
) -> None:
    """
    GIVEN a sale command referencing unknown product or salesman ids
    WHEN record_sale is called
    THEN MissingReferenceError is propagated
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    if missing_reference != "product":
        _seed_product(workbook, "P1")
    if missing_reference != "salesman":
        _seed_salesman(workbook, "S1")

    # Act / Assert
    with pytest.raises(MissingReferenceError):
        transactions.record_sale(
            context,
            transactions.SaleCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=Decimal("1"),
                total_revenue=Decimal("7"),
                payment_type=constants.PaymentType.CASH,
            ),
        )


@pytest.mark.parametrize("invalid_quantity", [Decimal("0"), Decimal("-1")])
def test_record_sale_rejects_nonpositive_quantity(invalid_quantity: Decimal) -> None:
    """
    GIVEN a sale command with zero or negative quantity
    WHEN record_sale is called
    THEN ValueError is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1")
    _seed_salesman(workbook, "S1")

    # Act / Assert
    with pytest.raises(ValueError):
        transactions.record_sale(
            context,
            transactions.SaleCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=invalid_quantity,
                total_revenue=Decimal("7"),
                payment_type=constants.PaymentType.CASH,
            ),
        )


@pytest.mark.parametrize("invalid_revenue", [Decimal("-0.01"), Decimal("-10")])
def test_record_sale_rejects_negative_revenue(invalid_revenue: Decimal) -> None:
    """
    GIVEN a sale command with negative total_revenue
    WHEN record_sale is called
    THEN ValueError is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1")
    _seed_salesman(workbook, "S1")

    # Act / Assert
    with pytest.raises(ValueError):
        transactions.record_sale(
            context,
            transactions.SaleCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=Decimal("1"),
                total_revenue=invalid_revenue,
                payment_type=constants.PaymentType.CASH,
            ),
        )


def test_build_sale_transaction_applies_expected_field_mapping() -> None:
    """
    GIVEN a SaleCommand transaction id and timestamp
    WHEN _build_sale_transaction is called
    THEN SALE row fields are mapped with negative quantity and zero cost
    """
    # Arrange
    command = transactions.SaleCommand(
        product_id="P1",
        salesman_id="S1",
        quantity=Decimal("2"),
        total_revenue=Decimal("10"),
        payment_type=constants.PaymentType.CASH,
        notes="sale note",
    )
    ts = datetime.datetime(2026, 3, 15, 10, 0, tzinfo=datetime.UTC)

    # Act
    row = transactions._build_sale_transaction(
        command, transaction_id="TX1", timestamp=ts)

    # Assert
    assert row.transaction_id == "TX1"
    assert row.timestamp_iso == ts.isoformat()
    assert row.transaction_type == constants.TransactionType.SALE.value
    assert row.quantity_change == Decimal("-2")
    assert row.total_revenue == Decimal("10")
    assert row.total_cost == Decimal("0.00")
    assert row.linked_transaction_id is None
    assert row.notes == "sale note"


def test_build_sale_transaction_serializes_payment_type_value() -> None:
    """
    GIVEN a SaleCommand with PaymentType enum value
    WHEN _build_sale_transaction is called
    THEN payment_type column stores the enum string value
    """
    # Arrange
    command = transactions.SaleCommand(
        product_id="P1",
        salesman_id="S1",
        quantity=Decimal("1"),
        total_revenue=Decimal("5"),
        payment_type=constants.PaymentType.PIX,
    )

    # Act
    row = transactions._build_sale_transaction(
        command,
        transaction_id="TX1",
        timestamp=datetime.datetime(2026, 3, 15, tzinfo=datetime.UTC),
    )

    # Assert
    assert row.payment_type == constants.PaymentType.PIX.value


def test_record_restock_appends_transaction_and_invalidates_cache() -> None:
    """
    GIVEN an active product active salesman and valid restock command
    WHEN record_restock is called
    THEN RESTOCK transaction is appended and transactions cache is invalidated
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1")
    _seed_salesman(workbook, "S1")
    transactions.list_transactions(context)

    # Act
    created = transactions.record_restock(
        context,
        transactions.RestockCommand(
            product_id="P1",
            salesman_id="S1",
            quantity=Decimal("3"),
            total_cost=Decimal("12"),
        ),
    )

    # Assert
    assert created.transaction_type == constants.TransactionType.RESTOCK.value
    assert created.quantity_change == Decimal("3")
    assert created.total_cost == Decimal("-12")
    assert "transactions" not in context._cache


def test_record_restock_rejects_inactive_product_or_salesman() -> None:
    """
    GIVEN a restock command referencing inactive product or salesman
    WHEN record_restock is called
    THEN BusinessRuleViolation is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", is_active=False)
    _seed_salesman(workbook, "S1")

    # Act / Assert
    with pytest.raises(BusinessRuleViolation):
        transactions.record_restock(
            context,
            transactions.RestockCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=Decimal("1"),
                total_cost=Decimal("2"),
            ),
        )


@pytest.mark.parametrize("invalid_quantity", [Decimal("0"), Decimal("-1")])
def test_record_restock_rejects_nonpositive_quantity(invalid_quantity: Decimal) -> None:
    """
    GIVEN a restock command with zero or negative quantity
    WHEN record_restock is called
    THEN ValueError is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1")
    _seed_salesman(workbook, "S1")

    # Act / Assert
    with pytest.raises(ValueError):
        transactions.record_restock(
            context,
            transactions.RestockCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=invalid_quantity,
                total_cost=Decimal("2"),
            ),
        )


def test_record_restock_accepts_negative_input_cost_by_normalizing_with_abs() -> None:
    """
    GIVEN a restock command with negative total_cost input
    WHEN record_restock is called
    THEN validation accepts input by using absolute value semantics
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1")
    _seed_salesman(workbook, "S1")

    # Act
    created = transactions.record_restock(
        context,
        transactions.RestockCommand(
            product_id="P1",
            salesman_id="S1",
            quantity=Decimal("2"),
            total_cost=Decimal("-5"),
        ),
    )

    # Assert
    assert created.total_cost == Decimal("-5")


def test_build_restock_transaction_applies_expected_field_mapping() -> None:
    """
    GIVEN a RestockCommand transaction id and timestamp
    WHEN _build_restock_transaction is called
    THEN RESTOCK row fields are mapped with positive quantity and negative cost
    """
    # Arrange
    command = transactions.RestockCommand(
        product_id="P1",
        salesman_id="S1",
        quantity=Decimal("2"),
        total_cost=Decimal("7"),
        notes="restock note",
    )

    # Act
    row = transactions._build_restock_transaction(
        command,
        transaction_id="TX1",
        timestamp=datetime.datetime(2026, 3, 15, tzinfo=datetime.UTC),
    )

    # Assert
    assert row.transaction_type == constants.TransactionType.RESTOCK.value
    assert row.quantity_change == Decimal("2")
    assert row.total_revenue == Decimal("0.00")
    assert row.total_cost == Decimal("-7")
    assert row.notes == "restock note"


def test_build_restock_transaction_enforces_negative_cost_sign() -> None:
    """
    GIVEN a RestockCommand with positive or negative total_cost magnitude
    WHEN _build_restock_transaction is called
    THEN total_cost is always stored as a negative amount
    """
    # Arrange
    positive = transactions.RestockCommand(
        product_id="P1",
        salesman_id="S1",
        quantity=Decimal("1"),
        total_cost=Decimal("4"),
    )
    negative = transactions.RestockCommand(
        product_id="P1",
        salesman_id="S1",
        quantity=Decimal("1"),
        total_cost=Decimal("-4"),
    )
    ts = datetime.datetime(2026, 3, 15, tzinfo=datetime.UTC)

    # Act
    positive_row = transactions._build_restock_transaction(
        positive, transaction_id="A", timestamp=ts)
    negative_row = transactions._build_restock_transaction(
        negative, transaction_id="B", timestamp=ts)

    # Assert
    assert positive_row.total_cost == Decimal("-4")
    assert negative_row.total_cost == Decimal("-4")


def test_record_write_off_appends_transaction_and_invalidates_cache() -> None:
    """
    GIVEN an active product active salesman and valid write-off command
    WHEN record_write_off is called
    THEN WRITE_OFF transaction is appended and transactions cache is invalidated
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1")
    _seed_salesman(workbook, "S1")
    transactions.list_transactions(context)

    # Act
    created = transactions.record_write_off(
        context,
        transactions.WriteOffCommand(
            product_id="P1",
            salesman_id="S1",
            quantity=Decimal("2"),
        ),
    )

    # Assert
    assert created.transaction_type == constants.TransactionType.WRITE_OFF.value
    assert created.quantity_change == Decimal("-2")
    assert created.total_revenue == Decimal("0.00")
    assert created.total_cost == Decimal("0.00")
    assert "transactions" not in context._cache


@pytest.mark.parametrize("inactive_party", ["product", "salesman"])
def test_record_write_off_rejects_inactive_product_or_salesman(inactive_party) -> None:
    """
    GIVEN a write-off command referencing inactive product or salesman
    WHEN record_write_off is called
    THEN BusinessRuleViolation is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", is_active=(inactive_party != "product"))
    _seed_salesman(workbook, "S1", is_active=(inactive_party != "salesman"))

    # Act / Assert
    with pytest.raises(BusinessRuleViolation):
        transactions.record_write_off(
            context,
            transactions.WriteOffCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=Decimal("1"),
            ),
        )


@pytest.mark.parametrize("invalid_quantity", [Decimal("0"), Decimal("-1")])
def test_record_write_off_rejects_nonpositive_quantity(invalid_quantity: Decimal) -> None:
    """
    GIVEN a write-off command with zero or negative quantity
    WHEN record_write_off is called
    THEN ValueError is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1")
    _seed_salesman(workbook, "S1")

    # Act / Assert
    with pytest.raises(ValueError):
        transactions.record_write_off(
            context,
            transactions.WriteOffCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=invalid_quantity,
            ),
        )


def test_build_write_off_transaction_applies_expected_field_mapping() -> None:
    """
    GIVEN a WriteOffCommand transaction id and timestamp
    WHEN _build_write_off_transaction is called
    THEN WRITE_OFF row fields are mapped with negative quantity and zero amounts
    """
    # Arrange
    command = transactions.WriteOffCommand(
        product_id="P1",
        salesman_id="S1",
        quantity=Decimal("3"),
        notes="damaged",
    )

    # Act
    row = transactions._build_write_off_transaction(
        command,
        transaction_id="TX1",
        timestamp=datetime.datetime(2026, 3, 15, tzinfo=datetime.UTC),
    )

    # Assert
    assert row.transaction_type == constants.TransactionType.WRITE_OFF.value
    assert row.quantity_change == Decimal("-3")
    assert row.total_revenue == Decimal("0.00")
    assert row.total_cost == Decimal("0.00")
    assert row.notes == "damaged"


def test_record_credit_payment_appends_transaction_and_invalidates_cache() -> None:
    """
    GIVEN a valid linked credit sale active salesman and payment command
    WHEN record_credit_payment is called
    THEN CREDIT_PAYMENT transaction is appended and transactions cache invalidated
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_salesman(workbook, "S1")
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "P1",
        "S1",
        constants.PaymentType.ON_CREDIT.value,
        Decimal("-1"),
        Decimal("0"),
        Decimal("0"),
    )
    transactions.list_transactions(context)

    # Act
    created = transactions.record_credit_payment(
        context,
        transactions.CreditPaymentCommand(
            linked_transaction_id="SALE1",
            salesman_id="S1",
            total_revenue=Decimal("4"),
            payment_type=constants.PaymentType.PIX,
        ),
    )

    # Assert
    assert created.transaction_type == constants.TransactionType.CREDIT_PAYMENT.value
    assert created.linked_transaction_id == "SALE1"
    assert created.quantity_change == Decimal("0")
    assert created.total_revenue == Decimal("4")
    assert "transactions" not in context._cache


def test_record_credit_payment_rejects_inactive_salesman() -> None:
    """
    GIVEN a credit payment command with inactive salesman
    WHEN record_credit_payment is called
    THEN BusinessRuleViolation is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_salesman(workbook, "S1", is_active=False)
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "P1",
        "S1",
        constants.PaymentType.ON_CREDIT.value,
        Decimal("-1"),
        Decimal("0"),
        Decimal("0"),
    )

    # Act / Assert
    with pytest.raises(BusinessRuleViolation):
        transactions.record_credit_payment(
            context,
            transactions.CreditPaymentCommand(
                linked_transaction_id="SALE1",
                salesman_id="S1",
                total_revenue=Decimal("4"),
                payment_type=constants.PaymentType.CASH,
            ),
        )


def test_record_credit_payment_propagates_unknown_linked_transaction() -> None:
    """
    GIVEN a credit payment command linked to unknown transaction id
    WHEN record_credit_payment is called
    THEN MissingReferenceError is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_salesman(workbook, "S1")

    # Act / Assert
    with pytest.raises(MissingReferenceError):
        transactions.record_credit_payment(
            context,
            transactions.CreditPaymentCommand(
                linked_transaction_id="UNKNOWN",
                salesman_id="S1",
                total_revenue=Decimal("4"),
                payment_type=constants.PaymentType.CASH,
            ),
        )


@pytest.mark.parametrize("invalid_revenue", [Decimal("-0.01"), Decimal("-2")])
def test_record_credit_payment_rejects_negative_revenue(invalid_revenue: Decimal) -> None:
    """
    GIVEN a credit payment command with negative total_revenue
    WHEN record_credit_payment is called
    THEN ValueError is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_salesman(workbook, "S1")
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "P1",
        "S1",
        constants.PaymentType.ON_CREDIT.value,
        Decimal("-1"),
        Decimal("0"),
        Decimal("0"),
    )

    # Act / Assert
    with pytest.raises(ValueError):
        transactions.record_credit_payment(
            context,
            transactions.CreditPaymentCommand(
                linked_transaction_id="SALE1",
                salesman_id="S1",
                total_revenue=invalid_revenue,
                payment_type=constants.PaymentType.CASH,
            ),
        )


def test_record_credit_payment_rejects_ineligible_linked_sale() -> None:
    """
    GIVEN a credit payment command linked to transaction that fails credit-link validation
    WHEN record_credit_payment is called
    THEN BusinessRuleViolation is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_salesman(workbook, "S1")
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "P1",
        "S1",
        constants.PaymentType.CASH.value,
        Decimal("-1"),
        Decimal("7"),
        Decimal("0"),
    )

    # Act / Assert
    with pytest.raises(BusinessRuleViolation):
        transactions.record_credit_payment(
            context,
            transactions.CreditPaymentCommand(
                linked_transaction_id="SALE1",
                salesman_id="S1",
                total_revenue=Decimal("4"),
                payment_type=constants.PaymentType.CASH,
            ),
        )


def test_build_credit_payment_transaction_applies_expected_field_mapping() -> None:
    """
    GIVEN a CreditPaymentCommand transaction id timestamp and linked product id
    WHEN _build_credit_payment_transaction is called
    THEN CREDIT_PAYMENT row fields are mapped with zero quantity and linked id
    """
    # Arrange
    command = transactions.CreditPaymentCommand(
        linked_transaction_id="SALE1",
        salesman_id="S1",
        total_revenue=Decimal("6"),
        payment_type=constants.PaymentType.PIX,
        notes="partial",
    )

    # Act
    row = transactions._build_credit_payment_transaction(
        command,
        transaction_id="CP1",
        timestamp=datetime.datetime(2026, 3, 15, tzinfo=datetime.UTC),
        product_id="P1",
    )

    # Assert
    assert row.transaction_type == constants.TransactionType.CREDIT_PAYMENT.value
    assert row.product_id == "P1"
    assert row.linked_transaction_id == "SALE1"
    assert row.quantity_change == Decimal("0")
    assert row.total_revenue == Decimal("6")
    assert row.total_cost == Decimal("0.00")
    assert row.payment_type == constants.PaymentType.PIX.value


@pytest.mark.parametrize(
    "link_case",
    [
        dal.TransactionRow(
            transaction_id="X1",
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=constants.TransactionType.RESTOCK.value,
            product_id="P1",
            salesman_id="S1",
            payment_type=None,
            quantity_change=Decimal("1"),
            total_revenue=Decimal("0"),
            total_cost=Decimal("-1"),
            linked_transaction_id=None,
            notes=None,
        ),
        dal.TransactionRow(
            transaction_id="X2",
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=constants.TransactionType.SALE.value,
            product_id="P1",
            salesman_id="S1",
            payment_type=constants.PaymentType.CASH.value,
            quantity_change=Decimal("-1"),
            total_revenue=Decimal("0"),
            total_cost=Decimal("0"),
            linked_transaction_id=None,
            notes=None,
        ),
        dal.TransactionRow(
            transaction_id="X3",
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=constants.TransactionType.SALE.value,
            product_id="P1",
            salesman_id="S1",
            payment_type=constants.PaymentType.ON_CREDIT.value,
            quantity_change=Decimal("-1"),
            total_revenue=Decimal("1"),
            total_cost=Decimal("0"),
            linked_transaction_id=None,
            notes=None,
        ),
        dal.TransactionRow(
            transaction_id="X4",
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=constants.TransactionType.SALE.value,
            product_id="P1",
            salesman_id="S1",
            payment_type=constants.PaymentType.ON_CREDIT.value,
            quantity_change=Decimal("-1"),
            total_revenue=Decimal("0"),
            total_cost=Decimal("0"),
            linked_transaction_id="OTHER",
            notes=None,
        ),
    ],
)
def test_validate_credit_sale_link_rejects_ineligible_transactions(link_case) -> None:
    """
    GIVEN a linked transaction that violates credit linkage constraints
    WHEN _validate_credit_sale_link is called
    THEN BusinessRuleViolation is raised
    """
    # Arrange / Act / Assert
    with pytest.raises(BusinessRuleViolation):
        transactions._validate_credit_sale_link(link_case)


def test_validate_credit_sale_link_accepts_valid_credit_sale() -> None:
    """
    GIVEN a SALE transaction recorded as OnCredit with allowed revenue and linkage state
    WHEN _validate_credit_sale_link is called
    THEN no exception is raised
    """
    # Arrange
    valid = dal.TransactionRow(
        transaction_id="SALE1",
        timestamp_iso="2026-03-15T10:00:00+00:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P1",
        salesman_id="S1",
        payment_type=constants.PaymentType.ON_CREDIT.value,
        quantity_change=Decimal("-1"),
        total_revenue=Decimal("0"),
        total_cost=Decimal("0"),
        linked_transaction_id=None,
        notes=None,
    )

    # Act / Assert
    transactions._validate_credit_sale_link(valid)


def test_record_open_stock_appends_transaction_and_invalidates_cache() -> None:
    """
    GIVEN an active product active salesman and valid open stock command
    WHEN record_open_stock is called
    THEN OPEN_STOCK transaction is appended and transactions cache invalidated
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1")
    _seed_salesman(workbook, "S1")
    transactions.list_transactions(context)

    # Act
    created = transactions.record_open_stock(
        context,
        transactions.OpenStockCommand(
            product_id="P1",
            salesman_id="S1",
            quantity=Decimal("5"),
            total_revenue=Decimal("25"),
        ),
    )

    # Assert
    assert created.transaction_type == constants.TransactionType.OPEN_STOCK.value
    assert created.quantity_change == Decimal("5")
    assert created.total_revenue == Decimal("25")
    assert created.total_cost == Decimal("0.00")
    assert "transactions" not in context._cache


@pytest.mark.parametrize("inactive_party", ["product", "salesman"])
def test_record_open_stock_rejects_inactive_product_or_salesman(inactive_party) -> None:
    """
    GIVEN an open stock command referencing inactive product or salesman
    WHEN record_open_stock is called
    THEN BusinessRuleViolation is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1", is_active=(inactive_party != "product"))
    _seed_salesman(workbook, "S1", is_active=(inactive_party != "salesman"))

    # Act / Assert
    with pytest.raises(BusinessRuleViolation):
        transactions.record_open_stock(
            context,
            transactions.OpenStockCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=Decimal("1"),
                total_revenue=Decimal("1"),
            ),
        )


@pytest.mark.parametrize("invalid_quantity", [Decimal("0"), Decimal("-1")])
def test_record_open_stock_rejects_nonpositive_quantity(invalid_quantity: Decimal) -> None:
    """
    GIVEN an open stock command with zero or negative quantity
    WHEN record_open_stock is called
    THEN ValueError is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1")
    _seed_salesman(workbook, "S1")

    # Act / Assert
    with pytest.raises(ValueError):
        transactions.record_open_stock(
            context,
            transactions.OpenStockCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=invalid_quantity,
                total_revenue=Decimal("1"),
            ),
        )


@pytest.mark.parametrize("invalid_revenue", [Decimal("-0.01"), Decimal("-3")])
def test_record_open_stock_rejects_negative_revenue(invalid_revenue: Decimal) -> None:
    """
    GIVEN an open stock command with negative total_revenue
    WHEN record_open_stock is called
    THEN ValueError is raised
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_product(workbook, "P1")
    _seed_salesman(workbook, "S1")

    # Act / Assert
    with pytest.raises(ValueError):
        transactions.record_open_stock(
            context,
            transactions.OpenStockCommand(
                product_id="P1",
                salesman_id="S1",
                quantity=Decimal("1"),
                total_revenue=invalid_revenue,
            ),
        )


def test_build_open_stock_transaction_applies_expected_field_mapping() -> None:
    """
    GIVEN an OpenStockCommand transaction id and timestamp
    WHEN _build_open_stock_transaction is called
    THEN OPEN_STOCK row fields are mapped with positive quantity and zero cost
    """
    # Arrange
    command = transactions.OpenStockCommand(
        product_id="P1",
        salesman_id="S1",
        quantity=Decimal("4"),
        total_revenue=Decimal("20"),
    )

    # Act
    row = transactions._build_open_stock_transaction(
        command,
        transaction_id="OS1",
        timestamp=datetime.datetime(2026, 3, 15, tzinfo=datetime.UTC),
    )

    # Assert
    assert row.transaction_type == constants.TransactionType.OPEN_STOCK.value
    assert row.quantity_change == Decimal("4")
    assert row.total_revenue == Decimal("20")
    assert row.total_cost == Decimal("0.00")


def test_record_void_appends_reversal_and_invalidates_cache() -> None:
    """
    GIVEN a valid target transaction and void command
    WHEN record_void is called
    THEN VOID reversal is appended and transactions cache invalidated
    """
    # Arrange
    workbook = _make_workbook()
    context = _make_context(workbook)
    _seed_transaction(
        workbook,
        "SALE1",
        constants.TransactionType.SALE.value,
        "P1",
        "S1",
        constants.PaymentType.CASH.value,
        Decimal("-2"),
        Decimal("10"),
        Decimal("0"),
    )
    transactions.list_transactions(context)

    # Act
    reversal = transactions.record_void(
        context,
        transactions.VoidCommand(
            linked_transaction_id="SALE1", notes="wrong sale"),
    )

    # Assert
    assert reversal.transaction_type == constants.TransactionType.VOID.value
    assert reversal.linked_transaction_id == "SALE1"
    assert reversal.quantity_change == Decimal("2")
    assert reversal.total_revenue == Decimal("-10")
    assert "transactions" not in context._cache


def test_record_void_propagates_unknown_target_transaction() -> None:
    """
    GIVEN a void command linked to unknown transaction id
    WHEN record_void is called
    THEN MissingReferenceError is raised
    """
    # Arrange
    context = _make_context(_make_workbook())

    # Act / Assert
    with pytest.raises(MissingReferenceError):
        transactions.record_void(
            context,
            transactions.VoidCommand(linked_transaction_id="UNKNOWN"),
        )


def test_record_void_rejects_ineligible_target_types() -> None:
    """
    GIVEN a void command targeting transaction type VOID or CREDIT_PAYMENT
    WHEN record_void is called
    THEN BusinessRuleViolation is raised
    """
    # Arrange
    for tx_type in [
        constants.TransactionType.VOID.value,
        constants.TransactionType.CREDIT_PAYMENT.value,
    ]:
        workbook = _make_workbook()
        context = _make_context(workbook)
        _seed_transaction(
            workbook,
            "TARGET",
            tx_type,
            "P1",
            "S1",
            constants.PaymentType.CASH.value,
            Decimal("0"),
            Decimal("1"),
            Decimal("0"),
        )

        # Act / Assert
        with pytest.raises(BusinessRuleViolation):
            transactions.record_void(
                context,
                transactions.VoidCommand(linked_transaction_id="TARGET"),
            )


def test_build_void_transaction_negates_numeric_fields_and_links_original_id() -> None:
    """
    GIVEN an original transaction timestamp and optional notes
    WHEN _build_void_transaction is called
    THEN numeric deltas are negated and linked_transaction_id references original id
    """
    # Arrange
    original = dal.TransactionRow(
        transaction_id="SALE1",
        timestamp_iso="2026-03-15T10:00:00+00:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P1",
        salesman_id="S1",
        payment_type=constants.PaymentType.CASH.value,
        quantity_change=Decimal("-2"),
        total_revenue=Decimal("10"),
        total_cost=Decimal("0"),
        linked_transaction_id=None,
        notes=None,
    )

    # Act
    row = transactions._build_void_transaction(
        original,
        timestamp=datetime.datetime(2026, 3, 15, tzinfo=datetime.UTC),
        notes="void note",
    )

    # Assert
    assert row.transaction_type == constants.TransactionType.VOID.value
    assert row.linked_transaction_id == "SALE1"
    assert row.quantity_change == Decimal("2")
    assert row.total_revenue == Decimal("-10")
    assert row.total_cost == Decimal("0")
    assert row.notes == "void note"


@pytest.mark.parametrize(
    "ineligible_type",
    [constants.TransactionType.VOID.value,
        constants.TransactionType.CREDIT_PAYMENT.value],
)
def test_validate_void_target_rejects_ineligible_types(ineligible_type) -> None:
    """
    GIVEN a transaction whose type is VOID or CREDIT_PAYMENT
    WHEN _validate_void_target is called
    THEN BusinessRuleViolation is raised
    """
    # Arrange
    row = dal.TransactionRow(
        transaction_id="X1",
        timestamp_iso="2026-03-15T10:00:00+00:00",
        transaction_type=ineligible_type,
        product_id="P1",
        salesman_id="S1",
        payment_type=None,
        quantity_change=Decimal("0"),
        total_revenue=Decimal("0"),
        total_cost=Decimal("0"),
        linked_transaction_id=None,
        notes=None,
    )

    # Act / Assert
    with pytest.raises(BusinessRuleViolation):
        transactions._validate_void_target(row)


def test_validate_void_target_accepts_other_transaction_types() -> None:
    """
    GIVEN a transaction type that is eligible for voiding
    WHEN _validate_void_target is called
    THEN no exception is raised
    """
    # Arrange
    row = dal.TransactionRow(
        transaction_id="X1",
        timestamp_iso="2026-03-15T10:00:00+00:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P1",
        salesman_id="S1",
        payment_type=constants.PaymentType.CASH.value,
        quantity_change=Decimal("-1"),
        total_revenue=Decimal("5"),
        total_cost=Decimal("0"),
        linked_transaction_id=None,
        notes=None,
    )

    # Act / Assert
    transactions._validate_void_target(row)
