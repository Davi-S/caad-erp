"""Unit tests outlining the expected behavior of the business logic layer."""

from __future__ import annotations

from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from caad_erp import constants, core_logic, data_manager  # noqa: E402


@pytest.fixture
def runtime_context(config_file):
    """Load the runtime context through the public API."""

    context = core_logic.load_runtime_context(config_file)
    core_logic.ensure_schema_version(context)
    return context


def _append_product(workbook_path: Path, *, product_id: str, name: str, price: Decimal, is_active: bool = True) -> None:
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook[constants.SheetName.PRODUCTS.value]
    sheet.append([product_id, name, str(price), is_active])
    workbook.save(workbook_path)


def _append_salesman(workbook_path: Path, *, salesman_id: str, name: str, is_active: bool = True) -> None:
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook[constants.SheetName.SALESMEN.value]
    sheet.append([salesman_id, name, is_active])
    workbook.save(workbook_path)


def _append_transaction(
    workbook_path: Path,
    *,
    transaction_id: str,
    timestamp: str,
    transaction_type: str,
    product_id: str | None,
    salesman_id: str | None,
    payment_type: str | None,
    quantity_change: str,
    total_revenue: str,
    total_cost: str,
    linked_transaction_id: str | None,
    notes: str | None,
) -> None:
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook[constants.SheetName.TRANSACTION_LOG.value]
    sheet.append(
        [
            transaction_id,
            timestamp,
            transaction_type,
            product_id,
            salesman_id,
            payment_type,
            quantity_change,
            total_revenue,
            total_cost,
            linked_transaction_id,
            notes,
        ]
    )
    workbook.save(workbook_path)


# ---------------------------------------------------------------------------
# Runtime/context management
# ---------------------------------------------------------------------------


def test_load_runtime_context_returns_context(config_file):
    """load_runtime_context should assemble settings and workbook into a context."""

    context = core_logic.load_runtime_context(config_file)
    assert isinstance(context, core_logic.RuntimeContext)
    assert context.settings.data_file.exists()


def test_ensure_schema_version_rejects_mismatch(config_factory):
    """Schema mismatches should surface a RuntimeError with clear messaging."""

    bundle = config_factory(schema_version="0.9")
    context = core_logic.load_runtime_context(bundle.config_path)
    with pytest.raises(RuntimeError):
        core_logic.ensure_schema_version(context)


def test_list_products_excludes_inactive_by_default(config_file):
    """list_products should hide inactive rows unless explicitly requested."""

    _append_product(config_file.parent / "master_workbook.xlsx", product_id="P1", name="Active", price=Decimal("1.00"))
    _append_product(
        config_file.parent / "master_workbook.xlsx",
        product_id="P2",
        name="Inactive",
        price=Decimal("2.00"),
        is_active=False,
    )

    context = core_logic.load_runtime_context(config_file)
    products = core_logic.list_products(context)
    ids = {product.product_id for product in products}
    assert "P1" in ids and "P2" not in ids


def test_list_products_can_include_inactive(config_file):
    """A caller should be able to include inactive products when needed."""

    _append_product(config_file.parent / "master_workbook.xlsx", product_id="P3", name="Old", price=Decimal("3.00"), is_active=False)

    context = core_logic.load_runtime_context(config_file)
    products = core_logic.list_products(context, include_inactive=True)
    assert any(product.product_id == "P3" for product in products)


def test_list_salesmen_excludes_inactive_by_default(config_file):
    """list_salesmen should filter inactive rows unless instructed otherwise."""

    _append_salesman(config_file.parent / "master_workbook.xlsx", salesman_id="S2", name="Active", is_active=True)
    _append_salesman(config_file.parent / "master_workbook.xlsx", salesman_id="S3", name="Retired", is_active=False)

    context = core_logic.load_runtime_context(config_file)
    salesmen = core_logic.list_salesmen(context)
    ids = {salesman.salesman_id for salesman in salesmen}
    assert "S2" in ids and "S3" not in ids


def test_list_transactions_returns_all_rows(config_file):
    """list_transactions should return every ledger entry in order."""

    _append_transaction(
        config_file.parent / "master_workbook.xlsx",
        transaction_id="T1",
        timestamp="2025-10-30T00:00:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P1",
        salesman_id="S-DEFAULT",
        payment_type=constants.PaymentType.CASH.value,
        quantity_change="-1",
        total_revenue="1.00",
        total_cost="0.00",
        linked_transaction_id=None,
        notes=None,
    )

    context = core_logic.load_runtime_context(config_file)
    transactions = core_logic.list_transactions(context)
    assert transactions[0].transaction_id == "T1"


def test_get_product_returns_match(config_file):
    """get_product should hydrate a ProductRow for the requested ID."""

    _append_product(config_file.parent / "master_workbook.xlsx", product_id="P10", name="Cookie", price=Decimal("4.00"))

    context = core_logic.load_runtime_context(config_file)
    product = core_logic.get_product(context, "P10")
    assert product.product_name == "Cookie"


def test_get_product_missing_raises(config_file):
    """Unknown ProductIDs should raise MissingReferenceError."""

    context = core_logic.load_runtime_context(config_file)
    with pytest.raises(core_logic.MissingReferenceError):
        core_logic.get_product(context, "NOPE")


def test_get_salesman_returns_match(config_file):
    """get_salesman should fetch active salesmen."""

    _append_salesman(config_file.parent / "master_workbook.xlsx", salesman_id="S8", name="Jordan")

    context = core_logic.load_runtime_context(config_file)
    salesman = core_logic.get_salesman(context, "S8")
    assert salesman.salesman_name == "Jordan"


def test_get_transaction_returns_match(config_file):
    """get_transaction should retrieve ledger rows by ID."""

    _append_transaction(
        config_file.parent / "master_workbook.xlsx",
        transaction_id="T55",
        timestamp="2025-10-30T01:00:00",
        transaction_type=constants.TransactionType.RESTOCK.value,
        product_id="P10",
        salesman_id=None,
        payment_type=None,
        quantity_change="10",
        total_revenue="0.00",
        total_cost="-20.00",
        linked_transaction_id=None,
        notes="Bulk",
    )

    context = core_logic.load_runtime_context(config_file)
    transaction = core_logic.get_transaction(context, "T55")
    assert transaction.transaction_type == constants.TransactionType.RESTOCK.value


def test_calculate_inventory_rolls_up_quantities(config_file):
    """calculate_inventory should return total on-hand per ProductID."""

    _append_transaction(
        config_file.parent / "master_workbook.xlsx",
        transaction_id="T100",
        timestamp="2025-10-30T02:00:00",
        transaction_type=constants.TransactionType.RESTOCK.value,
        product_id="P10",
        salesman_id=None,
        payment_type=constants.PaymentType.CASH.value,
        quantity_change="5",
        total_revenue="0.00",
        total_cost="-10.00",
        linked_transaction_id=None,
        notes=None,
    )
    _append_transaction(
        config_file.parent / "master_workbook.xlsx",
        transaction_id="T101",
        timestamp="2025-10-30T02:30:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P10",
        salesman_id="S-DEFAULT",
        payment_type=constants.PaymentType.CASH.value,
        quantity_change="-2",
        total_revenue="4.00",
        total_cost="0.00",
        linked_transaction_id=None,
        notes=None,
    )

    context = core_logic.load_runtime_context(config_file)
    inventory = core_logic.calculate_inventory(context)
    assert inventory["P10"] == Decimal("3")


def test_calculate_profit_summary_returns_totals(config_file):
    """calculate_profit_summary should return total revenue, cost, and profit."""

    _append_transaction(
        config_file.parent / "master_workbook.xlsx",
        transaction_id="T110",
        timestamp="2025-10-30T03:00:00",
        transaction_type=constants.TransactionType.RESTOCK.value,
        product_id="P10",
        salesman_id=None,
        payment_type=constants.PaymentType.CASH.value,
        quantity_change="5",
        total_revenue="0.00",
        total_cost="-15.00",
        linked_transaction_id=None,
        notes=None,
    )
    _append_transaction(
        config_file.parent / "master_workbook.xlsx",
        transaction_id="T111",
        timestamp="2025-10-30T03:15:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P10",
        salesman_id="S-DEFAULT",
        payment_type=constants.PaymentType.CASH.value,
        quantity_change="-5",
        total_revenue="25.00",
        total_cost="0.00",
        linked_transaction_id=None,
        notes=None,
    )

    context = core_logic.load_runtime_context(config_file)
    summary = core_logic.calculate_profit_summary(context)
    assert summary == {
        "total_revenue": Decimal("25.00"),
        "total_cost": Decimal("-15.00"),
        "profit": Decimal("10.00"),
    }


def test_record_sale_appends_transaction(config_file):
    """record_sale should validate inputs and append a SALE row."""

    _append_product(config_file.parent / "master_workbook.xlsx", product_id="P200", name="Drink", price=Decimal("3.50"))

    context = core_logic.load_runtime_context(config_file)
    command = core_logic.SaleCommand(
        product_id="P200",
        salesman_id="S-DEFAULT",
        quantity=Decimal("2"),
        total_revenue=Decimal("7.00"),
        payment_type=constants.PaymentType.CASH,
        notes="Evening sale",
    )
    transaction = core_logic.record_sale(context, command)
    assert transaction.transaction_type == constants.TransactionType.SALE.value


def test_record_restock_appends_transaction(config_file):
    """record_restock should log incoming inventory with TotalCost."""

    _append_product(config_file.parent / "master_workbook.xlsx", product_id="P201", name="Snack", price=Decimal("2.50"))

    context = core_logic.load_runtime_context(config_file)
    command = core_logic.RestockCommand(
        product_id="P201",
        quantity=Decimal("10"),
        total_cost=Decimal("-12.00"),
        notes="Morning restock",
    )
    transaction = core_logic.record_restock(context, command)
    assert transaction.transaction_type == constants.TransactionType.RESTOCK.value
    assert transaction.quantity_change == Decimal("10")


def test_record_write_off_appends_transaction(config_file):
    """record_write_off should log shrink events with zero revenue/cost."""

    _append_product(config_file.parent / "master_workbook.xlsx", product_id="P202", name="Fruit", price=Decimal("1.25"))

    context = core_logic.load_runtime_context(config_file)
    command = core_logic.WriteOffCommand(
        product_id="P202",
        quantity=Decimal("1"),
        notes="Spoiled",
    )
    transaction = core_logic.record_write_off(context, command)
    assert transaction.transaction_type == constants.TransactionType.WRITE_OFF.value


def test_record_credit_payment_appends_transaction(config_file):
    """record_credit_payment should log cash collection for credit sales."""

    _append_product(config_file.parent / "master_workbook.xlsx", product_id="P203", name="Candy", price=Decimal("1.00"))
    _append_transaction(
        config_file.parent / "master_workbook.xlsx",
        transaction_id="T-credit",
        timestamp="2025-10-30T04:00:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P203",
        salesman_id="S-DEFAULT",
        payment_type=constants.PaymentType.ON_CREDIT.value,
        quantity_change="-2",
        total_revenue="0.00",
        total_cost="0.00",
        linked_transaction_id=None,
        notes="Credit sale",
    )

    context = core_logic.load_runtime_context(config_file)
    command = core_logic.CreditPaymentCommand(
        linked_transaction_id="T-credit",
        total_revenue=Decimal("2.00"),
        notes="Settled",
    )
    transaction = core_logic.record_credit_payment(context, command)
    assert transaction.transaction_type == constants.TransactionType.CREDIT_PAYMENT.value


def test_record_open_stock_appends_transaction(config_file):
    """record_open_stock should log baseline stock during rollover."""

    _append_product(config_file.parent / "master_workbook.xlsx", product_id="P204", name="Water", price=Decimal("1.50"))

    context = core_logic.load_runtime_context(config_file)
    command = core_logic.OpenStockCommand(
        product_id="P204",
        quantity=Decimal("20"),
        total_revenue=Decimal("30.00"),
    )
    transaction = core_logic.record_open_stock(context, command)
    assert transaction.transaction_type == constants.TransactionType.OPEN_STOCK.value


def test_record_void_creates_reversal_and_replacement(config_file):
    """record_void should produce a VOID plus replacement transaction."""

    _append_product(config_file.parent / "master_workbook.xlsx", product_id="P205", name="Tea", price=Decimal("2.00"))
    _append_transaction(
        config_file.parent / "master_workbook.xlsx",
        transaction_id="T-original",
        timestamp="2025-10-30T05:00:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P205",
        salesman_id="S-DEFAULT",
        payment_type=constants.PaymentType.CASH.value,
        quantity_change="-3",
        total_revenue="6.00",
        total_cost="0.00",
        linked_transaction_id=None,
        notes="Incorrect quantity",
    )

    context = core_logic.load_runtime_context(config_file)
    command = core_logic.VoidCommand(
        linked_transaction_id="T-original",
        replacement_command=core_logic.SaleCommand(
            product_id="P205",
            salesman_id="S-DEFAULT",
            quantity=Decimal("1"),
            total_revenue=Decimal("2.00"),
            payment_type=constants.PaymentType.CASH,
            notes="Corrected",
        ),
        notes="Fix entry",
    )
    void_rows = core_logic.record_void(context, command)
    assert len(void_rows) == 2


def test_generate_transaction_id_uses_timestamp():
    """Transaction IDs should be sortable and include the timestamp."""

    when = datetime(2025, 10, 30, 12, 30, 0)
    tx_id = core_logic.generate_transaction_id(when=when)
    assert tx_id.startswith("T20251030")


def test_require_positive_quantity_rejects_nonpositive():
    """Quantities of zero or less should raise ValueError."""

    with pytest.raises(ValueError):
        core_logic.require_positive_quantity(Decimal("0"))


def test_require_positive_quantity_accepts_positive():
    """Positive quantities should pass validation."""

    core_logic.require_positive_quantity(Decimal("1"))


def test_require_nonnegative_money_rejects_negative():
    """Negative currency values should raise ValueError."""

    with pytest.raises(ValueError):
        core_logic.require_nonnegative_money(Decimal("-0.01"))


def test_require_nonnegative_money_accepts_zero():
    """Zero or positive currency values should pass validation."""

    core_logic.require_nonnegative_money(Decimal("0.00"))


def test_persist_context_writes_to_disk(config_file):
    """persist_context should flush workbook changes to disk."""

    context = core_logic.load_runtime_context(config_file)
    context.workbook[constants.SheetName.PRODUCTS.value].append(["P301", "Granola", "3.00", True])

    core_logic.persist_context(context)

    reloaded = openpyxl.load_workbook(config_file.parent / "master_workbook.xlsx")
    values = list(
        reloaded[constants.SheetName.PRODUCTS.value].iter_rows(min_row=2, values_only=True)
    )
    assert ("P301", "Granola", "3.00", True) in values


def test_refresh_context_reloads_from_disk(config_file):
    """refresh_context should discard in-memory workbook state and reload."""

    context = core_logic.load_runtime_context(config_file)
    context.workbook[constants.SheetName.PRODUCTS.value].append(["P302", "Brownie", "2.25", True])
    core_logic.persist_context(context)

    reloaded_context = core_logic.refresh_context(context)
    values = list(
        reloaded_context.workbook[constants.SheetName.PRODUCTS.value].iter_rows(min_row=2, values_only=True)
    )
    assert ("P302", "Brownie", "2.25", True) in values


def test_validate_credit_sale_link_accepts_credit_sale():
    """validate_credit_sale_link should accept undisturbed credit sales."""

    sale = data_manager.TransactionRow(
        transaction_id="Tcredit",
        timestamp_iso="2025-10-30T07:00:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P205",
        salesman_id="S-DEFAULT",
        payment_type=constants.PaymentType.ON_CREDIT.value,
        quantity_change=Decimal("-1"),
        total_revenue=Decimal("0.00"),
        total_cost=Decimal("0.00"),
        linked_transaction_id=None,
        notes=None,
    )
    core_logic.validate_credit_sale_link(sale)


def test_validate_credit_sale_link_rejects_non_credit_sale():
    """validate_credit_sale_link should reject cash sales or mismatched entries."""

    sale = data_manager.TransactionRow(
        transaction_id="Tcash",
        timestamp_iso="2025-10-30T07:30:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P205",
        salesman_id="S-DEFAULT",
        payment_type=constants.PaymentType.CASH.value,
        quantity_change=Decimal("-1"),
        total_revenue=Decimal("2.00"),
        total_cost=Decimal("0.00"),
        linked_transaction_id=None,
        notes=None,
    )
    with pytest.raises(core_logic.BusinessRuleViolation):
        core_logic.validate_credit_sale_link(sale)


def test_validate_void_target_rejects_void_or_credit_payment():
    """validate_void_target should reject transactions that cannot be voided."""

    void_txn = data_manager.TransactionRow(
        transaction_id="Tvoid",
        timestamp_iso="2025-10-30T08:00:00",
        transaction_type=constants.TransactionType.VOID.value,
        product_id="P205",
        salesman_id="S-DEFAULT",
        payment_type=constants.PaymentType.CASH.value,
        quantity_change=Decimal("1"),
        total_revenue=Decimal("-2.00"),
        total_cost=Decimal("0.00"),
        linked_transaction_id="Torig",
        notes=None,
    )
    with pytest.raises(core_logic.BusinessRuleViolation):
        core_logic.validate_void_target(void_txn)


def test_build_void_reversal_inverts_original():
    """build_void_reversal should produce a transaction that cancels the original."""

    original = data_manager.TransactionRow(
        transaction_id="Torig",
        timestamp_iso="2025-10-30T09:00:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P205",
        salesman_id="S-DEFAULT",
        payment_type=constants.PaymentType.CASH.value,
        quantity_change=Decimal("-2"),
        total_revenue=Decimal("4.00"),
        total_cost=Decimal("0.00"),
        linked_transaction_id=None,
        notes="Original",
    )
    reversal_time = datetime(2025, 10, 30, 9, 30, 0)
    reversal = core_logic.build_void_reversal(original, timestamp=reversal_time, notes="Fix")
    assert reversal.transaction_type == constants.TransactionType.VOID.value
    assert reversal.quantity_change == Decimal("2")
    assert reversal.total_revenue == Decimal("-4.00")


def test_build_sale_transaction_constructs_row():
    """build_sale_transaction should convert commands into TransactionRow objects."""

    command = core_logic.SaleCommand(
        product_id="P205",
        salesman_id="S-DEFAULT",
        quantity=Decimal("2"),
        total_revenue=Decimal("6.00"),
        payment_type=constants.PaymentType.CASH,
        timestamp=datetime(2025, 10, 30, 10, 0, 0),
        notes="Morning",
    )
    row = core_logic.build_sale_transaction(command, transaction_id="T-build", timestamp=datetime(2025, 10, 30, 10, 0, 0))
    assert row.transaction_type == constants.TransactionType.SALE.value
    assert row.quantity_change == Decimal("-2")


def test_build_restock_transaction_constructs_row():
    """build_restock_transaction should log positive quantities and negative cost."""

    command = core_logic.RestockCommand(
        product_id="P205",
        quantity=Decimal("5"),
        total_cost=Decimal("-8.00"),
        timestamp=datetime(2025, 10, 30, 11, 0, 0),
        notes="Vendor delivery",
    )
    row = core_logic.build_restock_transaction(command, transaction_id="T-restock", timestamp=datetime(2025, 10, 30, 11, 0, 0))
    assert row.transaction_type == constants.TransactionType.RESTOCK.value
    assert row.quantity_change == Decimal("5")


def test_build_write_off_transaction_constructs_row():
    """build_write_off_transaction should log negative quantity with zero revenue/cost."""

    command = core_logic.WriteOffCommand(
        product_id="P205",
        quantity=Decimal("1"),
        timestamp=datetime(2025, 10, 30, 12, 0, 0),
        notes="Spoilage",
    )
    row = core_logic.build_write_off_transaction(command, transaction_id="T-writeoff", timestamp=datetime(2025, 10, 30, 12, 0, 0))
    assert row.transaction_type == constants.TransactionType.WRITE_OFF.value
    assert row.total_revenue == Decimal("0")
    assert row.total_cost == Decimal("0")


def test_build_credit_payment_transaction_constructs_row():
    """build_credit_payment_transaction should log zero quantity with positive revenue."""

    command = core_logic.CreditPaymentCommand(
        linked_transaction_id="Tcredit",
        total_revenue=Decimal("5.00"),
        timestamp=datetime(2025, 10, 30, 13, 0, 0),
        notes="Payment",
    )
    row = core_logic.build_credit_payment_transaction(command, transaction_id="T-payment", timestamp=datetime(2025, 10, 30, 13, 0, 0))
    assert row.transaction_type == constants.TransactionType.CREDIT_PAYMENT.value
    assert row.quantity_change == Decimal("0")
    assert row.total_revenue == Decimal("5.00")
    assert row.linked_transaction_id == "Tcredit"


def test_build_open_stock_transaction_constructs_row():
    """build_open_stock_transaction should seed balances with positive quantity and revenue."""

    command = core_logic.OpenStockCommand(
        product_id="P205",
        quantity=Decimal("15"),
        total_revenue=Decimal("30.00"),
        timestamp=datetime(2025, 10, 30, 14, 0, 0),
    )
    row = core_logic.build_open_stock_transaction(command, transaction_id="T-open", timestamp=datetime(2025, 10, 30, 14, 0, 0))
    assert row.transaction_type == constants.TransactionType.OPEN_STOCK.value
    assert row.total_revenue == Decimal("30.00")
