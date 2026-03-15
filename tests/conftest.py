import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook import Workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))


"""
Workbook Test Data Plan:

Right now, we already have dynamic workbook fixtures in conftest.py:
make_workbook
products_workbook
salesmen_workbook
transactions_workbook
tmp_workbook_path
These currently create header-only workbooks. For implementation, I will add seeded fixtures dynamically (in-memory) such as:
products_workbook_with_rows
salesmen_workbook_with_rows
transactions_workbook_with_rows
mixed_edge_rows fixtures (blank optional fields, duplicate keys, invalid candidates)
Why dynamic instead of checked-in xlsx files:
faster and deterministic for unit tests
easier to vary per test
no binary fixture maintenance overhead
If we later add integration-style tests for real-world workbook quirks, then I will add a small static fixture set under tests fixtures directory with a few curated xlsx files. For DAL unit tests, dynamic data is the better default.
"""


@pytest.fixture
def make_workbook():
    """Return a factory that creates a workbook with a single configured sheet."""

    def _build(sheet_name: str, headers: list[str]) -> Workbook:
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        return wb

    return _build


@pytest.fixture
def products_workbook(make_workbook) -> Workbook:
    """Create an in-memory workbook with a canonical Products sheet header."""

    return make_workbook(
        "Products",
        ["ProductID", "ProductName", "SellPrice", "IsActive"],
    )


@pytest.fixture
def salesmen_workbook(make_workbook) -> Workbook:
    """Create an in-memory workbook with a canonical Salesmen sheet header."""

    return make_workbook(
        "Salesmen",
        ["SalesmanID", "SalesmanName", "IsActive"],
    )


@pytest.fixture
def transactions_workbook(make_workbook) -> Workbook:
    """Create an in-memory workbook with a canonical TransactionLog sheet header."""

    return make_workbook(
        "TransactionLog",
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ],
    )


@pytest.fixture
def tmp_workbook_path(tmp_path: Path) -> Path:
    """Persist a minimal workbook on disk and return its path for file IO tests."""

    path = tmp_path / "master_workbook.xlsx"
    wb = openpyxl.Workbook()
    wb.save(path)
    return path
