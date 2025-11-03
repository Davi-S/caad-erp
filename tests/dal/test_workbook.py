import openpyxl
import pytest
from openpyxl.workbook import Workbook as OpenpyxlWorkbook

from caad_erp import dal, constants


def test_open_workbook_returns_openpyxl_instance(master_workbook_path):
    """Given an existing workbook path When open_workbook executes Then an openpyxl Workbook returns."""

    # Arrange
    workbook_path = master_workbook_path

    # Act
    workbook = dal.open_workbook(workbook_path)

    # Assert
    assert isinstance(workbook, OpenpyxlWorkbook)


def test_open_workbook_missing_file_raises(tmp_path):
    """Given a missing workbook path When open_workbook runs Then FileNotFoundError is raised."""

    # Arrange
    missing_path = tmp_path / "missing.xlsx"

    # Act / Assert
    with pytest.raises(FileNotFoundError):
        dal.open_workbook(missing_path)


def test_save_workbook_persists_changes(master_workbook_path):
    """Given workbook mutations When save_workbook persists them Then disk reads include the updates."""

    # Arrange
    workbook = dal.open_workbook(master_workbook_path)
    sheet = workbook[constants.SheetName.PRODUCTS.value]
    sheet.append(["P100", "Chips", "2.50", True])

    # Act
    dal.save_workbook(workbook, master_workbook_path)
    reloaded = dal.open_workbook(master_workbook_path)
    values = list(
        reloaded[constants.SheetName.PRODUCTS.value].iter_rows(
            min_row=2, values_only=True)
    )

    # Assert
    assert values == [("P100", "Chips", "2.50", True)]


def test_save_workbook_with_destination_creates_copy(master_workbook_path, tmp_path):
    """Given a destination path When save_workbook writes to it Then a copied file exists independently."""

    # Arrange
    workbook = dal.open_workbook(master_workbook_path)
    sheet = workbook[constants.SheetName.SALESMEN.value]
    sheet.append(["S2", "Jordan", True])
    copy_path = tmp_path / "copy.xlsx"

    # Act
    dal.save_workbook(workbook, destination=copy_path)
    copy = openpyxl.load_workbook(copy_path)
    rows = list(copy[constants.SheetName.SALESMEN.value].iter_rows(
        min_row=2, values_only=True))

    # Assert
    assert ("S2", "Jordan", True) in rows


def test_refresh_workbook_returns_new_instance(master_workbook_path):
    """Given prior workbook changes When refresh_workbook runs Then a new instance reflecting disk state is returned."""

    # Arrange
    original = dal.open_workbook(master_workbook_path)
    sheet = original[constants.SheetName.PRODUCTS.value]
    sheet.append(["P200", "Bars", "4.00", True])
    dal.save_workbook(original, master_workbook_path)

    # Act
    refreshed = dal.refresh_workbook(master_workbook_path)
    values = list(
        refreshed[constants.SheetName.PRODUCTS.value].iter_rows(
            min_row=2, values_only=True)
    )

    # Assert
    assert refreshed is not original
    assert ("P200", "Bars", "4.00", True) in values


def test_locate_row_returns_row_index(master_workbook_path):
    """Given a matching key When locate_row executes Then the correct worksheet index is returned."""

    # Arrange
    workbook = dal.open_workbook(master_workbook_path)
    sheet = workbook[constants.SheetName.PRODUCTS.value]
    sheet.append(["P600", "Snack", "2.00", True])
    dal.save_workbook(workbook, master_workbook_path)

    # Act
    reloaded = dal.open_workbook(master_workbook_path)
    row_index = dal.locate_row(
        reloaded,
        constants.SheetName.PRODUCTS.value,
        "ProductID",
        "P600",
    )

    # Assert
    assert row_index == 2


def test_locate_row_returns_none_when_missing(master_workbook_path):
    """Given an absent key When locate_row executes Then None is returned."""

    # Arrange
    workbook = dal.open_workbook(master_workbook_path)

    # Act
    result = dal.locate_row(
        workbook,
        constants.SheetName.PRODUCTS.value,
        "ProductID",
        "NOPE",
    )

    # Assert
    assert result is None
