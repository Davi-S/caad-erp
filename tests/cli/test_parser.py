import argparse
import importlib
import io
import sys
import types
from pathlib import Path

import openpyxl
import pytest

from caad_erp import constants, exceptions
from caad_erp.cli import command_spec, parser


def _make_data_and_config(tmp_path: Path) -> tuple[Path, Path]:
    data_file = tmp_path / "data.xlsx"
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    salesmen.append(["S001", "Default", True])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append([
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ])
    wb.save(data_file)

    config_path = tmp_path / "config.ini"
    config_path.write_text(
        "\n".join(
            [
                "[System]",
                f"DataFile = {data_file}",
                "LoungeName = Test Lounge",
                f"SchemaVersion = {constants.EXPECTED_SCHEMA_VERSION}",
                "",
                "[Defaults]",
                "DefaultSalesman = S001",
                "",
            ]
        ),
        encoding="utf-8",
    )
    return data_file, config_path


def _commands_dir() -> Path:
    package = importlib.import_module("caad_erp.cli.commands")
    return Path(next(iter(package.__path__)))


def _temp_command_module(name: str, content: str):
    commands_dir = _commands_dir()
    path = commands_dir / f"{name}.py"
    path.write_text(content, encoding="utf-8")
    importlib.invalidate_caches()
    return path


def _cleanup_temp_module(name: str, path: Path) -> None:
    try:
        path.unlink(missing_ok=True)
    finally:
        importlib.invalidate_caches()
        sys.modules.pop(f"caad_erp.cli.commands.{name}", None)


def test_build_parser_includes_config_option() -> None:
    """
    GIVEN no parser instance exists
    WHEN build_parser is called
    THEN returned parser exposes optional --config argument with Path typing
    """
    # Arrange / Act
    parse = parser.build_parser()
    args = parse.parse_args([])

    # Assert
    assert hasattr(args, "config")
    assert args.config is None


def test_configure_subcommands_registers_discovered_specs_and_repl() -> None:
    """
    GIVEN a base parser and discovered command specs
    WHEN configure_subcommands is called
    THEN specs are registered and repl command is available
    """
    # Arrange
    parse = parser.build_parser()

    # Act
    command_table = parser.configure_subcommands(parse)

    # Assert
    assert "repl" not in command_table
    help_text = parse.format_help()
    assert "repl" in help_text


def test_configure_subcommands_returns_command_table_from_specs() -> None:
    """
    GIVEN discovered command specs
    WHEN configure_subcommands is called
    THEN returned mapping indexes every command by name
    """
    # Arrange
    parse = parser.build_parser()

    # Act
    command_table = parser.configure_subcommands(parse)

    # Assert
    assert "sale" in command_table
    assert "stock" in command_table
    assert command_table["sale"].name == "sale"


def test_discover_command_specs_returns_sorted_tuple() -> None:
    """
    GIVEN command modules with valid register factories
    WHEN discover_command_specs is called
    THEN returned command specs are sorted by command name
    """
    # Arrange / Act
    specs = parser.discover_command_specs()

    # Assert
    names = [spec.name for spec in specs]
    assert names == sorted(names)


def test_discover_command_specs_skips_package_entries() -> None:
    """
    GIVEN pkgutil iteration containing nested package entries
    WHEN discover_command_specs is called
    THEN package entries are skipped and only modules are considered
    """
    # Arrange
    commands_dir = _commands_dir()
    package_dir = commands_dir / "zzz_tmp_nested_pkg"
    package_dir.mkdir(exist_ok=True)
    (package_dir / "__init__.py").write_text("", encoding="utf-8")
    importlib.invalidate_caches()

    # Act / Assert
    try:
        specs = parser.discover_command_specs()
        assert isinstance(specs, tuple)
    finally:
        (package_dir / "__init__.py").unlink(missing_ok=True)
        package_dir.rmdir()
        importlib.invalidate_caches()


def test_discover_command_specs_raises_for_missing_register_factory() -> None:
    """
    GIVEN a command module without expected register_<module>_command function
    WHEN discover_command_specs is called
    THEN ValueError is raised
    """
    # Arrange
    name = "zzz_tmp_missing_register"
    path = _temp_command_module(name, "X = 1\n")

    # Act / Assert
    try:
        with pytest.raises(ValueError):
            parser.discover_command_specs()
    finally:
        _cleanup_temp_module(name, path)


def test_discover_command_specs_raises_for_non_callable_factory() -> None:
    """
    GIVEN a command module whose register attribute is non-callable
    WHEN discover_command_specs is called
    THEN TypeError is raised
    """
    # Arrange
    name = "zzz_tmp_non_callable"
    content = "register_zzz_tmp_non_callable_command = 123\n"
    path = _temp_command_module(name, content)

    # Act / Assert
    try:
        with pytest.raises(TypeError):
            parser.discover_command_specs()
    finally:
        _cleanup_temp_module(name, path)


def test_discover_command_specs_raises_for_invalid_factory_return_type() -> None:
    """
    GIVEN a register factory returning a non-CommandSpec value
    WHEN discover_command_specs is called
    THEN TypeError is raised
    """
    # Arrange
    name = "zzz_tmp_invalid_return"
    content = "def register_zzz_tmp_invalid_return_command():\n    return 42\n"
    path = _temp_command_module(name, content)

    # Act / Assert
    try:
        with pytest.raises(TypeError):
            parser.discover_command_specs()
    finally:
        _cleanup_temp_module(name, path)


def test_discover_command_specs_raises_when_package_lacks_path() -> None:
    """
    GIVEN imported commands package object without __path__ attribute
    WHEN discover_command_specs is called
    THEN ValueError is raised
    """
    # Arrange
    original = sys.modules.get("caad_erp.cli.commands")
    fake_package = types.ModuleType("caad_erp.cli.commands")
    sys.modules["caad_erp.cli.commands"] = fake_package

    # Act / Assert
    try:
        with pytest.raises(ValueError):
            parser.discover_command_specs()
    finally:
        if original is None:
            sys.modules.pop("caad_erp.cli.commands", None)
        else:
            sys.modules["caad_erp.cli.commands"] = original


def test_dispatch_command_executes_selected_spec() -> None:
    """
    GIVEN parsed args containing a known command and command table entry
    WHEN dispatch_command is called
    THEN selected command spec execute function is invoked and exit code returned
    """
    # Arrange
    calls = []

    def _execute(context, args):
        calls.append((context, args.command))
        return 7

    spec = command_spec.CommandSpec(
        name="x",
        help_text="x",
        register=lambda action: action.add_parser("x"),
        execute=_execute,
    )
    table = {"x": spec}
    args = argparse.Namespace(command="x")

    # Act
    exit_code = parser.dispatch_command(
        context=object(), args=args, command_table=table)

    # Assert
    assert exit_code == 7
    assert calls[0][1] == "x"


def test_dispatch_command_raises_when_command_missing_in_args() -> None:
    """
    GIVEN parsed args without command attribute
    WHEN dispatch_command is called
    THEN KeyError is raised
    """
    # Arrange
    args = argparse.Namespace()

    # Act / Assert
    with pytest.raises(KeyError):
        parser.dispatch_command(context=object(), args=args, command_table={})


def test_dispatch_command_raises_when_command_is_none() -> None:
    """
    GIVEN parsed args with command set to None
    WHEN dispatch_command is called
    THEN KeyError is raised
    """
    # Arrange
    args = argparse.Namespace(command=None)

    # Act / Assert
    with pytest.raises(KeyError):
        parser.dispatch_command(context=object(), args=args, command_table={})


def test_dispatch_command_raises_for_unknown_command_name() -> None:
    """
    GIVEN parsed args with command not present in command table
    WHEN dispatch_command is called
    THEN KeyError is raised
    """
    # Arrange
    args = argparse.Namespace(command="unknown")

    # Act / Assert
    with pytest.raises(KeyError):
        parser.dispatch_command(context=object(), args=args, command_table={})


def test_build_command_table_indexes_specs_by_name() -> None:
    """
    GIVEN iterable of unique command specs
    WHEN build_command_table is called
    THEN mapping is returned keyed by command names
    """
    # Arrange
    spec_a = command_spec.CommandSpec(
        name="a",
        help_text="a",
        register=lambda action: action.add_parser("a"),
        execute=lambda context, args: 0,
    )
    spec_b = command_spec.CommandSpec(
        name="b",
        help_text="b",
        register=lambda action: action.add_parser("b"),
        execute=lambda context, args: 0,
    )

    # Act
    table = parser.build_command_table([spec_a, spec_b])

    # Assert
    assert set(table.keys()) == {"a", "b"}


def test_build_command_table_raises_on_duplicate_names() -> None:
    """
    GIVEN iterable of command specs with duplicate names
    WHEN build_command_table is called
    THEN ValueError is raised
    """
    # Arrange
    spec_a1 = command_spec.CommandSpec(
        name="a",
        help_text="a",
        register=lambda action: action.add_parser("a"),
        execute=lambda context, args: 0,
    )
    spec_a2 = command_spec.CommandSpec(
        name="a",
        help_text="a2",
        register=lambda action: action.add_parser("a2"),
        execute=lambda context, args: 0,
    )

    # Act / Assert
    with pytest.raises(ValueError):
        parser.build_command_table([spec_a1, spec_a2])


@pytest.mark.parametrize(
    "error_obj, expected_code",
    [
        (exceptions.BusinessRuleViolation("rule"), 2),
        (FileNotFoundError("missing"), 3),
        (RuntimeError("boom"), 1),
    ],
)
def test_handle_cli_error_maps_exception_categories(error_obj, expected_code) -> None:
    """
    GIVEN a CLI exception category
    WHEN handle_cli_error is called
    THEN corresponding shell exit code is returned
    """
    # Arrange / Act
    code = parser.handle_cli_error(error_obj)

    # Assert
    assert code == expected_code


def test_main_runs_repl_when_no_command_is_provided(tmp_path: Path) -> None:
    """
    GIVEN parsed arguments with command omitted
    WHEN main is called
    THEN REPL is invoked and its exit code is returned
    """
    # Arrange
    _, config = _make_data_and_config(tmp_path)
    original_stdin = sys.stdin
    sys.stdin = io.StringIO("exit\n")

    # Act
    try:
        exit_code = parser.main(["--config", str(config)])
    finally:
        sys.stdin = original_stdin

    # Assert
    assert exit_code == 0


def test_main_runs_repl_when_command_is_repl(tmp_path: Path) -> None:
    """
    GIVEN parsed arguments with command equal to repl
    WHEN main is called
    THEN REPL is invoked and its exit code is returned
    """
    # Arrange
    _, config = _make_data_and_config(tmp_path)
    original_stdin = sys.stdin
    sys.stdin = io.StringIO("exit\n")

    # Act
    try:
        exit_code = parser.main(["--config", str(config), "repl"])
    finally:
        sys.stdin = original_stdin

    # Assert
    assert exit_code == 0


def test_main_dispatches_command_and_persists_on_successful_mutation(tmp_path: Path) -> None:
    """
    GIVEN parsed mutating command returning zero exit status
    WHEN main is called
    THEN command is dispatched and context is persisted
    """
    # Arrange
    data_file, config = _make_data_and_config(tmp_path)

    # Act
    exit_code = parser.main(
        [
            "--config",
            str(config),
            "add-salesman",
            "--salesman-id",
            "S002",
            "--salesman-name",
            "Bob",
        ]
    )

    # Assert
    assert exit_code == 0
    reloaded = openpyxl.load_workbook(data_file)
    salesmen_sheet = reloaded[constants.SheetName.SALESMEN.value]
    values = [row[0]
              for row in salesmen_sheet.iter_rows(min_row=2, values_only=True)]
    assert "S002" in values


def test_main_skips_persist_for_non_mutating_commands(tmp_path: Path) -> None:
    """
    GIVEN parsed non-mutating command returning zero exit status
    WHEN main is called
    THEN command is dispatched without persistence
    """
    # Arrange
    data_file, config = _make_data_and_config(tmp_path)
    before = data_file.read_bytes()

    # Act
    exit_code = parser.main(["--config", str(config), "profit"])

    # Assert
    assert exit_code == 0
    after = data_file.read_bytes()
    assert before == after


def test_main_skips_persist_for_nonzero_exit_code(tmp_path: Path) -> None:
    """
    GIVEN parsed mutating command returning nonzero exit status
    WHEN main is called
    THEN command exit code is returned without persistence
    """
    # Arrange
    data_file, config = _make_data_and_config(tmp_path)
    before = data_file.read_bytes()

    # Act
    exit_code = parser.main(
        [
            "--config",
            str(config),
            "add-salesman",
            "--salesman-id",
            "S001",
            "--salesman-name",
            "Duplicate",
        ]
    )

    # Assert
    assert exit_code == 2
    after = data_file.read_bytes()
    assert before == after


def test_main_routes_exceptions_to_cli_error_handler(tmp_path: Path) -> None:
    """
    GIVEN an exception raised during context load or command execution
    WHEN main is called
    THEN handle_cli_error return value becomes the CLI exit code
    """
    # Arrange
    missing_config = tmp_path / "missing.ini"

    # Act
    exit_code = parser.main(["--config", str(missing_config), "profit"])

    # Assert
    assert exit_code == 3
