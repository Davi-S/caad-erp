from decimal import Decimal
from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants, dal
from caad_erp.cli.commands import debts
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    products.append(["P001", "Cookie", Decimal("10.00"), True])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    salesmen.append(["S001", "Ana", True])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append([
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ])
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_debts_command_returns_non_mutating_command_spec() -> None:
    """
    GIVEN debts command module
    WHEN register_debts_command is called
    THEN a non-mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = debts.register_debts_command()

    # Assert
    assert spec.name == "debts"
    assert spec.is_mutating is False


def test_register_debts_registrar_configures_command_default() -> None:
    """
    GIVEN subparser factory instance
    WHEN debts registrar is executed
    THEN parser sets command default without additional required args
    """
    # Arrange
    spec = debts.register_debts_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(["debts"])

    # Assert
    assert args.command == "debts"


def test_display_debts_report_prints_empty_state_when_no_balances(capsys) -> None:
    """
    GIVEN summary mapping without outstanding balances
    WHEN _display_debts_report is called
    THEN empty-state message and total outstanding are printed
    """
    # Arrange / Act
    debts._display_debts_report(
        {"balances": [], "total_outstanding": Decimal("0")})

    # Assert
    output = capsys.readouterr().out
    assert "No outstanding credit balances." in output
    assert "Total outstanding" in output


@pytest.mark.parametrize(
    "debt_rows",
    [
        [
            bll.OutstandingDebt(
                transaction_id="SALE1",
                timestamp_iso="2026-03-15T10:00:00+00:00",
                product_id="P001",
                salesman_id="S001",
                quantity=Decimal("2"),
                expected_amount=Decimal("20"),
                amount_paid=Decimal("5"),
                balance=Decimal("15"),
            )
        ]
    ],
)
def test_display_debts_report_prints_table_for_outstanding_balances(debt_rows, capsys) -> None:
    """
    GIVEN summary mapping containing outstanding debt rows
    WHEN _display_debts_report is called
    THEN tabular debt lines and total outstanding are printed
    """
    # Arrange
    summary = {"balances": debt_rows, "total_outstanding": Decimal("15")}

    # Act
    debts._display_debts_report(summary)

    # Assert
    output = capsys.readouterr().out
    assert "Outstanding credit balances:" in output
    assert "SALE1" in output
    assert "Total outstanding: 15" in output


def test_run_debts_report_calls_bll_and_returns_zero(tmp_path: Path, capsys) -> None:
    """
    GIVEN runtime context and parsed debts args
    WHEN _run_debts_report is called
    THEN outstanding debt summary is computed displayed and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    dal.append_transaction(
        context.workbook,
        dal.TransactionRow(
            transaction_id="SALE1",
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=constants.TransactionType.SALE.value,
            product_id="P001",
            salesman_id="S001",
            payment_type=constants.PaymentType.ON_CREDIT.value,
            quantity_change=Decimal("-1"),
            total_revenue=Decimal("0"),
            total_cost=Decimal("0"),
            linked_transaction_id=None,
            notes=None,
        ),
    )

    # Act
    exit_code = debts._run_debts_report(context, argparse.Namespace())

    # Assert
    assert exit_code == 0
    assert "Outstanding credit balances:" in capsys.readouterr().out
