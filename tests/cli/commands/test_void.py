from pathlib import Path

import argparse
import openpyxl

from caad_erp import bll, constants, dal
from caad_erp.cli.commands import void
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    products.append(["P001", "Cookie", 250, True])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    salesmen.append(["S001", "Ana", True])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append([
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ])
    dal.append_transaction(
        wb,
        dal.TransactionRow(
            transaction_id="SALE1",
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=constants.TransactionType.SALE.value,
            product_id="P001",
            salesman_id="S001",
            payment_type=constants.PaymentType.CASH.value,
            quantity_change=-1,
            total_revenue=250,
            total_cost=0,
            linked_transaction_id=None,
            notes=None,
        ),
    )
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_void_command_returns_command_spec() -> None:
    """
    GIVEN void command module
    WHEN register_void_command is called
    THEN a mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = void.register_void_command()

    # Assert
    assert spec.name == "void"
    assert spec.is_mutating is True


def test_register_void_registrar_configures_required_arguments() -> None:
    """
    GIVEN subparser factory instance
    WHEN void registrar is executed
    THEN parser includes required linked-transaction-id and optional notes
    """
    # Arrange
    spec = void.register_void_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(["void", "--linked-transaction-id", "SALE1"])

    # Assert
    assert args.command == "void"
    assert args.linked_transaction_id == "SALE1"


def test_translate_void_maps_args_to_void_command() -> None:
    """
    GIVEN parsed void args with valid values
    WHEN _translate_void is called
    THEN VoidCommand is built with linked transaction id and notes
    """
    # Arrange
    args = argparse.Namespace(linked_transaction_id="SALE1", notes="fix")

    # Act
    command = void._translate_void(args)

    # Assert
    assert command.linked_transaction_id == "SALE1"
    assert command.notes == "fix"


def test_run_void_calls_bll_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed void args
    WHEN _run_void is called
    THEN bll.record_void is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(linked_transaction_id="SALE1", notes="fix")

    # Act
    exit_code = void._run_void(context, args)

    # Assert
    assert exit_code == 0
    rows = bll.list_transactions(context)
    assert len(rows) == 2
    assert rows[-1].transaction_type == constants.TransactionType.VOID.value
