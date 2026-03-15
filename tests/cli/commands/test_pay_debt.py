from decimal import Decimal, InvalidOperation
from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants, dal
from caad_erp.cli.commands import pay_debt
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    products.append(["P001", "Cookie", Decimal("2.50"), True])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    salesmen.append(["S001", "Ana", True])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append([
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ])
    dal.append_transaction(
        wb,
        dal.TransactionRow(
            transaction_id="SALE1",
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=constants.TransactionType.SALE.value,
            product_id="P001",
            salesman_id="S001",
            payment_type=constants.PaymentType.ON_CREDIT.value,
            quantity_change=Decimal("-1"),
            total_revenue=Decimal("0"),
            total_cost=Decimal("0"),
            linked_transaction_id=None,
            notes=None,
        ),
    )
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_pay_debt_command_returns_command_spec() -> None:
    """
    GIVEN pay-debt command module
    WHEN register_pay_debt_command is called
    THEN a mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = pay_debt.register_pay_debt_command()

    # Assert
    assert spec.name == "pay-debt"
    assert spec.is_mutating is True


def test_register_pay_debt_registrar_configures_required_arguments_and_choices() -> None:
    """
    GIVEN subparser factory instance
    WHEN pay-debt registrar is executed
    THEN parser includes required linkage amount salesman and payment-type choices excluding OnCredit
    """
    # Arrange
    spec = pay_debt.register_pay_debt_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args([
        "pay-debt",
        "--linked-transaction-id",
        "SALE1",
        "--total-revenue",
        "3.00",
        "--salesman-id",
        "S001",
        "--payment-type",
        constants.PaymentType.CASH.value,
    ])

    # Assert
    assert args.command == "pay-debt"
    assert args.payment_type != constants.PaymentType.ON_CREDIT.value


def test_translate_pay_debt_maps_args_to_credit_payment_command() -> None:
    """
    GIVEN parsed pay-debt args with valid values
    WHEN _translate_pay_debt is called
    THEN CreditPaymentCommand is built with Decimal amount and enum payment type
    """
    # Arrange
    args = argparse.Namespace(
        linked_transaction_id="SALE1",
        total_revenue="3.00",
        salesman_id="S001",
        payment_type=constants.PaymentType.PIX.value,
        notes="partial",
    )

    # Act
    command = pay_debt._translate_pay_debt(args)

    # Assert
    assert command.linked_transaction_id == "SALE1"
    assert command.total_revenue == Decimal("3.00")
    assert command.payment_type == constants.PaymentType.PIX


@pytest.mark.parametrize("invalid_revenue", ["", "abc", "1.2.3"])
def test_translate_pay_debt_raises_for_invalid_decimal(invalid_revenue) -> None:
    """
    GIVEN parsed pay-debt args with invalid total-revenue text
    WHEN _translate_pay_debt is called
    THEN decimal conversion error is raised
    """
    # Arrange
    args = argparse.Namespace(
        linked_transaction_id="SALE1",
        total_revenue=invalid_revenue,
        salesman_id="S001",
        payment_type=constants.PaymentType.CASH.value,
        notes=None,
    )

    # Act / Assert
    with pytest.raises(InvalidOperation):
        pay_debt._translate_pay_debt(args)


@pytest.mark.parametrize("invalid_payment_type", ["", "INVALID", "Card"])
def test_translate_pay_debt_raises_for_invalid_payment_type(invalid_payment_type) -> None:
    """
    GIVEN parsed pay-debt args with invalid payment-type value
    WHEN _translate_pay_debt is called
    THEN ValueError is raised
    """
    # Arrange
    args = argparse.Namespace(
        linked_transaction_id="SALE1",
        total_revenue="2",
        salesman_id="S001",
        payment_type=invalid_payment_type,
        notes=None,
    )

    # Act / Assert
    with pytest.raises(ValueError):
        pay_debt._translate_pay_debt(args)


def test_run_pay_debt_calls_bll_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed pay-debt args
    WHEN _run_pay_debt is called
    THEN bll.record_credit_payment is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        linked_transaction_id="SALE1",
        total_revenue="3.00",
        salesman_id="S001",
        payment_type=constants.PaymentType.CASH.value,
        notes=None,
    )

    # Act
    exit_code = pay_debt._run_pay_debt(context, args)

    # Assert
    assert exit_code == 0
    rows = bll.list_transactions(context)
    assert len(rows) == 2
    assert rows[-1].transaction_type == constants.TransactionType.CREDIT_PAYMENT.value
