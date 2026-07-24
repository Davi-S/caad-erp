from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants, exceptions
from caad_erp.cli.commands import write_off
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    products.append(["P001", "Cookie", 250, True])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    salesmen.append(["S001", "Ana", True])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append(
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ]
    )
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_write_off_command_returns_command_spec() -> None:
    """
    GIVEN write-off command module
    WHEN register_write_off_command is called
    THEN a mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = write_off.register_write_off_command()

    # Assert
    assert spec.name == "write-off"
    assert spec.is_mutating is True


def test_register_write_off_registrar_configures_required_arguments() -> None:
    """
    GIVEN subparser factory instance
    WHEN write-off registrar is executed
    THEN parser includes required product quantity salesman and optional notes
    """
    # Arrange
    spec = write_off.register_write_off_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(
        [
            "write-off",
            "--product-id",
            "P001",
            "--quantity",
            "1",
            "--salesman-id",
            "S001",
        ]
    )

    # Assert
    assert args.command == "write-off"
    assert args.quantity == "1"


def test_run_write_off_calls_bll_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed write-off args
    WHEN _run_write_off is called
    THEN bll.record_write_off is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        product_id="P001",
        quantity="1",
        salesman_id="S001",
        notes=None,
    )

    # Act
    exit_code = write_off._run_write_off(context, args)

    # Assert
    assert exit_code == 0
    rows = bll.list_transactions(context)
    assert len(rows) == 1
    assert rows[0].transaction_type == constants.TransactionType.WRITE_OFF.value


def test_run_write_off_returns_nonzero_exit_code_on_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """
    GIVEN runtime context and write-off args when bll raises BusinessRuleViolation
    WHEN _run_write_off is called
    THEN non-zero exit code 2 is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        product_id="P001",
        quantity="1",
        salesman_id="S001",
        notes=None,
    )

    def _mock_record_write_off(*args, **kwargs):
        raise exceptions.BusinessRuleViolation("Product inactive")

    monkeypatch.setattr(bll, "record_write_off", _mock_record_write_off)

    # Act
    exit_code = write_off._run_write_off(context, args)

    # Assert
    assert exit_code == 2
