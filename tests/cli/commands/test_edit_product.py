from pathlib import Path
import argparse
import openpyxl

from caad_erp import bll, constants
from caad_erp.cli.commands import edit_product
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    products.append(["P001", "Cookie", 250, True])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append(
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ]
    )
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_edit_product_command_returns_command_spec() -> None:
    """
    GIVEN edit-product command module
    WHEN register_edit_product_command is called
    THEN a CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = edit_product.register_edit_product_command()

    # Assert
    assert spec.name == "edit-product"


def test_register_edit_product_registrar_configures_arguments() -> None:
    """
    GIVEN subparser factory instance
    WHEN edit-product registrar is executed
    THEN parser includes expected product arguments and command default
    """
    # Arrange
    spec = edit_product.register_edit_product_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(
        [
            "edit-product",
            "--product-id",
            "P001",
            "--product-name",
            "New Cookie",
            "--product-sell-price",
            "300",
            "--product-is-active",
        ]
    )

    # Assert
    assert args.command == "edit-product"
    assert args.product_id == "P001"
    assert args.product_name == "New Cookie"
    assert args.product_sell_price == "300"
    assert args.product_is_active is True


def test_translate_edit_product_sets_fields_and_trims_id() -> None:
    """
    GIVEN parsed edit-product args with product values
    WHEN _translate_edit_product is called
    THEN ProductCommand is produced with trimmed id and updated fields
    """
    # Arrange
    # Note: providing 'inactive' to match the current implementation in edit_product.py
    args = argparse.Namespace(
        product_id="  P001  ",
        product_name="New Cookie",
        product_sell_price="300",
        inactive=True,
    )

    # Act
    command = edit_product._translate_edit_product(args)

    # Assert
    assert command.product_id == "P001"
    assert command.product_name == "New Cookie"
    assert command.sell_price == 300
    assert command.is_active is True


def test_run_edit_product_calls_bll_update_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed edit-product args
    WHEN _run_edit_product is called
    THEN bll.update_product is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        product_id="P001",
        product_name="New Cookie",
        product_sell_price="300",
        inactive=False,
    )

    # Act
    exit_code = edit_product._run_edit_product(context, args)

    # Assert
    assert exit_code == 0
    updated = bll.get_product(context, "P001")
    assert updated.product_name == "New Cookie"
    assert updated.sell_price == 300
    assert updated.is_active is False
