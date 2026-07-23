from pathlib import Path

import argparse
import openpyxl

from caad_erp import bll, constants
from caad_erp.cli.commands import list_salesmen
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    salesmen.append(["S001", "Ana", True])
    salesmen.append(["S002", "Bob", False])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append(
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ]
    )
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_list_salesmen_command_returns_non_mutating_command_spec() -> None:
    """
    GIVEN list-salesmen command module
    WHEN register_list_salesmen_command is called
    THEN a non-mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = list_salesmen.register_list_salesmen_command()

    # Assert
    assert spec.name == "list-salesmen"
    assert spec.is_mutating is False


def test_run_list_salesmen_report_calls_bll_and_returns_zero(
    tmp_path: Path, capsys
) -> None:
    """
    GIVEN runtime context and parsed list-salesmen
    WHEN _run_list_salesmen_report is called
    THEN salesmen are listed and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(salesman_id=None)

    # Act
    exit_code = list_salesmen._run_list_salesmen_report(context, args)

    # Assert
    assert exit_code == 0
    output = capsys.readouterr().out
    assert "S001" in output
    assert "S002" in output


def test_run_list_salesmen_report_reports_only_specific_salesman(
    tmp_path: Path, capsys
) -> None:
    """
    GIVEN runtime context and parsed list-salesmen with salesman id
    WHEN _run_list_salesmen_report is called
    THEN only the salesman with the given id is listed and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(salesman_id="S001")

    # Act
    exit_code = list_salesmen._run_list_salesmen_report(context, args)

    # Assert
    assert exit_code == 0
    output = capsys.readouterr().out
    assert "S001" in output
    assert "S002" not in output
