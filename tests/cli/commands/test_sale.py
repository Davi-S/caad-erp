from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants, dal
from caad_erp.cli.commands import sale
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    products.append(["P001", "Cookie", 250, True])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    salesmen.append(["S001", "Ana", True])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append([
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ])
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_sale_command_returns_command_spec() -> None:
    """
    GIVEN sale command module
    WHEN register_sale_command is called
    THEN a mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = sale.register_sale_command()

    # Assert
    assert spec.name == "sale"
    assert spec.is_mutating is True


def test_register_sale_registrar_configures_required_arguments_and_choices() -> None:
    """
    GIVEN subparser factory instance
    WHEN sale registrar is executed
    THEN parser includes required fields and payment-type choices from PaymentType
    """
    # Arrange
    spec = sale.register_sale_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args([
        "sale",
        "--product-id",
        "P001",
        "--quantity",
        "2",
        "--salesman-id",
        "S001",
        "--total-revenue",
        "5.00",
        "--payment-type",
        constants.PaymentType.CASH.value,
    ])

    # Assert
    assert args.command == "sale"
    assert args.payment_type == constants.PaymentType.CASH.value


def test_translate_sale_maps_args_to_sale_command() -> None:
    """
    GIVEN parsed sale args with valid values
    WHEN _translate_sale is called
    THEN SaleCommand is built with integer conversions and enum payment type
    """
    # Arrange
    args = argparse.Namespace(
        product_id="P001",
        quantity=2,
        salesman_id="S001",
        total_revenue=500,
        payment_type=constants.PaymentType.PIX.value,
        notes="note",
    )

    # Act
    command = sale._translate_sale(args)

    # Assert
    assert command.product_id == "P001"
    assert command.quantity == 2
    assert command.total_revenue == 500
    assert command.payment_type == constants.PaymentType.PIX


@pytest.mark.parametrize("invalid_payment_type", ["", "CreditCard", "INVALID"])
def test_translate_sale_raises_for_invalid_payment_type(invalid_payment_type) -> None:
    """
    GIVEN parsed sale args with invalid payment-type value
    WHEN _translate_sale is called
    THEN ValueError is raised
    """
    # Arrange
    args = argparse.Namespace(
        product_id="P001",
        quantity="1",
        salesman_id="S001",
        total_revenue="5",
        payment_type=invalid_payment_type,
        notes=None,
    )

    # Act / Assert
    with pytest.raises(ValueError):
        sale._translate_sale(args)


def test_run_sale_calls_bll_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed sale args
    WHEN _run_sale is called
    THEN bll.record_sale is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        product_id="P001",
        quantity="2",
        salesman_id="S001",
        total_revenue=500,
        payment_type=constants.PaymentType.CASH.value,
        notes=None,
    )

    # Act
    exit_code = sale._run_sale(context, args)

    # Assert
    assert exit_code == 0
    rows = bll.list_transactions(context)
    assert len(rows) == 1
    assert rows[0].transaction_type == constants.TransactionType.SALE.value
