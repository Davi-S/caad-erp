from decimal import Decimal
from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants, dal
from caad_erp.cli.commands import log
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append([
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ])
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_log_command_returns_non_mutating_command_spec() -> None:
    """
    GIVEN log command module
    WHEN register_log_command is called
    THEN a non-mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = log.register_log_command()

    # Assert
    assert spec.name == "log"
    assert spec.is_mutating is False


def test_register_log_registrar_configures_command_default() -> None:
    """
    GIVEN subparser factory instance
    WHEN log registrar is executed
    THEN parser sets command default without additional required args
    """
    # Arrange
    spec = log.register_log_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(["log"])

    # Assert
    assert args.command == "log"


def test_display_transaction_log_prints_empty_message_for_no_entries(capsys) -> None:
    """
    GIVEN empty transaction iterable
    WHEN _display_transaction_log is called
    THEN empty-state message is printed
    """
    # Arrange / Act
    log._display_transaction_log([])

    # Assert
    assert "Transaction log is empty." in capsys.readouterr().out


@pytest.mark.parametrize("notes_value", [None, "", "short", "x" * 60])
def test_display_transaction_log_formats_rows_and_notes(notes_value, capsys) -> None:
    """
    GIVEN transaction rows with optional product salesman timestamp and notes values
    WHEN _display_transaction_log is called
    THEN row formatting applies defaults timestamp truncation and notes truncation rules
    """
    # Arrange
    row = dal.TransactionRow(
        transaction_id="T1",
        timestamp_iso="2026-03-15T10:00:00+00:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id=None,
        salesman_id=None,
        payment_type=constants.PaymentType.CASH.value,
        quantity_change=Decimal("-1"),
        total_revenue=Decimal("2.50"),
        total_cost=Decimal("0"),
        linked_transaction_id=None,
        notes=notes_value,
    )

    # Act
    log._display_transaction_log([row])

    # Assert
    output = capsys.readouterr().out
    assert "Transaction log:" in output
    assert "2026-03-15T10:00:00" in output
    if notes_value and len(notes_value) > 40:
        assert "..." in output


def test_run_log_report_calls_bll_and_returns_zero(tmp_path: Path, capsys) -> None:
    """
    GIVEN runtime context and parsed log args
    WHEN _run_log_report is called
    THEN transaction list is fetched displayed and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    dal.append_transaction(
        context.workbook,
        dal.TransactionRow(
            transaction_id="T1",
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=constants.TransactionType.SALE.value,
            product_id="P001",
            salesman_id="S001",
            payment_type=constants.PaymentType.CASH.value,
            quantity_change=Decimal("-1"),
            total_revenue=Decimal("2.50"),
            total_cost=Decimal("0"),
            linked_transaction_id=None,
            notes=None,
        ),
    )

    # Act
    exit_code = log._run_log_report(context, argparse.Namespace())

    # Assert
    assert exit_code == 0
    assert "Transaction log:" in capsys.readouterr().out
