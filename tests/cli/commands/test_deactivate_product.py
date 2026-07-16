from pathlib import Path

import argparse
import openpyxl

from caad_erp import bll, constants
from caad_erp.cli.commands import deactivate_product
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    products.append(["P001", "Cookie", 250, True])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append([
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ])
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_deactivate_product_command_returns_command_spec() -> None:
    """
    GIVEN deactivate-product command module
    WHEN register_deactivate_product_command is called
    THEN a CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = deactivate_product.register_deactivate_product_command()

    # Assert
    assert spec.name == "deactivate-product"
    assert spec.is_mutating is True


def test_register_deactivate_product_registrar_configures_required_id() -> None:
    """
    GIVEN subparser factory instance
    WHEN deactivate-product registrar is executed
    THEN parser includes required product-id and command default
    """
    # Arrange
    spec = deactivate_product.register_deactivate_product_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(["deactivate-product", "--product-id", "P001"])

    # Assert
    assert args.command == "deactivate-product"
    assert args.product_id == "P001"


def test_translate_deactivate_product_sets_is_active_false_and_trims_id() -> None:
    """
    GIVEN parsed deactivate-product args with product id value
    WHEN _translate_deactivate_product is called
    THEN ProductCommand is produced with trimmed id and is_active false
    """
    # Arrange
    args = argparse.Namespace(product_id="  P001  ")

    # Act
    command = deactivate_product._translate_deactivate_product(args)

    # Assert
    assert command.product_id == "P001"
    assert command.is_active is False


def test_run_deactivate_product_calls_bll_update_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed deactivate-product args
    WHEN _run_deactivate_product is called
    THEN bll.update_product is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(product_id="P001")

    # Act
    exit_code = deactivate_product._run_deactivate_product(context, args)

    # Assert
    assert exit_code == 0
    updated = bll.get_product(context, "P001")
    assert updated.is_active is False
