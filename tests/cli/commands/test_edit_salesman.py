from pathlib import Path
import argparse
import openpyxl

from caad_erp import bll, constants
from caad_erp.cli.commands import edit_salesman
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    salesmen.append(["S001", "Ana", True])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append(
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ]
    )
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_edit_salesman_command_returns_command_spec() -> None:
    """
    GIVEN edit-salesman command module
    WHEN register_edit_salesman_command is called
    THEN a CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = edit_salesman.register_edit_salesman_command()

    # Assert
    assert spec.name == "edit-salesman"


def test_register_edit_salesman_registrar_configures_arguments() -> None:
    """
    GIVEN subparser factory instance
    WHEN edit-salesman registrar is executed
    THEN parser includes expected salesman arguments and command default
    """
    # Arrange
    spec = edit_salesman.register_edit_salesman_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(
        [
            "edit-salesman",
            "--salesman-id",
            "S001",
            "--salesman-name",
            "John",
            "--salesman-is-active",
            "False",
        ]
    )

    # Assert
    assert args.command == "edit-salesman"
    assert args.salesman_id == "S001"
    assert args.salesman_name == "John"
    assert args.salesman_is_active is False


def test_translate_edit_salesman_sets_fields_and_trims_id() -> None:
    """
    GIVEN parsed edit-salesman args with salesman values
    WHEN _translate_edit_salesman is called
    THEN SalesmanCommand is produced with trimmed id and updated fields
    """
    # Arrange
    args = argparse.Namespace(
        salesman_id="  S001  ", salesman_name="John", salesman_is_active=False
    )

    # Act
    command = edit_salesman._translate_edit_salesman(args)

    # Assert
    assert command.salesman_id == "S001"
    assert command.salesman_name == "John"
    assert command.is_active is False


def test_run_edit_salesman_calls_bll_update_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed edit-salesman args
    WHEN _run_edit_salesman is called
    THEN bll.update_salesman is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        salesman_id="S001", salesman_name="John", salesman_is_active=False
    )

    # Act
    exit_code = edit_salesman._run_edit_salesman(context, args)

    # Assert
    assert exit_code == 0
    updated = bll.get_salesman(context, "S001")
    assert updated.salesman_name == "John"
    assert updated.is_active is False
