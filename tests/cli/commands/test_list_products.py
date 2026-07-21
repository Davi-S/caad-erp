from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants, dal
from caad_erp.cli.commands import list_products
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    products.append(["P001", "Cookie", 250, True])
    products.append(["P002", "Soda", 300, False])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append([
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ])
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_list_products_command_returns_non_mutating_command_spec() -> None:
    """
    GIVEN list-products command module
    WHEN register_list_products_command is called
    THEN a non-mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = list_products.register_list_products_command()

    # Assert
    assert spec.name == "list-products"
    assert spec.is_mutating is False


@pytest.mark.parametrize("include_inactive", [False, True])
def test_run_list_products_report_calls_bll_and_returns_zero(include_inactive, tmp_path: Path, capsys) -> None:
    """
    GIVEN runtime context and parsed list-products args with all flag mode
    WHEN _run_list_products_report is called
    THEN products are listed with include_inactive forwarded and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(all=include_inactive)

    # Act
    exit_code = list_products._run_list_products_report(context, args)

    # Assert
    assert exit_code == 0
    output = capsys.readouterr().out
    assert "P001" in output
    if include_inactive:
        assert "P002" in output
