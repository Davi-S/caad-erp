from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants
from caad_erp.cli.commands import add_salesman
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append(
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ]
    )
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_add_salesman_command_returns_command_spec() -> None:
    """
    GIVEN add-salesman command module
    WHEN register_add_salesman_command is called
    THEN a CommandSpec is returned with expected name help and mutating flag
    """
    # Arrange / Act
    spec = add_salesman.register_add_salesman_command()

    # Assert
    assert spec.name == "add-salesman"
    assert spec.is_mutating is True


def test_register_add_salesman_registrar_configures_parser_arguments() -> None:
    """
    GIVEN subparser factory instance
    WHEN add-salesman registrar is executed
    THEN parser includes required id name and optional inactive arguments
    """
    # Arrange
    spec = add_salesman.register_add_salesman_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(
        [
            "add-salesman",
            "--salesman-id",
            "S001",
            "--salesman-name",
            "Ana",
        ]
    )

    # Assert
    assert args.command == "add-salesman"
    assert args.salesman_id == "S001"
    assert args.salesman_name == "Ana"
    assert args.inactive is False


@pytest.mark.parametrize(
    "inactive_flag, expected_is_active", [(False, True), (True, False)]
)
def test_translate_add_salesman_maps_args_to_salesman_command(
    inactive_flag,
    expected_is_active,
) -> None:
    """
    GIVEN parsed add-salesman args
    WHEN _translate_add_salesman is called
    THEN SalesmanCommand fields are mapped and inactive flag is inverted
    """
    # Arrange
    args = argparse.Namespace(
        salesman_id="S001",
        salesman_name="Ana",
        inactive=inactive_flag,
    )

    # Act
    command = add_salesman._translate_add_salesman(args)

    # Assert
    assert command.salesman_id == "S001"
    assert command.salesman_name == "Ana"
    assert command.is_active is expected_is_active


def test_run_add_salesman_calls_bll_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed add-salesman args
    WHEN _run_add_salesman is called
    THEN bll.add_salesman is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        salesman_id="S001",
        salesman_name="Ana",
        inactive=False,
    )

    # Act
    exit_code = add_salesman._run_add_salesman(context, args)

    # Assert
    assert exit_code == 0
    created = bll.get_salesman(context, "S001")
    assert created.salesman_name == "Ana"
