from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants
from caad_erp.cli.commands import restock
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    products.append(["P001", "Cookie", 250, True])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    salesmen.append(["S001", "Ana", True])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append([
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ])
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_restock_command_returns_command_spec() -> None:
    """
    GIVEN restock command module
    WHEN register_restock_command is called
    THEN a mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = restock.register_restock_command()

    # Assert
    assert spec.name == "restock"
    assert spec.is_mutating is True


def test_register_restock_registrar_configures_required_arguments() -> None:
    """
    GIVEN subparser factory instance
    WHEN restock registrar is executed
    THEN parser includes required product quantity cost salesman and optional notes
    """
    # Arrange
    spec = restock.register_restock_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args([
        "restock",
        "--product-id",
        "P001",
        "--quantity",
        "3",
        "--total-cost",
        "4.50",
        "--salesman-id",
        "S001",
    ])

    # Assert
    assert args.command == "restock"
    assert args.total_cost == "4.50"


def test_run_restock_calls_bll_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed restock args
    WHEN _run_restock is called
    THEN bll.record_restock is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        product_id="P001",
        quantity="3",
        total_cost=450,
        salesman_id="S001",
        notes=None,
    )

    # Act
    exit_code = restock._run_restock(context, args)

    # Assert
    assert exit_code == 0
    rows = bll.list_transactions(context)
    assert len(rows) == 1
    assert rows[0].transaction_type == constants.TransactionType.RESTOCK.value
