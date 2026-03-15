from decimal import Decimal
from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants, dal
from caad_erp.cli.commands import profit
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append([
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ])
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
        default_salesman_id="S001",
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_profit_command_returns_non_mutating_command_spec() -> None:
    """
    GIVEN profit command module
    WHEN register_profit_command is called
    THEN a non-mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = profit.register_profit_command()

    # Assert
    assert spec.name == "profit"
    assert spec.is_mutating is False


def test_register_profit_registrar_configures_command_default() -> None:
    """
    GIVEN subparser factory instance
    WHEN profit registrar is executed
    THEN parser sets command default without additional required args
    """
    # Arrange
    spec = profit.register_profit_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(["profit"])

    # Assert
    assert args.command == "profit"


@pytest.mark.parametrize(
    "summary_payload",
    [
        {},
        {"total_revenue": Decimal("10.00"), "total_cost": Decimal("-4.00")},
        {
            "total_revenue": Decimal("10.00"),
            "total_cost": Decimal("-4.00"),
            "profit": Decimal("9.00"),
        },
    ],
)
def test_display_profit_summary_handles_defaults_and_prints_values(summary_payload, capsys) -> None:
    """
    GIVEN summary mapping with optional missing keys
    WHEN _display_profit_summary is called
    THEN output includes revenue cost and computed or provided profit values
    """
    # Arrange / Act
    profit._display_profit_summary(summary_payload)

    # Assert
    output = capsys.readouterr().out
    assert "Total revenue" in output
    assert "Total cost" in output
    assert "Profit" in output


def test_run_profit_report_calls_bll_and_returns_zero(tmp_path: Path, capsys) -> None:
    """
    GIVEN runtime context and parsed profit args
    WHEN _run_profit_report is called
    THEN profit summary is computed displayed and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    dal.append_transaction(
        context.workbook,
        dal.TransactionRow(
            transaction_id="T1",
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=constants.TransactionType.SALE.value,
            product_id="P001",
            salesman_id="S001",
            payment_type=constants.PaymentType.CASH.value,
            quantity_change=Decimal("-1"),
            total_revenue=Decimal("5"),
            total_cost=Decimal("0"),
            linked_transaction_id=None,
            notes=None,
        ),
    )

    # Act
    exit_code = profit._run_profit_report(context, argparse.Namespace())

    # Assert
    assert exit_code == 0
    assert "Profit summary" in capsys.readouterr().out
