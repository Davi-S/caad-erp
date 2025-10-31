"""Unit tests documenting the expected behavior of the data access layer."""

from __future__ import annotations

import configparser
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook as OpenpyxlWorkbook
import pytest

from caad_erp import constants, data_manager  # noqa: E402


def test_find_config_file_respects_explicit_path(config_file: Path):
    """Supplying an explicit path should be treated as the winning answer."""

    result = data_manager.find_config_file(config_file)
    assert result == config_file


def test_find_config_file_discovers_in_cwd(tmp_path, monkeypatch):
    """Auto-discovery should locate config.ini in the working directory tree."""

    config_dir = tmp_path / "nested"
    config_dir.mkdir(parents=True)
    config_file = config_dir / "config.ini"
    config_file.write_text("[System]\nDataFile=master_workbook.xlsx")
    monkeypatch.chdir(config_dir)

    result = data_manager.find_config_file()
    assert result == config_file


def test_find_config_file_raises_when_missing(tmp_path, monkeypatch):
    """Absent configuration should surface a clear FileNotFoundError."""

    monkeypatch.chdir(tmp_path)
    with pytest.raises(FileNotFoundError):
        data_manager.find_config_file()


def test_read_config_loads_sections(config_file: Path):
    """read_config should return a populated ConfigParser."""

    parser = data_manager.read_config(config_file)
    assert parser.get("System", "LoungeName") == "Test Lounge"
    assert parser.get("Defaults", "DefaultSalesman") == "S-DEFAULT"


def test_read_config_missing_file_raises(tmp_path):
    """Missing files should propagate a FileNotFoundError."""

    with pytest.raises(FileNotFoundError):
        data_manager.read_config(tmp_path / "not_there.ini")


def test_parse_settings_resolves_relative_paths(config_factory):
    """Relative DataFile entries should be anchored to the config location."""

    parser = configparser.ConfigParser()
    bundle = config_factory(make_relative=True)
    parser.read(bundle.config_path)
    settings = data_manager.parse_settings(parser, base_path=bundle.config_path.parent)
    assert settings.data_file == (bundle.config_path.parent / bundle.workbook_path.name).resolve()
    assert settings.default_salesman_id == "S-DEFAULT"


def test_parse_settings_requires_expected_sections(tmp_path):
    """Missing keys should result in a descriptive KeyError."""

    parser = configparser.ConfigParser()
    parser.read_string("[Other]\nvalue=1")
    with pytest.raises(KeyError):
        data_manager.parse_settings(parser, base_path=tmp_path)


def test_open_workbook_returns_openpyxl_instance(master_workbook_path):
    """open_workbook should hand back a loaded Workbook object."""

    workbook = data_manager.open_workbook(master_workbook_path)
    assert isinstance(workbook, OpenpyxlWorkbook)


def test_open_workbook_missing_file_raises(tmp_path):
    """Missing workbook files should yield FileNotFoundError."""

    with pytest.raises(FileNotFoundError):
        data_manager.open_workbook(tmp_path / "missing.xlsx")


def test_save_workbook_persists_changes(master_workbook_path):
    """save_workbook should persist changes to the provided destination path."""

    workbook = data_manager.open_workbook(master_workbook_path)
    sheet = workbook[constants.SheetName.PRODUCTS.value]
    sheet.append(["P100", "Chips", "2.50", True])
    data_manager.save_workbook(workbook, master_workbook_path)

    reloaded = data_manager.open_workbook(master_workbook_path)
    values = list(reloaded[constants.SheetName.PRODUCTS.value].iter_rows(min_row=2, values_only=True))
    assert values == [("P100", "Chips", "2.50", True)]


def test_save_workbook_with_destination_creates_copy(master_workbook_path, tmp_path):
    """Providing a destination should create a new file independent of the source."""

    workbook = data_manager.open_workbook(master_workbook_path)
    sheet = workbook[constants.SheetName.SALESMEN.value]
    sheet.append(["S2", "Jordan", True])
    copy_path = tmp_path / "copy.xlsx"
    data_manager.save_workbook(workbook, destination=copy_path)

    copy = openpyxl.load_workbook(copy_path)
    rows = list(copy[constants.SheetName.SALESMEN.value].iter_rows(min_row=2, values_only=True))
    assert ("S2", "Jordan", True) in rows


def test_refresh_workbook_returns_new_instance(master_workbook_path):
    """refresh_workbook should return a freshly loaded workbook from disk."""

    original = data_manager.open_workbook(master_workbook_path)
    sheet = original[constants.SheetName.PRODUCTS.value]
    sheet.append(["P200", "Bars", "4.00", True])
    data_manager.save_workbook(original, master_workbook_path)

    refreshed = data_manager.refresh_workbook(master_workbook_path)
    assert refreshed is not original
    values = list(refreshed[constants.SheetName.PRODUCTS.value].iter_rows(min_row=2, values_only=True))
    assert ("P200", "Bars", "4.00", True) in values


def test_iter_products_yields_product_rows(master_workbook_path):
    """iter_products should yield ProductRow instances for worksheet data."""

    workbook = data_manager.open_workbook(master_workbook_path)
    products = workbook[constants.SheetName.PRODUCTS.value]
    products.append(["P300", "Soda", "5.00", True])
    data_manager.save_workbook(workbook, master_workbook_path)

    refreshed = data_manager.open_workbook(master_workbook_path)
    rows = list(data_manager.iter_products(refreshed))
    assert rows == [
        data_manager.ProductRow(
            product_id="P300",
            product_name="Soda",
            sell_price=Decimal("5.00"),
            is_active=True,
        )
    ]


def test_iter_salesmen_yields_salesman_rows(master_workbook_path):
    """iter_salesmen should expose SalesmanRow objects."""

    workbook = data_manager.open_workbook(master_workbook_path)
    salesmen = workbook[constants.SheetName.SALESMEN.value]
    salesmen.append(["S2", "Morgan", True])
    data_manager.save_workbook(workbook, master_workbook_path)

    refreshed = data_manager.open_workbook(master_workbook_path)
    rows = list(data_manager.iter_salesmen(refreshed))
    assert rows[0].salesman_id == "S-DEFAULT"
    assert any(row.salesman_id == "S2" for row in rows)


def test_iter_transactions_yields_transaction_rows(master_workbook_path):
    """iter_transactions should convert worksheet rows to TransactionRow objects."""

    workbook = data_manager.open_workbook(master_workbook_path)
    transactions = workbook[constants.SheetName.TRANSACTION_LOG.value]
    transactions.append(
        [
            "T1",
            "2025-10-29T20:00:00",
            constants.TransactionType.SALE.value,
            "P300",
            "S-DEFAULT",
            constants.PaymentType.CASH.value,
            "-1",
            "5.00",
            "0.00",
            None,
            "Notes",
        ]
    )
    data_manager.save_workbook(workbook, master_workbook_path)

    refreshed = data_manager.open_workbook(master_workbook_path)
    rows = list(data_manager.iter_transactions(refreshed))
    assert rows[0].transaction_id == "T1"
    assert rows[0].quantity_change == Decimal("-1")


def test_append_product_adds_row(master_workbook_path):
    """append_product should add the provided row to the worksheet."""

    workbook = data_manager.open_workbook(master_workbook_path)
    record = data_manager.ProductRow(
        product_id="P400",
        product_name="Juice",
        sell_price=Decimal("6.00"),
        is_active=False,
    )
    data_manager.append_product(workbook, record)
    data_manager.save_workbook(workbook, master_workbook_path)

    refreshed = data_manager.open_workbook(master_workbook_path)
    rows = list(data_manager.iter_products(refreshed))
    assert any(row.product_id == "P400" and row.is_active is False for row in rows)


def test_append_salesman_adds_row(master_workbook_path):
    """append_salesman should append the record to the Salesmen sheet."""

    workbook = data_manager.open_workbook(master_workbook_path)
    record = data_manager.SalesmanRow(
        salesman_id="S9",
        salesman_name="Jamie",
        is_active=False,
    )
    data_manager.append_salesman(workbook, record)
    data_manager.save_workbook(workbook, master_workbook_path)

    refreshed = data_manager.open_workbook(master_workbook_path)
    rows = list(data_manager.iter_salesmen(refreshed))
    assert any(row.salesman_id == "S9" for row in rows)


def test_append_transaction_adds_row(master_workbook_path):
    """append_transaction should add a ledger row to TransactionLog."""

    workbook = data_manager.open_workbook(master_workbook_path)
    record = data_manager.TransactionRow(
        transaction_id="T2",
        timestamp_iso="2025-10-29T21:00:00",
        transaction_type=constants.TransactionType.RESTOCK.value,
        product_id="P400",
        salesman_id=None,
        payment_type=constants.PaymentType.CASH.value,
        quantity_change=Decimal("10"),
        total_revenue=Decimal("0.00"),
        total_cost=Decimal("-20.00"),
        linked_transaction_id=None,
        notes="Restock",
    )
    data_manager.append_transaction(workbook, record)
    data_manager.save_workbook(workbook, master_workbook_path)

    refreshed = data_manager.open_workbook(master_workbook_path)
    rows = list(data_manager.iter_transactions(refreshed))
    assert any(row.transaction_id == "T2" for row in rows)


def test_update_product_modifies_existing_row(master_workbook_path):
    """update_product should mutate values for the matching ProductID."""

    workbook = data_manager.open_workbook(master_workbook_path)
    sheet = workbook[constants.SheetName.PRODUCTS.value]
    sheet.append(["P500", "Old", "1.00", True])
    data_manager.save_workbook(workbook, master_workbook_path)

    reloaded = data_manager.open_workbook(master_workbook_path)
    data_manager.update_product(
        reloaded,
        "P500",
        field_values={"ProductName": "New", "SellPrice": Decimal("2.00")},
    )
    data_manager.save_workbook(reloaded, master_workbook_path)

    final = data_manager.open_workbook(master_workbook_path)
    rows = list(data_manager.iter_products(final))
    assert any(row.product_id == "P500" and row.product_name == "New" for row in rows)


def test_update_product_missing_raises(master_workbook_path):
    """Updating a nonexistent product should surface a KeyError."""

    workbook = data_manager.open_workbook(master_workbook_path)
    with pytest.raises(KeyError):
        data_manager.update_product(workbook, "NOPE", field_values={"ProductName": "X"})


def test_update_salesman_modifies_existing_row(master_workbook_path):
    """update_salesman should support partial field merges."""

    workbook = data_manager.open_workbook(master_workbook_path)
    sheet = workbook[constants.SheetName.SALESMEN.value]
    sheet.append(["S500", "Taylor", True])
    data_manager.save_workbook(workbook, master_workbook_path)

    reloaded = data_manager.open_workbook(master_workbook_path)
    data_manager.update_salesman(reloaded, "S500", field_values={"IsActive": False})
    data_manager.save_workbook(reloaded, master_workbook_path)

    final = data_manager.open_workbook(master_workbook_path)
    rows = list(data_manager.iter_salesmen(final))
    assert any(row.salesman_id == "S500" and row.is_active is False for row in rows)


def test_update_salesman_missing_raises(master_workbook_path):
    """Attempting to update a missing salesman should raise KeyError."""

    workbook = data_manager.open_workbook(master_workbook_path)
    with pytest.raises(KeyError):
        data_manager.update_salesman(workbook, "MISSING", field_values={"IsActive": False})


def test_locate_row_returns_row_index(master_workbook_path):
    """locate_row should return the worksheet index of the matching key."""

    workbook = data_manager.open_workbook(master_workbook_path)
    sheet = workbook[constants.SheetName.PRODUCTS.value]
    sheet.append(["P600", "Snack", "2.00", True])
    data_manager.save_workbook(workbook, master_workbook_path)

    reloaded = data_manager.open_workbook(master_workbook_path)
    row_index = data_manager.locate_row(
        reloaded,
        constants.SheetName.PRODUCTS.value,
        "ProductID",
        "P600",
    )
    assert row_index == 2


def test_locate_row_returns_none_when_missing(master_workbook_path):
    """locate_row should return None if the key is not present."""

    workbook = data_manager.open_workbook(master_workbook_path)
    assert (
        data_manager.locate_row(
            workbook,
            constants.SheetName.PRODUCTS.value,
            "ProductID",
            "NOPE",
        )
        is None
    )


def test_serialize_product_preserves_order():
    """serialize_product should follow the column ordering defined by setup."""

    record = data_manager.ProductRow("P1", "Name", Decimal("1.25"), True)
    assert data_manager.serialize_product(record) == ["P1", "Name", Decimal("1.25"), True]


def test_serialize_salesman_preserves_order():
    """serialize_salesman should output values in sheet order."""

    record = data_manager.SalesmanRow("S1", "Sam", False)
    assert data_manager.serialize_salesman(record) == ["S1", "Sam", False]


def test_serialize_transaction_preserves_order():
    """serialize_transaction should output the TransactionLog column order."""

    record = data_manager.TransactionRow(
        transaction_id="T3",
        timestamp_iso="2025-10-29T22:00:00",
        transaction_type=constants.TransactionType.SALE.value,
        product_id="P1",
        salesman_id="S1",
        payment_type=constants.PaymentType.CASH.value,
        quantity_change=Decimal("-1"),
        total_revenue=Decimal("3.00"),
        total_cost=Decimal("0.00"),
        linked_transaction_id=None,
        notes="Note",
    )
    assert data_manager.serialize_transaction(record) == [
        "T3",
        "2025-10-29T22:00:00",
        constants.TransactionType.SALE.value,
        "P1",
        "S1",
        constants.PaymentType.CASH.value,
        Decimal("-1"),
        Decimal("3.00"),
        Decimal("0.00"),
        None,
        "Note",
    ]


def test_deserialize_product_constructs_dataclass():
    """deserialize_product should coerce worksheet values into ProductRow."""

    record = data_manager.deserialize_product(["P9", "Bar", "2.75", True])
    assert record.product_id == "P9"
    assert record.sell_price == Decimal("2.75")


def test_deserialize_salesman_constructs_dataclass():
    """deserialize_salesman should create a SalesmanRow from raw values."""

    record = data_manager.deserialize_salesman(["S9", "Alex", False])
    assert record.salesman_name == "Alex"
    assert record.is_active is False


def test_deserialize_transaction_constructs_dataclass():
    """deserialize_transaction should parse decimals and optional fields."""

    record = data_manager.deserialize_transaction(
        [
            "T9",
            "2025-10-29T23:00:00",
            constants.TransactionType.WRITE_OFF.value,
            "P9",
            "S9",
            constants.PaymentType.CASH.value,
            "-2",
            "0.00",
            "0.00",
            "T8",
            None,
        ]
    )
    assert record.transaction_id == "T9"
    assert record.quantity_change == Decimal("-2")
    assert record.linked_transaction_id == "T8"
