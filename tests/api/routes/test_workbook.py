import io
import openpyxl
from fastapi.testclient import TestClient


def test_download_workbook_report_returns_valid_excel_file(api_client: TestClient) -> None:
    """
    GIVEN an initialized API runtime context
    WHEN GET /reports/workbook is called
    THEN endpoint returns 200, excel content type header, and a valid openpyxl readable workbook binary
    """
    response = api_client.get("/reports/workbook")

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="master_workbook.xlsx"' in response.headers["content-disposition"]

    # Verify content can be loaded by openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Products" in wb.sheetnames
    assert "Salesmen" in wb.sheetnames
    assert "TransactionLog" in wb.sheetnames


def test_download_workbook_report_contains_persisted_mutations(
    api_client: TestClient,
) -> None:
    """
    GIVEN products created via API mutation endpoints
    WHEN GET /reports/workbook is requested
    THEN the returned workbook file contains the newly added records in its sheets
    """
    # Create product via API
    create_res = api_client.post(
        "/products",
        json={
            "product_id": "DL-P100",
            "product_name": "Downloaded Product",
            "sell_price": 1200,
            "is_active": True,
        },
    )
    assert create_res.status_code == 201

    # Download workbook
    response = api_client.get("/reports/workbook")
    assert response.status_code == 200

    # Read sheet content
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    products_sheet = wb["Products"]
    product_ids = [cell.value for cell in products_sheet["A"][1:]]
    assert "DL-P100" in product_ids


def test_download_workbook_report_returns_503_when_runtime_uninitialized(
    api_client_without_runtime: TestClient,
) -> None:
    """
    GIVEN runtime context dependency is not initialized
    WHEN GET /reports/workbook is requested
    THEN response status is 503 Service Unavailable
    """
    response = api_client_without_runtime.get("/reports/workbook")

    assert response.status_code == 503
    assert response.json()["error_type"] == "RuntimeError"
