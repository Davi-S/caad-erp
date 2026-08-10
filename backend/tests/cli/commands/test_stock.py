from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants, dal, exceptions
from caad_erp.cli.commands import stock
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append(
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ]
    )
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_stock_command_returns_non_mutating_command_spec() -> None:
    """
    GIVEN stock command module
    WHEN register_stock_command is called
    THEN a non-mutating CommandSpec is returned with expected metadata
    """
    # Arrange / Act
    spec = stock.register_stock_command()

    # Assert
    assert spec.name == "stock"
    assert spec.is_mutating is False


def test_register_stock_registrar_configures_command_default() -> None:
    """
    GIVEN subparser factory instance
    WHEN stock registrar is executed
    THEN parser sets command default without additional required args
    """
    # Arrange
    spec = stock.register_stock_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(["stock"])

    # Assert
    assert args.command == "stock"


def test_display_inventory_report_prints_empty_message_for_no_data(capsys) -> None:
    """
    GIVEN empty inventory mapping
    WHEN _display_inventory_report is called
    THEN empty-state message is printed
    """
    # Arrange / Act
    stock._display_inventory_report({})

    # Assert
    assert "No stock data available." in capsys.readouterr().out


def test_display_inventory_report_prints_sorted_rows_for_inventory(capsys) -> None:
    """
    GIVEN inventory mapping containing multiple product quantities
    WHEN _display_inventory_report is called
    THEN table header and sorted rows are printed
    """
    # Arrange
    inventory = {"P002": 5, "P001": 3}

    # Act
    stock._display_inventory_report(inventory)

    # Assert
    output = capsys.readouterr().out
    assert "Product ID" in output
    assert output.find("P001") < output.find("P002")


def test_run_stock_report_calls_bll_and_returns_zero(tmp_path: Path, capsys) -> None:
    """
    GIVEN runtime context and parsed stock args
    WHEN _run_stock_report is called
    THEN inventory is computed displayed and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    dal.append_transaction(
        context.workbook,
        dal.TransactionRow(
            transaction_id="T1",
            timestamp_iso="2026-03-15T10:00:00+00:00",
            transaction_type=constants.TransactionType.RESTOCK.value,
            product_id="P001",
            salesman_id="S001",
            payment_type=None,
            quantity_change=5,
            total_revenue=0,
            total_cost=-500,
            linked_transaction_id=None,
            notes=None,
        ),
    )

    # Act
    exit_code = stock._run_stock_report(context, argparse.Namespace())

    # Assert
    assert exit_code == 0
    assert "P001" in capsys.readouterr().out


def test_run_stock_report_returns_nonzero_exit_code_on_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """
    GIVEN runtime context when bll raises BusinessRuleViolation
    WHEN _run_stock_report is called
    THEN non-zero exit code 2 is returned
    """
    # Arrange
    context = _make_context(tmp_path)

    def _mock_calculate_inventory(*args, **kwargs):
        raise exceptions.BusinessRuleViolation("Stock calculation error")

    monkeypatch.setattr(bll, "calculate_inventory", _mock_calculate_inventory)

    # Act
    exit_code = stock._run_stock_report(context, argparse.Namespace())

    # Assert
    assert exit_code == 2
