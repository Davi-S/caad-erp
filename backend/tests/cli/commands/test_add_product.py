from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants, exceptions
from caad_erp.cli.commands import add_product
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append(
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ]
    )
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_add_product_command_returns_command_spec() -> None:
    """
    GIVEN add-product command module
    WHEN register_add_product_command is called
    THEN a CommandSpec is returned with expected name help and mutating flag
    """
    # Arrange / Act
    spec = add_product.register_add_product_command()

    # Assert
    assert spec.name == "add-product"
    assert "Register a new product" in spec.help_text
    assert spec.is_mutating is True


def test_register_add_product_registrar_configures_parser_arguments() -> None:
    """
    GIVEN subparser factory instance
    WHEN add-product registrar is executed
    THEN parser includes required id name price and optional inactive arguments
    """
    # Arrange
    spec = add_product.register_add_product_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(
        [
            "add-product",
            "--product-id",
            "P001",
            "--product-name",
            "Cookie",
            "--sell-price",
            "2.50",
        ]
    )

    # Assert
    assert args.command == "add-product"
    assert args.product_id == "P001"
    assert args.product_name == "Cookie"
    assert args.sell_price == "2.50"
    assert args.inactive is False


@pytest.mark.parametrize(
    "inactive_flag, expected_is_active", [(False, True), (True, False)]
)
def test_translate_add_product_maps_args_to_product_command(
    inactive_flag,
    expected_is_active,
) -> None:
    """
    GIVEN parsed add-product args
    WHEN _translate_add_product is called
    THEN ProductCommand fields are mapped and inactive flag is inverted
    """
    # Arrange
    args = argparse.Namespace(
        product_id="P001",
        product_name="Cookie",
        sell_price=250,
        inactive=inactive_flag,
    )

    # Act
    command = add_product._translate_add_product(args)

    # Assert
    assert command.product_id == "P001"
    assert command.product_name == "Cookie"
    assert command.sell_price == 250
    assert command.is_active is expected_is_active


def test_run_add_product_calls_bll_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed add-product args
    WHEN _run_add_product is called
    THEN bll.add_product is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        product_id="P001",
        product_name="Cookie",
        sell_price=250,
        inactive=False,
    )

    # Act
    exit_code = add_product._run_add_product(context, args)

    # Assert
    assert exit_code == 0
    created = bll.get_product(context, "P001")
    assert created.product_name == "Cookie"
    assert created.sell_price == 250


def test_run_add_product_returns_nonzero_exit_code_on_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """
    GIVEN runtime context and add-product args when bll raises BusinessRuleViolation
    WHEN _run_add_product is called
    THEN non-zero exit code 2 is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        product_id="P001",
        product_name="Cookie",
        sell_price=250,
        inactive=False,
    )

    def _mock_add_product(*args, **kwargs):
        raise exceptions.BusinessRuleViolation("Duplicate product")

    monkeypatch.setattr(bll, "add_product", _mock_add_product)

    # Act
    exit_code = add_product._run_add_product(context, args)

    # Assert
    assert exit_code == 2
