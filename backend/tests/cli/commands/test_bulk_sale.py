from pathlib import Path

import argparse
import openpyxl
import pytest

from caad_erp import bll, constants, exceptions
from caad_erp.cli.commands import bulk_sale
from caad_erp.settings import AppSettings


def _make_context(tmp_path: Path) -> bll.RuntimeContext:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    products = wb.create_sheet(constants.SheetName.PRODUCTS.value)
    products.append(["ProductID", "ProductName", "SellPrice", "IsActive"])
    products.append(["P001", "Cookie", 250, True])
    products.append(["P002", "Soda", 300, True])
    salesmen = wb.create_sheet(constants.SheetName.SALESMEN.value)
    salesmen.append(["SalesmanID", "SalesmanName", "IsActive"])
    salesmen.append(["S001", "Ana", True])
    tx = wb.create_sheet(constants.SheetName.TRANSACTION_LOG.value)
    tx.append(
        [
            "TransactionID",
            "Timestamp",
            "TransactionType",
            "ProductID",
            "SalesmanID",
            "PaymentType",
            "QuantityChange",
            "TotalRevenue",
            "TotalCost",
            "LinkedTransactionID",
            "Notes",
        ]
    )
    settings = AppSettings(
        data_file=tmp_path / "data.xlsx",
        lounge_name="Test",
        schema_version=constants.EXPECTED_SCHEMA_VERSION,
    )
    return bll.RuntimeContext(settings=settings, workbook=wb)


def test_register_bulk_sale_command_returns_command_spec() -> None:
    """
    GIVEN bulk-sale command module
    WHEN register_bulk_sale_command is called
    THEN a mutating CommandSpec is returned with name 'bulk-sale'
    """
    # Arrange / Act
    spec = bulk_sale.register_bulk_sale_command()

    # Assert
    assert spec.name == "bulk-sale"
    assert spec.is_mutating is True


def test_register_bulk_sale_registrar_configures_required_arguments() -> None:
    """
    GIVEN subparser factory instance
    WHEN bulk-sale registrar is executed
    THEN parser configures salesman-id payment-type and repeated item flags
    """
    # Arrange
    spec = bulk_sale.register_bulk_sale_command()
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command")
    spec.register(subparsers)

    # Act
    args = parser.parse_args(
        [
            "bulk-sale",
            "-s",
            "S001",
            "-p",
            "Cash",
            "-n",
            "Bulk order",
            "-i",
            "P001",
            "2",
            "500",
            "-i",
            "P002",
            "1",
            "300",
        ]
    )

    # Assert
    assert args.command == "bulk-sale"
    assert args.salesman_id == "S001"
    assert args.payment_type == "Cash"
    assert args.notes == "Bulk order"
    assert args.item == [["P001", "2", "500"], ["P002", "1", "300"]]


def test_translate_bulk_sale_maps_args_to_list_of_sale_commands() -> None:
    """
    GIVEN parsed bulk-sale args
    WHEN _translate_bulk_sale is called
    THEN list of SaleCommand objects is constructed
    """
    # Arrange
    args = argparse.Namespace(
        salesman_id="S001",
        payment_type="Cash",
        notes="Shared note",
        item=[["P001", "2", "500"], ["P002", "1", "300"]],
    )

    # Act
    commands = bulk_sale._translate_bulk_sale(args)

    # Assert
    assert len(commands) == 2
    assert commands[0].product_id == "P001"
    assert commands[0].salesman_id == "S001"
    assert commands[0].quantity == 2
    assert commands[0].total_revenue == 500
    assert commands[0].payment_type == constants.PaymentType.CASH
    assert commands[0].notes == "Shared note"

    assert commands[1].product_id == "P002"
    assert commands[1].quantity == 1
    assert commands[1].total_revenue == 300


def test_run_bulk_sale_calls_bll_and_returns_zero(tmp_path: Path) -> None:
    """
    GIVEN runtime context and parsed bulk-sale args
    WHEN _run_bulk_sale is called
    THEN bll.record_bulk_sale is invoked and zero is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        salesman_id="S001",
        payment_type="Cash",
        notes="Bulk sale notes",
        item=[["P001", "2", "500"], ["P002", "1", "300"]],
    )

    # Act
    exit_code = bulk_sale._run_bulk_sale(context, args)

    # Assert
    assert exit_code == 0
    rows = bll.list_transactions(context)
    assert len(rows) == 2


def test_run_bulk_sale_returns_nonzero_exit_code_on_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """
    GIVEN runtime context and bulk-sale args when bll raises BusinessRuleViolation
    WHEN _run_bulk_sale is called
    THEN non-zero exit code 2 is returned
    """
    # Arrange
    context = _make_context(tmp_path)
    args = argparse.Namespace(
        salesman_id="S001",
        payment_type="Cash",
        notes=None,
        item=[["P999", "1", "100"]],
    )

    def _mock_record_bulk_sale(*args, **kwargs):
        raise exceptions.MissingReferenceError("Product not found")

    monkeypatch.setattr(bll, "record_bulk_sale", _mock_record_bulk_sale)

    # Act
    exit_code = bulk_sale._run_bulk_sale(context, args)

    # Assert
    assert exit_code == 2
