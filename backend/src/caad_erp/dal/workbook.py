"""Workbook lifecycle utilities for the CAAD ERP Excel datastore.

Functions in this module abstract the ``openpyxl`` mechanics for opening,
saving, refreshing, and searching worksheets. Centralizing these operations
keeps higher layers focused on domain logic instead of file-system and
spreadsheet plumbing concerns.
"""

import logging
import typing as t
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook

logger = logging.getLogger(__name__)


def open_workbook(data_file: Path) -> Workbook:
    """Open the master Excel workbook and return a live ``openpyxl`` workbook.

    The provided path is expanded (supporting ``~``), resolved to its absolute
    form, and verified for existence. Successful calls load the workbook via
    :func:`openpyxl.load_workbook`. Callers must retain the resolved path and
    provide it back to :func:`save_workbook` when persisting changes.

    Args:
        data_file (Path): Filesystem path to the ``master_workbook.xlsx`` file.

    Returns:
        Workbook: ``openpyxl`` workbook instance backed by the provided file.

    Raises:
        FileNotFoundError: If ``data_file`` does not exist after expansion and
            resolution.
    """

    data_file = Path(data_file).expanduser().resolve()
    logger.debug("Opening workbook '%s'", data_file)
    if not data_file.exists():
        logger.error("Workbook not found at '%s'", data_file)
        raise FileNotFoundError(f"Workbook not found: {data_file}")

    wb = openpyxl.load_workbook(data_file)
    logger.debug("Opened workbook '%s'", data_file)
    return wb


def save_workbook(workbook: Workbook, destination: Path) -> None:
    """Persist the workbook to disk at an explicitly provided destination.

    The destination path is expanded (supporting ``~``) and resolved to an
    absolute location. Parent directories are created on demand to match the
    previous behavior of the data layer always producing the target folder.

    Args:
        workbook (Workbook): Workbook instance to persist.
        destination (Path): Filesystem path that should receive the serialized
            workbook.
    """

    dest = Path(destination).expanduser().resolve()
    dest.parent.mkdir(parents=True, exist_ok=True)
    logger.debug("Saving workbook to '%s'", dest)
    workbook.save(dest)


def locate_row(
    workbook: Workbook, sheet_name: str, key_column: str, key_value: str
) -> t.Optional[int]:
    """Find a row by matching a key value within the specified worksheet.

    The function constructs a mapping from header titles to column indices,
    verifies that ``key_column`` exists, and scans the worksheet for the first
    row whose value equals ``key_value``. The header row itself is not
    considered during matching.

    Args:
        workbook (Workbook): Workbook providing access to ``sheet_name``.
        sheet_name (str): Name of the worksheet to search.
        key_column (str): Header title identifying the column that stores the
            lookup key.
        key_value (str): Value to match within the key column.

    Returns:
        int | None: 1-based Excel row index when a match is found, otherwise
            ``None``.

    Raises:
        KeyError: If ``key_column`` is not present in the worksheet header.
    """

    sheet = workbook[sheet_name]
    # Build header -> column index map
    header_cells = list(sheet[1])
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(header_cells)}
    if key_column not in header_map:
        logger.error("Column '%s' not found in worksheet '%s'", key_column, sheet_name)
        raise KeyError(f"Unknown column: {key_column}")

    key_col_index = header_map[key_column]

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        cell_value = row[key_col_index - 1]
        if cell_value == key_value:
            return row_idx

    return None
