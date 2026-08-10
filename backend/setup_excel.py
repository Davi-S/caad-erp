"""Utility for initializing the CAAD ERP master workbook.

The module doubles as a script (``python setup_excel.py``) and as a library
used by tests or other tooling. Shared helpers keep the workbook bootstrap
logic consistent regardless of the execution path.
"""

from __future__ import annotations

import argparse
import configparser
from dataclasses import dataclass
from pathlib import Path
from typing import Mapping, Sequence
import sys

import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

# Define the schema exactly as specified in the architecture
SHEET_COLUMNS: Mapping[str, Sequence[str]] = {
    "Products": [
        "ProductID",
        "ProductName",
        "SellPrice",
        "IsActive",
    ],
    "Salesmen": [
        "SalesmanID",
        "SalesmanName",
        "IsActive",
    ],
    "TransactionLog": [
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ],
}

CONFIG_FILE = "config.ini"


@dataclass(frozen=True)
class SetupSettings:
    """Type-safe representation of configuration values used during setup."""

    data_file: Path
    default_salesman_id: str


@dataclass(frozen=True)
class DashboardStyles:
    """Centralized formatting tokens used across Dashboard builders."""

    title_font: Font = Font(bold=True, size=14)
    header_font: Font = Font(bold=True, size=11)
    bold_font: Font = Font(bold=True)
    currency_format: str = "$#,##0.00"
    percentage_format: str = "0.0%"


def load_settings(config_path: Path) -> SetupSettings:
    """Read ``config.ini`` and produce :class:`SetupSettings`.

    Relative paths inside the config file are resolved against the config file's
    directory so the behavior matches the previous script.
    """

    if not config_path.exists():
        raise FileNotFoundError(f"Configuration file not found: {config_path}")

    parser = configparser.ConfigParser()
    parser.read(config_path)

    try:
        data_file_raw = parser.get("System", "DataFile")
        default_salesman_id = parser.get("Defaults", "DefaultSalesman")
    except (configparser.NoSectionError, configparser.NoOptionError) as exc:
        raise KeyError(f"Missing required configuration entry: {exc}") from exc

    data_file_path = Path(data_file_raw)
    if not data_file_path.is_absolute():
        data_file_path = (config_path.parent / data_file_path).resolve()

    return SetupSettings(
        data_file=data_file_path,
        default_salesman_id=default_salesman_id,
    )


def _build_kpi_cards(ws: Worksheet, styles: DashboardStyles) -> None:
    """Build Section 1: Executive KPI Summary Cards (A1:B6)."""
    ws["A1"] = "CAAD ERP Executive Summary"
    ws["A1"].font = styles.title_font

    kpis = [
        ("Total Revenue", "=SUM(TransactionLog!H:H)/100", styles.currency_format),
        ("Total Costs / Expenses", "=SUM(TransactionLog!I:I)/100", styles.currency_format),
        ("Net Profit", "=B2+B3", styles.currency_format),
        ("Profit Margin", "=IFERROR((B2+B3)/B2, 0)", styles.percentage_format),
        (
            "Outstanding Debts",
            '=(SUMIF(TransactionLog!F:F, "OnCredit", TransactionLog!H:H) - SUMIF(TransactionLog!C:C, "CREDIT_PAYMENT", TransactionLog!H:H))/100',
            styles.currency_format,
        ),
    ]

    for idx, (label, formula, num_format) in enumerate(kpis, start=2):
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = styles.bold_font
        val_cell = ws.cell(row=idx, column=2, value=formula)
        val_cell.number_format = num_format


def _build_payment_breakdown(ws: Worksheet, styles: DashboardStyles) -> None:
    """Build Section 2: Payment Method Breakdown Table & Pie Chart (D1:H14)."""
    headers = [("D1", "Payment Type"), ("E1", "Revenue"), ("F1", "Share %")]
    for cell_ref, text in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = styles.header_font

    payment_types = ["Cash", "PIX", "OnCredit", "Other"]
    for idx, ptype in enumerate(payment_types, start=2):
        ws.cell(row=idx, column=4, value=ptype)
        rev_cell = ws.cell(
            row=idx,
            column=5,
            value=f'=SUMIFS(TransactionLog!H:H, TransactionLog!F:F, "{ptype}", TransactionLog!C:C, "SALE")/100',
        )
        rev_cell.number_format = styles.currency_format
        share_cell = ws.cell(
            row=idx,
            column=6,
            value=f"=IFERROR(E{idx}/$B$2, 0)",
        )
        share_cell.number_format = styles.percentage_format

    chart_payment = PieChart()
    chart_payment.title = "Revenue Distribution by Payment Method"
    labels_payment = Reference(ws, min_col=4, min_row=2, max_row=5)
    data_payment = Reference(ws, min_col=5, min_row=1, max_row=5)
    chart_payment.add_data(data_payment, titles_from_data=True)
    chart_payment.set_categories(labels_payment)
    chart_payment.width = 14
    chart_payment.height = 7
    ws.add_chart(chart_payment, "H1")


def _build_sales_leaderboard(ws: Worksheet, styles: DashboardStyles) -> None:
    """Build Section 3: Sales Leaderboard Table with Dual Rankings (A16:F26)."""
    headers = [
        ("A16", "Salesman Name"),
        ("B16", "Deals Closed"),
        ("C16", "Deals Rank"),
        ("D16", "Total Revenue"),
        ("E16", "Revenue Rank"),
        ("F16", "% of Total"),
    ]
    for cell_ref, text in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = styles.header_font

    for r in range(2, 12):
        k = 15 + r
        ws.cell(row=k, column=1, value=f'=IF(Salesmen!B{r}="","",Salesmen!B{r})')
        ws.cell(
            row=k,
            column=2,
            value=f'=IF(A{k}="","",COUNTIFS(TransactionLog!E:E, Salesmen!A{r}, TransactionLog!C:C, "SALE"))',
        )
        ws.cell(
            row=k,
            column=3,
            value=f'=IF(A{k}="","",RANK(B{k}, $B$17:$B$26))',
        )
        rev_cell = ws.cell(
            row=k,
            column=4,
            value=f'=IF(A{k}="","",SUMIFS(TransactionLog!H:H, TransactionLog!E:E, Salesmen!A{r}, TransactionLog!C:C, "SALE")/100)',
        )
        rev_cell.number_format = styles.currency_format
        ws.cell(
            row=k,
            column=5,
            value=f'=IF(A{k}="","",RANK(D{k}, $D$17:$D$26))',
        )
        share_cell = ws.cell(
            row=k, column=6, value=f'=IF(A{k}="","",IFERROR(D{k}/$B$2, 0))'
        )
        share_cell.number_format = styles.percentage_format


def _build_product_table(ws: Worksheet, styles: DashboardStyles) -> None:
    """Build Section 4: Dynamic Product Inventory & Performance Table (A28:E78)."""
    headers = [
        ("A28", "ProductID"),
        ("B28", "ProductName"),
        ("C28", "Stock On Hand"),
        ("D28", "Total Sales Revenue"),
        ("E28", "Stock Status Alert"),
    ]
    for cell_ref, text in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = styles.header_font

    for r in range(2, 52):
        i = 27 + r
        ws.cell(row=i, column=1, value=f'=IF(Products!A{r}="","",Products!A{r})')
        ws.cell(row=i, column=2, value=f'=IF(A{i}="","",Products!B{r})')
        ws.cell(
            row=i,
            column=3,
            value=f'=IF(A{i}="","",SUMIF(TransactionLog!D:D, A{i}, TransactionLog!G:G))',
        )
        rev_cell = ws.cell(
            row=i,
            column=4,
            value=f'=IF(A{i}="","",SUMIFS(TransactionLog!H:H, TransactionLog!D:D, A{i}, TransactionLog!C:C, "SALE")/100)',
        )
        rev_cell.number_format = styles.currency_format
        ws.cell(
            row=i,
            column=5,
            value=f'=IF(A{i}="","",IF(C{i}<=0, "OUT OF STOCK", IF(C{i}<=5, "LOW STOCK", "OK")))',
        )


def _apply_column_widths(ws: Worksheet) -> None:
    """Set optimal column widths for clean visual rendering."""
    column_widths = {
        "A": 22,
        "B": 24,
        "C": 14,
        "D": 22,
        "E": 18,
        "F": 14,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


def _create_dashboard_sheet(workbook: openpyxl.Workbook) -> None:
    """Create and format the executive Dashboard sheet as the first tab in the workbook."""
    ws = workbook.create_sheet(title="Dashboard", index=0)
    styles = DashboardStyles()

    _build_kpi_cards(ws, styles)
    _build_payment_breakdown(ws, styles)
    _build_sales_leaderboard(ws, styles)
    _build_product_table(ws, styles)
    _apply_column_widths(ws)

    workbook.active = ws


def create_master_workbook(
    destination: Path,
    *,
    default_salesman_id: str | None = None,
    sheet_columns: Mapping[str, Sequence[str]] = SHEET_COLUMNS,
    default_salesman_template: Mapping[str, object] | None = None,
    overwrite: bool = False,
) -> Path:
    """Create the CAAD ERP master workbook at ``destination``.

    Parameters are overridable to facilitate testing. When ``overwrite`` is
    ``False`` (the default) this function raises ``FileExistsError`` if the
    target already exists.
    """

    destination = destination.expanduser().resolve()
    if destination.exists() and not overwrite:
        raise FileExistsError(
            f"Refusing to overwrite existing master workbook: {destination}"
        )

    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()

    # Remove the default sheet openpyxl generates so we can create ours.
    if workbook.active and workbook.active.title == "Sheet":
        workbook.remove(workbook.active)

    bold_font = Font(bold=True)

    for sheet_name, columns in sheet_columns.items():
        worksheet = workbook.create_sheet(title=sheet_name)
        for column_index, column_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=column_index)
            cell.value = column_name
            cell.font = bold_font

    _create_dashboard_sheet(workbook)

    workbook.save(destination)
    return destination


def run_from_config(config_path: Path, *, overwrite: bool = False) -> Path:
    """Convenience helper mirroring the original CLI behavior."""

    settings = load_settings(config_path)
    return create_master_workbook(
        settings.data_file,
        default_salesman_id=settings.default_salesman_id,
        overwrite=overwrite,
    )


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    """Parse CLI arguments for the setup script."""

    parser = argparse.ArgumentParser(description="Initialize CAAD ERP data file")
    parser.add_argument(
        "--config",
        default=CONFIG_FILE,
        help="Path to configuration file (default: config.ini)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite the target workbook if it already exists.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    """Entry point for the CLI script."""

    args = parse_args(argv)
    config_path = Path(args.config).expanduser().resolve()

    print("--- CAAD ERP Setup Script ---")
    print(f"Using configuration: {config_path}")

    try:
        output_path = run_from_config(config_path, overwrite=args.force)
    except FileNotFoundError as exc:
        print(f"\n[ERROR] {exc}")
        return 1
    except KeyError as exc:
        print(f"\n[ERROR] {exc}")
        return 1
    except FileExistsError as exc:
        print(f"\n[ERROR] {exc}")
        print("Run with --force to overwrite the existing file if appropriate.")
        return 1
    except (PermissionError, OSError) as exc:
        print(f"\n[ERROR] Unable to write workbook: {exc}")
        return 1

    print(f"\n[SUCCESS] Created master workbook at '{output_path}'.")
    return 0


if __name__ == "__main__":  # pragma: no cover - exercised via manual runs
    sys.exit(main())
