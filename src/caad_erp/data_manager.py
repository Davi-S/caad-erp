"""Data access layer for Lounge ERP.

This module provides low-level helpers that read from and write to the
``master_workbook.xlsx`` workbook. Business logic belongs elsewhere.

The public API is designed around three responsibilities:

1. Configuration handling: finding and parsing ``config.ini``.
2. Workbook lifecycle: opening, validating, and persisting the Excel file.
3. Sheet operations: loading structured records and appending or updating
   individual rows.
"""


from __future__ import annotations

import configparser
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Iterable, Optional, Sequence, Any

from openpyxl.workbook import Workbook
import openpyxl

from . import log
from .constants import SheetName


CONFIG_FILE_NAME = "config.ini"
PRODUCTS_SHEET = SheetName.PRODUCTS.value
SALESMEN_SHEET = SheetName.SALESMEN.value
TRANSACTION_LOG_SHEET = SheetName.TRANSACTION_LOG.value


@dataclass(frozen=True)
class ConfigSettings:
    """Typed representation of the ``config.ini`` settings we care about."""

    data_file: Path
    lounge_name: str
    schema_version: str
    default_salesman_id: str


@dataclass(frozen=True)
class ProductRow:
    """In-memory view of a row from the ``Products`` sheet."""

    product_id: str
    product_name: str
    sell_price: Decimal
    is_active: bool


@dataclass(frozen=True)
class SalesmanRow:
    """In-memory view of a row from the ``Salesmen`` sheet."""

    salesman_id: str
    salesman_name: str
    is_active: bool


@dataclass(frozen=True)
class TransactionRow:
    """In-memory view of a row from the ``TransactionLog`` sheet."""

    transaction_id: str
    timestamp_iso: str
    transaction_type: str
    product_id: Optional[str]
    salesman_id: Optional[str]
    payment_type: Optional[str]
    quantity_change: Decimal
    total_revenue: Decimal
    total_cost: Decimal
    linked_transaction_id: Optional[str]
    notes: Optional[str]


def find_config_file(explicit_path: Optional[Path] = None) -> Path:
    """Locate the configuration file that controls how the data layer behaves.

    If the caller provides ``explicit_path`` the value is returned immediately
    without any verification, which allows the caller to deliberately target a
    non-standard location. When no explicit path is given the function walks up
    from the current working directory toward the filesystem root looking for a
    file named ``CONFIG_FILE_NAME``. The first match that exists on disk is
    considered authoritative.

    Args:
        explicit_path (Path | None): Optional path to use instead of performing
            the upward search. May be relative to the current working directory.

    Returns:
        Path: The path provided by the caller or the discovered configuration
            file. The path is *not* resolved or validated when supplied
            explicitly.

    Raises:
        FileNotFoundError: If the search exhausts all parent directories without
            finding ``CONFIG_FILE_NAME``.
    """

    if explicit_path:
        return explicit_path

    # Walk upwards from the current working directory looking for CONFIG_FILE_NAME
    current = Path.cwd()
    root = current.anchor or Path(current.root) if hasattr(
        current, "root") else Path(current.root)
    for p in (current, *current.parents):
        candidate = p / CONFIG_FILE_NAME
        if candidate.exists():
            return candidate

    raise FileNotFoundError(
        f"Configuration file not found: {CONFIG_FILE_NAME}")


def read_config(config_path: Path) -> configparser.ConfigParser:
    """Load ``config.ini`` and return a populated ``ConfigParser`` instance.

    The function expands user home references (``~``), resolves the absolute
    path, and validates that the file exists before parsing it. Callers receive
    the ``ConfigParser`` even if individual sections are missing; validation of
    required entries happens in :func:`parse_settings`.

    Args:
        config_path (Path): Path to the configuration file, relative or
            absolute.

    Returns:
        configparser.ConfigParser: Initialized parser containing the raw
            configuration data.

    Raises:
        FileNotFoundError: If ``config_path`` does not exist after expansion and
            resolution.
    """

    config_path = config_path.expanduser().resolve()
    if not config_path.exists():
        raise FileNotFoundError(f"Configuration file not found: {config_path}")

    parser = configparser.ConfigParser()
    parser.read(config_path)
    return parser


def parse_settings(parser: configparser.ConfigParser, *, base_path: Optional[Path] = None) -> ConfigSettings:
    """Convert a ``ConfigParser`` into strongly typed :class:`ConfigSettings`.

    The function validates that all required options are present under the
    expected sections and normalizes the configured data file path. Relative
    paths are expanded against ``base_path`` when provided, or against the
    current working directory as a fallback. Resulting paths are resolved to an
    absolute form to ensure downstream consumers operate on canonical values.

    Args:
        parser (configparser.ConfigParser): Parsed configuration data.
        base_path (Path | None): Directory to use as the anchor for relative
            ``DataFile`` entries. Defaults to :func:`Path.cwd` when omitted.

    Returns:
        ConfigSettings: Immutable settings container with resolved data file
            path, lounge metadata, schema version, and default salesman id.

    Raises:
        KeyError: If one of the required sections or options is missing from the
            configuration.
    """

    try:
        data_file_raw = parser.get("System", "DataFile")
        lounge_name = parser.get("System", "LoungeName")
        schema_version = parser.get("System", "SchemaVersion")
        default_salesman = parser.get("Defaults", "DefaultSalesman")
    except (configparser.NoSectionError, configparser.NoOptionError) as exc:
        raise KeyError(f"Missing required configuration entry: {exc}") from exc

    data_file_path = Path(data_file_raw)
    if not data_file_path.is_absolute():
        if base_path is None:
            base_path = Path.cwd()
        data_file_path = (base_path / data_file_path).resolve()

    return ConfigSettings(
        data_file=data_file_path,
        lounge_name=lounge_name,
        schema_version=schema_version,
        default_salesman_id=default_salesman,
    )


def open_workbook(data_file: Path) -> Workbook:
    """Open the master Excel workbook and return a live ``openpyxl`` workbook.

    The provided path is expanded (supporting ``~``), resolved to its absolute
    form, and verified for existence. Successful calls load the workbook via
    :func:`openpyxl.load_workbook`. Callers must retain the resolved path and
    provide it back to :func:`save_workbook` when persisting changes.

    Args:
        data_file (Path): Filesystem path to the ``master_workbook.xlsx`` file.

    Returns:
        Workbook: ``openpyxl`` workbook instance backed by the provided file.

    Raises:
        FileNotFoundError: If ``data_file`` does not exist after expansion and
            resolution.
    """

    data_file = Path(data_file).expanduser().resolve()
    if not data_file.exists():
        raise FileNotFoundError(f"Workbook not found: {data_file}")

    wb = openpyxl.load_workbook(data_file)
    return wb


def save_workbook(workbook: Workbook, destination: Path) -> None:
    """Persist the workbook to disk at an explicitly provided destination.

    The destination path is expanded (supporting ``~``) and resolved to an
    absolute location. Parent directories are created on demand to match the
    previous behavior of the data layer always producing the target folder.

    Args:
        workbook (Workbook): Workbook instance to persist.
        destination (Path): Filesystem path that should receive the serialized
            workbook.
    """

    dest = Path(destination).expanduser().resolve()
    dest.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(dest)


def refresh_workbook(data_file: Path) -> Workbook:
    """Reload the workbook from disk, discarding any unsaved in-memory changes.

    This helper simply proxies to :func:`open_workbook` to obtain a pristine
    workbook instance. It is useful when callers want to abandon modifications
    made to an existing workbook object.

    Args:
        data_file (Path): Location of the workbook to reopen.

    Returns:
        Workbook: Freshly loaded workbook detached from the previous instance.
    """

    # Simply open a fresh workbook instance
    return open_workbook(data_file)


def iter_products(workbook: Workbook) -> Iterable[ProductRow]:
    """Iterate over product records stored on the ``Products`` worksheet.

    The iterator skips the header row and any fully empty rows to avoid
    producing meaningless values. Each non-empty row is converted into a
    :class:`ProductRow` dataclass via :func:`deserialize_product` to provide a
    structured, type-aware record.

    Args:
        workbook (Workbook): Workbook containing the ``Products`` sheet.

    Yields:
        ProductRow: One structured row for each meaningful record in the sheet.
    """

    sheet = workbook[PRODUCTS_SHEET]
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        # skip fully empty rows
        if any(cell is not None for cell in raw):
            yield deserialize_product(raw)


def iter_salesmen(workbook: Workbook) -> Iterable[SalesmanRow]:
    """Iterate over the ``Salesmen`` worksheet and yield typed records.

    Header and completely empty rows are ignored. Remaining rows are converted
    into :class:`SalesmanRow` instances which normalize the raw worksheet values
    into predictable Python types.

    Args:
        workbook (Workbook): Workbook containing the ``Salesmen`` sheet.

    Yields:
        SalesmanRow: Structured representation of each active row in the sheet.
    """

    sheet = workbook[SALESMEN_SHEET]
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in raw):
            yield deserialize_salesman(raw)


def iter_transactions(workbook: Workbook) -> Iterable[TransactionRow]:
    """Stream transaction records from the ``TransactionLog`` worksheet.

    The generator skips headers and rows whose cells are all ``None``. Each
    meaningful row is transformed into a :class:`TransactionRow` via
    :func:`deserialize_transaction`, ensuring numeric fields become
    :class:`~decimal.Decimal` instances and optional text fields remain ``None``
    when blank.

    Args:
        workbook (Workbook): Workbook containing the transaction log sheet.

    Yields:
        TransactionRow: Normalized transaction record for each populated row.
    """

    sheet = workbook[TRANSACTION_LOG_SHEET]
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in raw):
            yield deserialize_transaction(raw)


def append_product(workbook: Workbook, record: ProductRow) -> None:
    """Append a product record to the ``Products`` worksheet.

    The dataclass is serialized into the exact column ordering expected by the
    sheet before being appended. Row formulas or formatting are preserved by
    ``openpyxl`` as part of the append operation.

    Args:
        workbook (Workbook): Workbook whose products sheet should be modified.
        record (ProductRow): Structured product data ready for persistence.
    """

    sheet = workbook[PRODUCTS_SHEET]
    sheet.append(serialize_product(record))


def append_salesman(workbook: Workbook, record: SalesmanRow) -> None:
    """Append a salesman record to the ``Salesmen`` worksheet.

    The helper serializes the dataclass into the worksheet's column ordering and
    leverages :meth:`openpyxl.worksheet.worksheet.Worksheet.append` to add the
    new row.

    Args:
        workbook (Workbook): Workbook whose salesmen sheet should be modified.
        record (SalesmanRow): Salesman entry to append.
    """

    sheet = workbook[SALESMEN_SHEET]
    sheet.append(serialize_salesman(record))


def append_transaction(workbook: Workbook, record: TransactionRow) -> None:
    """Append a transaction record to the ``TransactionLog`` worksheet.

    Numerical fields remain :class:`~decimal.Decimal` instances after
    serialization, allowing Excel to preserve precision when the workbook is
    saved.

    Args:
        workbook (Workbook): Workbook containing the transaction log.
        record (TransactionRow): Transaction to persist.
    """

    sheet = workbook[TRANSACTION_LOG_SHEET]
    sheet.append(serialize_transaction(record))


def update_product(workbook: Workbook, product_id: str, *, field_values: dict[str, Any]) -> None:
    """Update selected columns for an existing product.

    The function locates the row whose ``ProductID`` matches ``product_id``,
    validates that each requested field exists in the header row, and then writes
    the provided values into the corresponding cells. Only the specified fields
    are modified, leaving other columns untouched.

    Args:
        workbook (Workbook): Workbook containing the products sheet.
        product_id (str): Identifier used to locate the target row.
        field_values (dict[str, Any]): Mapping of column names to replacement
            values.

    Raises:
        KeyError: If the product or any referenced column cannot be found.
    """

    sheet_name = PRODUCTS_SHEET
    row_index = locate_row(workbook, sheet_name, "ProductID", product_id)
    if row_index is None:
        raise KeyError(f"Product not found: {product_id}")

    sheet = workbook[sheet_name]
    headers = list(sheet[1])
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(headers)}

    for field, value in field_values.items():
        if field not in header_map:
            raise KeyError(f"Unknown product field: {field}")
        col = header_map[field]
        sheet.cell(row=row_index, column=col, value=value)


def update_salesman(workbook: Workbook, salesman_id: str, *, field_values: dict[str, Any]) -> None:
    """Update selected columns for an existing salesman.

    The function resolves the row by ``SalesmanID``, checks that each requested
    column exists, and updates only the specified fields with the supplied
    values.

    Args:
        workbook (Workbook): Workbook containing the salesmen sheet.
        salesman_id (str): Identifier used to locate the target row.
        field_values (dict[str, Any]): Mapping of column names to replacement
            values.

    Raises:
        KeyError: If the salesman or any referenced column is missing.
    """

    sheet_name = SALESMEN_SHEET
    row_index = locate_row(workbook, sheet_name, "SalesmanID", salesman_id)
    if row_index is None:
        raise KeyError(f"Salesman not found: {salesman_id}")

    sheet = workbook[sheet_name]
    headers = list(sheet[1])
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(headers)}

    for field, value in field_values.items():
        if field not in header_map:
            raise KeyError(f"Unknown salesman field: {field}")
        col = header_map[field]
        sheet.cell(row=row_index, column=col, value=value)


def locate_row(workbook: Workbook, sheet_name: str, key_column: str, key_value: str) -> Optional[int]:
    """Find a row by matching a key value within the specified worksheet.

    The function constructs a mapping from header titles to column indices,
    verifies that ``key_column`` exists, and scans the worksheet for the first
    row whose value equals ``key_value``. The header row itself is not
    considered during matching.

    Args:
        workbook (Workbook): Workbook providing access to ``sheet_name``.
        sheet_name (str): Name of the worksheet to search.
        key_column (str): Header title identifying the column that stores the
            lookup key.
        key_value (str): Value to match within the key column.

    Returns:
        int | None: 1-based Excel row index when a match is found, otherwise
            ``None``.

    Raises:
        KeyError: If ``key_column`` is not present in the worksheet header.
    """

    sheet = workbook[sheet_name]
    # Build header -> column index map
    header_cells = list(sheet[1])
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(header_cells)}
    if key_column not in header_map:
        raise KeyError(f"Unknown column: {key_column}")

    key_col_index = header_map[key_column]

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        cell_value = row[key_col_index - 1]
        if cell_value == key_value:
            return row_idx

    return None


def serialize_product(record: ProductRow) -> list[object]:
    """Convert a product dataclass into the worksheet column ordering.

    Args:
        record (ProductRow): Structured product data to transform.

    Returns:
        list[object]: Values arranged as ``[ProductID, ProductName,
        SellPrice, IsActive]`` suitable for worksheet insertion.
    """

    return [record.product_id, record.product_name, record.sell_price, record.is_active]


def serialize_salesman(record: SalesmanRow) -> list[object]:
    """Convert a salesman dataclass into the worksheet column ordering.

    Args:
        record (SalesmanRow): Structured salesman data to transform.

    Returns:
        list[object]: Values arranged as ``[SalesmanID, SalesmanName,
        IsActive]`` suitable for worksheet insertion.
    """

    return [record.salesman_id, record.salesman_name, record.is_active]


def serialize_transaction(record: TransactionRow) -> list[object]:
    """Convert a transaction dataclass into the transaction log column order.

    Args:
        record (TransactionRow): Structured transaction data to transform.

    Returns:
        list[object]: Values ordered to match the spreadsheet columns,
            preserving :class:`~decimal.Decimal` instances for numeric fields.
    """

    return [
        record.transaction_id,
        record.timestamp_iso,
        record.transaction_type,
        record.product_id,
        record.salesman_id,
        record.payment_type,
        record.quantity_change,
        record.total_revenue,
        record.total_cost,
        record.linked_transaction_id,
        record.notes,
    ]


def deserialize_product(raw_row: Sequence[object]) -> ProductRow:
    """Convert a raw worksheet row into a strongly typed product record.

    The converter normalizes numeric values into :class:`~decimal.Decimal`
    instances and coerces id/name fields to ``str`` to avoid surprises caused by
    Excel automatically interpreting numbers.

    Args:
        raw_row (Sequence[object]): Raw cell values from the worksheet row.

    Returns:
        ProductRow: Dataclass containing consistent Python representations of
            the row contents.
    """

    product_id = raw_row[0]
    product_name = raw_row[1]
    sell_raw = raw_row[2]
    is_active = raw_row[3]

    sell_price = Decimal(
        str(sell_raw)) if sell_raw is not None else Decimal("0.00")
    return ProductRow(product_id=str(product_id), product_name=str(product_name), sell_price=sell_price, is_active=bool(is_active))


def deserialize_salesman(raw_row: Sequence[object]) -> SalesmanRow:
    """Convert a raw worksheet row into a strongly typed salesman record.

    The function coerces identifier and name fields to ``str`` and uses ``bool``
    coercion for the active flag to hide underlying spreadsheet encodings.

    Args:
        raw_row (Sequence[object]): Raw cell values from the worksheet row.

    Returns:
        SalesmanRow: Dataclass populated with normalized values.
    """

    salesman_id = raw_row[0]
    salesman_name = raw_row[1]
    is_active = raw_row[2]
    return SalesmanRow(salesman_id=str(salesman_id), salesman_name=str(salesman_name), is_active=bool(is_active))


def deserialize_transaction(raw_row: Sequence[object]) -> TransactionRow:
    """Convert a raw worksheet row into a strongly typed transaction record.

    Decimal-compatible columns are normalized into :class:`~decimal.Decimal`
    instances to preserve precision, optional columns remain ``None`` when the
    sheet leaves them blank, and textual columns default to empty strings to
    avoid ``None`` values where downstream code expects text.

    Args:
        raw_row (Sequence[object]): Raw cell values from the transaction log row
            in their worksheet order.

    Returns:
        TransactionRow: Dataclass reflecting the row contents with consistent
            Python types.
    """

    (
        transaction_id,
        timestamp_iso,
        transaction_type,
        product_id,
        salesman_id,
        payment_type,
        quantity_change_raw,
        total_revenue_raw,
        total_cost_raw,
        linked_transaction_id,
        notes,
    ) = raw_row

    quantity_change = Decimal(
        str(quantity_change_raw)) if quantity_change_raw is not None else Decimal("0")
    total_revenue = Decimal(
        str(total_revenue_raw)) if total_revenue_raw is not None else Decimal("0.00")
    total_cost = Decimal(str(total_cost_raw)
                         ) if total_cost_raw is not None else Decimal("0.00")

    return TransactionRow(
        transaction_id=str(transaction_id),
        timestamp_iso=str(timestamp_iso) if timestamp_iso is not None else "",
        transaction_type=str(
            transaction_type) if transaction_type is not None else "",
        product_id=(str(product_id) if product_id is not None else None),
        salesman_id=(str(salesman_id) if salesman_id is not None else None),
        payment_type=(str(payment_type) if payment_type is not None else None),
        quantity_change=quantity_change,
        total_revenue=total_revenue,
        total_cost=total_cost,
        linked_transaction_id=(str(linked_transaction_id)
                               if linked_transaction_id is not None else None),
        notes=(str(notes) if notes is not None else None),
    )
