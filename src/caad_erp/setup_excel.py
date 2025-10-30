import os
import openpyxl
from openpyxl.styles import Font

DATA_FILE = "lounge_master_data.xlsx"

SHEET_COLUMNS = {
    "Products": ["ProductID", "ProductName", "SellPrice", "IsActive"],
    "Salesmen": ["SalesmanID", "SalesmanName", "IsActive"],
    "TransactionLog": [
        "TransactionID",
        "Timestamp",
        "TransactionType",
        "ProductID",
        "SalesmanID",
        "PaymentType",
        "QuantityChange",
        "TotalRevenue",
        "TotalCost",
        "LinkedTransactionID",
        "Notes",
    ],
}

DEFAULT_SALESMAN = ("tr", "Lounge Sale", True)
# DEFAULT_SALESMAN = None


def create_excel_file():
    """
    Creates a new, empty lounge_master_data.xlsx file
    with the correct sheets and column headers.
    """
    if os.path.exists(DATA_FILE):
        print(
            f"Error: '{DATA_FILE}' already exists. Please remove it to re-initialize."
        )
        return

    # Create a new workbook and remove the default "Sheet"
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    bold_font = Font(bold=True)

    # Create each sheet and add headers
    for sheet_name, columns in SHEET_COLUMNS.items():
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.font = bold_font
        print(f"Created sheet: '{sheet_name}'")

    # Add the default salesman, as per our "foo salesman" decision
    if DEFAULT_SALESMAN:
        try:
            ws_salesmen = wb["Salesmen"]
            ws_salesmen.append(DEFAULT_SALESMAN)
            print(f"Added default salesman: '{DEFAULT_SALESMAN[1]}'")
        except KeyError:
            print("Error: Could not add default salesman. 'Salesmen' sheet not found.")

    # Save the file
    try:
        wb.save(DATA_FILE)
        print(f"\nSuccessfully created '{DATA_FILE}'.")
        print("You can now run 'cli.py' to interact with the backend.")
    except Exception as e:
        print(f"An error occurred while saving the file: {e}")


if __name__ == "__main__":
    print("Initializing new ERP backend data file...")
    create_excel_file()
